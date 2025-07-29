import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from datetime import datetime
import customtkinter as ctk
from tkinter import messagebox, StringVar
import win32com.client as win32
import comtypes.client
from pathlib import Path
import winreg
from shutil import copyfile
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from collections import defaultdict
import numpy as np
import os
import threading
import glob
import subprocess
from collections import defaultdict, Counter
import yagmail
import time
import gc
import win32com.client as win32
import traceback
import shutil
import openpyxl

log_geral = []
def encontrar_pasta_onedrive_empresa():
    user_dir = os.environ["USERPROFILE"]
    possiveis = os.listdir(user_dir)
    for nome in possiveis:
        if "DIAS BRANCO" in nome.upper():
            caminho_completo = os.path.join(user_dir, nome)
            if os.path.isdir(caminho_completo) and "Gestão de Estoque - Documentos" in os.listdir(caminho_completo):
                return os.path.join(caminho_completo, "Gestão de Estoque - Documentos")
    return None

fonte_dir = encontrar_pasta_onedrive_empresa()
if not fonte_dir:
    raise FileNotFoundError("Não foi possível localizar a pasta sincronizada do SharePoint via OneDrive.")

caminho_base_fisica = os.path.join(fonte_dir, "SIMULADOR_BALANÇA_LIMPO_2.xlsx")
df_base_fisica = pd.read_excel(caminho_base_fisica, sheet_name="BASE FISICA")
df_base_familia = pd.read_excel(caminho_base_fisica, 'BASE_FAMILIA')

def converter_para_float_seguro(valor):
    if pd.isna(valor):
        return 0.0
    try:
        return float(valor)
    except (ValueError, TypeError):
        return 0.0

def salvar_em_base_auxiliar(df_remessa, remessa, log_callback, fonte_dir):
    caminho_aux = os.path.join(fonte_dir, "expedicao_edicoes.xlsx")
    try:
        if os.path.exists(caminho_aux):
            try:
                with pd.ExcelFile(caminho_aux) as xls:
                    if "dado_exp" in xls.sheet_names:
                        df_existente = pd.read_excel(xls, sheet_name="dado_exp")
                        df_existente = df_existente[
                            df_existente['REMESSA'].astype(str).str.replace(r'\.0$', '', regex=True) != str(remessa)
                        ]
                    else:
                        df_existente = pd.DataFrame()
            except Exception as e:
                log_callback(f"Erro ao ler arquivo existente, criando novo: {str(e)}")
                df_existente = pd.DataFrame()
        else:
            df_existente = pd.DataFrame()
        df_remessa['REMESSA'] = df_remessa['REMESSA'].astype(str).str.replace(r'\.0$', '', regex=True)
        df_remessa = df_remessa.dropna(subset=['ITEM'])
        df_remessa = df_remessa[df_remessa['ITEM'] != '']
        df_remessa = df_remessa.drop_duplicates(
            subset=['REMESSA', 'ITEM', 'CHAVE_PALETE'], 
            keep='last'
        )
        df_atualizado = pd.concat([df_existente, df_remessa], ignore_index=True)
        df_atualizado = df_atualizado.sort_values(by=['REMESSA', 'ITEM'])
        with pd.ExcelWriter(caminho_aux, engine='openpyxl') as writer:
            df_atualizado.to_excel(writer, sheet_name="dado_exp", index=False)
        
        log_callback(f"Remessa {remessa} salva na base auxiliar. Total de itens: {len(df_remessa)}")
        return df_atualizado

    except Exception as e:
        log_callback(f"[ERRO AO SALVAR NA BASE AUXILIAR]: {str(e)}")
        raise

def carregar_base_auxiliar(fonte_dir):
    caminho_aux = os.path.join(fonte_dir, "expedicao_edicoes.xlsx")
    try:
        if os.path.exists(caminho_aux):
            with pd.ExcelFile(caminho_aux) as xls:
                if "dado_exp" in xls.sheet_names:
                    return pd.read_excel(xls, sheet_name="dado_exp")
        return pd.DataFrame()
    except Exception as e:
        print(f"Erro ao carregar base auxiliar: {e}")
        return pd.DataFrame()

def remover_remessa_base_auxiliar(remessa, fonte_dir, log_callback):
    try:
        caminho_aux = os.path.join(fonte_dir, "expedicao_edicoes.xlsx")
        if not os.path.exists(caminho_aux):
            return
        
        with pd.ExcelFile(caminho_aux) as xls:
            if "dado_exp" in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name="dado_exp")
                df = df[df['REMESSA'].astype(str) != str(remessa)]
                
                with pd.ExcelWriter(caminho_aux, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name="dado_exp", index=False)
                
                log_callback(f"Remessa {remessa} removida da base auxiliar")
    except Exception as e:
        log_callback(f"Erro ao remover remessa da base auxiliar: {e}")

def tratar_erro_gen_py(e, log_callback=None):
    erro_str = str(e)
    if "gen_py" in erro_str and "CLSIDToClassMap" in erro_str:
        pasta_genpy = os.path.join(os.getenv("LOCALAPPDATA"), "Temp", "gen_py")
        try:
            if os.path.exists(pasta_genpy):
                shutil.rmtree(pasta_genpy)
                msg = f"[INFO] Pasta gen_py removida com sucesso de {pasta_genpy}. Por favor, reinicie o programa."
            else:
                msg = f"[INFO] gen_py não encontrada em {pasta_genpy}, mas o erro persiste. Tente rodar: python -m pywin32_postinstall -install"
        except Exception as erro_remocao:
            msg = f"[ERRO] Falha ao remover gen_py: {erro_remocao}"
    
        if log_callback:
            log_callback(msg)
        else:
            print(msg)
        return True 
    return False

def obter_dados_remessa(remessa, df_expedicao, log_callback):
    try:
        remessa_str = str(int(float(remessa))) if '.' in str(remessa) else str(remessa)
        
        caminho_aux = os.path.join(fonte_dir, "expedicao_edicoes.xlsx")
        if os.path.exists(caminho_aux):
            df_aux = pd.read_excel(caminho_aux, sheet_name="dado_exp")
            df_aux['REMESSA_COMPARACAO'] = df_aux['REMESSA'].astype(str).str.replace(r'\.0$', '', regex=True)
            df_filtrado_aux = df_aux[df_aux['REMESSA_COMPARACAO'] == remessa_str].copy()
            if not df_filtrado_aux.empty:
                log_callback(f"Remessa {remessa} encontrada na base auxiliar (edicoes)")
                return df_filtrado_aux.drop(columns=['REMESSA_COMPARACAO'])

        df_expedicao['REMESSA_COMPARACAO'] = df_expedicao['REMESSA'].astype(str).str.replace(r'\.0$', '', regex=True)
        df_filtrado = df_expedicao[df_expedicao['REMESSA_COMPARACAO'] == remessa_str].copy()
        
        if not df_filtrado.empty:
            log_callback(f"Remessa {remessa} encontrada na base original")
            return df_filtrado.drop(columns=['REMESSA_COMPARACAO'])
        
        log_callback(f"Remessa {remessa} não encontrada em nenhuma base")
        return pd.DataFrame()
        
    except Exception as e:
        log_callback(f"Erro ao buscar remessa {remessa}: {str(e)}")
        return pd.DataFrame()

def criar_copia_planilha(fonte_dir, nome_arquivo, log_callback):
    try:
        origem = os.path.join(fonte_dir, nome_arquivo)
        destino_dir = os.path.join(os.environ['USERPROFILE'], 'Downloads')
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        nome_copia = f"copia_temp_{timestamp}_{nome_arquivo}"
        destino = os.path.join(destino_dir, nome_copia)
        copyfile(origem, destino)
        log_callback(f"Cópia criada com sucesso: {destino}")
        return destino
    except Exception as e:
        log_callback(f"Erro ao criar cópia da planilha: {e}")
        raise

def enviar_email_com_log_e_pdf(caminho_pdf, remessa, log_callback=None, log_geral=None):
    import yagmail
    email_remetente = 'mdiasbrancoautomacao@gmail.com'
    token = 'secwygmzlibyxhhh'  
    email_destino = 'douglas.lins2@mdiasbranco.com.br'

    try:
        corpo_email = "\n".join(log_geral or ["(Sem logs técnicos disponíveis)"])
        yag = yagmail.SMTP(user=email_remetente, password=token)
        assunto = f"📦 Simulador Sobrepeso - Remessa {remessa}"
        yag.send(
            to=email_destino,
            subject=assunto,
            contents=corpo_email,
            attachments=[caminho_pdf]
        )
        if log_callback:
            log_callback(f"E-mail enviado com sucesso para {email_destino}")
    except Exception as e:
        if log_callback:
            log_callback(f"Erro ao enviar e-mail: {e}")

def print_pdf(file_path, impressora="VITLOG01A01", sumatra_path="C:\\Program Files\\SumatraPDF\\SumatraPDF.exe", log_callback=None):
    import subprocess
    args = [sumatra_path, "-print-to", impressora, "-silent", file_path]
    try:
        result = subprocess.run(args, check=True, capture_output=True, text=True)
        if log_callback:
            log_callback(f"Arquivo impresso com sucesso: {file_path}")
            if result.stdout:
                log_callback(f"stdout: {result.stdout}")
            if result.stderr:
                log_callback(f"stderr: {result.stderr}")
    except subprocess.CalledProcessError as e:
        if log_callback:
            log_callback(f"Erro ao imprimir {file_path}: {e}")
            log_callback(f"Output: {e.output}")
            log_callback(f"Stderr: {e.stderr}")

def calculo_sobrepeso_fixo(sku, df_base_fisica, peso_base_liq, log_callback):
    try:
        sp_row = df_base_fisica[df_base_fisica['CÓDIGO PRODUTO'] == sku]
        if not sp_row.empty:
            sp_fixo = pd.to_numeric(sp_row.iloc[0]['SOBRE PESO'], errors='coerce') / 100
            peso_base_liq = pd.to_numeric(peso_base_liq, errors='coerce')

            if isinstance(peso_base_liq, pd.Series):
                peso_base_liq = peso_base_liq.iloc[0]

            if pd.notna(peso_base_liq) and pd.notna(sp_fixo):
                ajuste = float(peso_base_liq) * float(sp_fixo)
            else:
                ajuste = 0.0
            return sp_fixo, ajuste
        else:
            log_callback(f"Nenhum sobrepeso fixo encontrado para SKU {sku}.")
            return 0.0, 0.0
    except Exception as e:
        log_callback(f"Erro ao buscar sobrepeso fixo para SKU {sku}: {e}")
        return 0.0, 0.0

def buscar_chave_por_endereco(endereco, df_estoque_sep):
    if pd.isna(endereco) or endereco == '':
        return None
    df_filtrado = df_estoque_sep[df_estoque_sep['endereco'] == endereco]
    if df_filtrado.empty:
        return None
    df_ordenado = df_filtrado.sort_values(by='Criado', ascending=False)
    return df_ordenado.iloc[0]['chave_pallete']

def processar_sobrepeso(chave_pallet, sku, peso_base_liq, df_sap, df_sobrepeso_real, df_base_fisica, log_callback):
    peso_base_liq = converter_para_float_seguro(peso_base_liq)
    sp = 0
    origem_sp = 'não encontrado'
    ajuste_sp = 0

    if pd.notna(chave_pallet) and chave_pallet in df_sap['Chave Pallet'].values:
        pallet_info = df_sap[df_sap['Chave Pallet'] == chave_pallet].iloc[0]
        lote = pallet_info['Lote']
        data_producao = pallet_info['Data de produção']
        hora_inicio = f"{pallet_info['Hora de criação'].hour:02d}:00:00"
        hora_fim = f"{pallet_info['Hora de modificação'].hour:02d}:00:00"
        linha_coluna = "L" + lote[-3:]
        if linha_coluna in ['LB06', 'LB07']:
            linha_coluna = 'LB06/07'

        log_callback(f"Processando pallet {chave_pallet} para SKU {sku}.")

        df_sp_filtro = df_sobrepeso_real[
            (df_sobrepeso_real['DataHora'] >= pd.to_datetime(f"{data_producao} {hora_inicio}")) &
            (df_sobrepeso_real['DataHora'] <= pd.to_datetime(f"{data_producao} {hora_fim}"))
        ]

        if linha_coluna in df_sp_filtro.columns:
            sp_valores = df_sp_filtro[linha_coluna].fillna(0)
            if not sp_valores.empty:
                media_sp = sp_valores.mean() / 100
                log_callback(f"[{linha_coluna}] Média SP: {media_sp:.4f}")

                peso_base_liq = pd.to_numeric(peso_base_liq, errors='coerce')
                sp = pd.to_numeric(media_sp, errors='coerce')

                if isinstance(peso_base_liq, pd.Series):
                    peso_base_liq = peso_base_liq.iloc[0]
                if isinstance(sp, pd.Series):
                    sp = sp.iloc[0]

                if pd.notna(peso_base_liq) and pd.notna(sp) and float(sp)>0:
                    ajuste_sp = float(peso_base_liq) * float(sp)
                    origem_sp = 'real'
                else:
                    sp=0
                    ajuste_sp = 0.0
                    origem_sp='fixo'
                    log_callback(f"Erro: peso_base_liq ou sp inválido (verificar se há a data ou chave pallete na base) para SKU {sku}. Ajuste SP definido como 0.")
        else:
            log_callback(f"Coluna {linha_coluna} não encontrada na base de sobrepeso.")

    if sp == 0:
        sp_valor, ajuste_fixo = calculo_sobrepeso_fixo(sku, df_base_fisica, peso_base_liq, log_callback)
        if sp_valor != 0:
            sp = sp_valor
            origem_sp = 'fixo'
            ajuste_sp = ajuste_fixo
            if origem_sp == 'fixo':
                log_callback(f"SOBREPESO FIXO aplicado para SKU {sku}: {sp:.4f} → Ajuste: {ajuste_sp:.2f} kg")
        else:
            log_callback(f"Nenhum sobrepeso encontrado para SKU {sku}.")

    return sp, origem_sp, ajuste_sp

def processar_sobrepeso_fixo_basico(sku, qtd, df_sku, df_base_fisica, peso_base_total, peso_base_total_liq, sp_total, itens_detalhados, log_callback):
    try:
        df_sku_filtrado = df_sku[df_sku['COD_PRODUTO'] == sku]
        if df_sku_filtrado.empty:
            log_callback(f"SKU {sku} não encontrado na base SKU.")
            return peso_base_total, peso_base_total_liq, sp_total

        peso_por_caixa_bruto = pd.to_numeric(df_sku_filtrado.iloc[0]['QTDE_PESO_BRU'], errors='coerce')
        peso_por_caixa_liquido = pd.to_numeric(df_sku_filtrado.iloc[0]['QTDE_PESO_LIQ'], errors='coerce')
        qtd_validada = pd.to_numeric(qtd, errors='coerce')

        peso_base = peso_base_liq = 0.0

        if pd.notna(peso_por_caixa_bruto) and pd.notna(qtd_validada):
            peso_base = float(peso_por_caixa_bruto) * float(qtd_validada)
            peso_base_total += peso_base

        if pd.notna(peso_por_caixa_liquido) and pd.notna(qtd_validada):
            peso_base_liq = float(peso_por_caixa_liquido) * float(qtd_validada)
            peso_base_total_liq += peso_base_liq

        sp_valor, ajuste_sp = calculo_sobrepeso_fixo(sku, df_base_fisica, peso_base_liq, log_callback)
        sp_total += ajuste_sp

        log_callback(f"Aplicando SP fixo para SKU {sku} → ajuste: {ajuste_sp:.2f}kg")

        itens_detalhados.append({
            'sku': sku,
            'chave_pallet': 'N/A',
            'sp': round(sp_valor, 4),
            'ajuste_sp': round(ajuste_sp, 2),
            'origem': 'fixo'
        })

        return peso_base_total, peso_base_total_liq, sp_total

    except Exception as e:
        log_callback(f"Erro ao aplicar SP fixo para SKU {sku}: {e}")
        return peso_base_total, peso_base_total_liq, sp_total
    
def calcular_peso_teorico_receb_ext(sku, qtd, df_sku):
    try:
        df_sku_filtrado = df_sku[df_sku['COD_PRODUTO'] == sku]
        if df_sku_filtrado.empty:
            return 0.0

        peso_liq = pd.to_numeric(df_sku_filtrado.iloc[0]['QTDE_PESO_LIQ'], errors='coerce')
        qtd = pd.to_numeric(qtd, errors='coerce')

        if pd.notna(peso_liq) and pd.notna(qtd):
            return float(peso_liq) * float(qtd)
        return 0.0
    except:
        return 0.0

def calcular_peso_receb_externo(chave, sku, row, df_externo_peso, df_estoque_sep, df_sku,
                                peso_base_total, peso_base_total_liq, itens_detalhados, log_callback):
    try:
        log_callback(f"🔎 Verificando recebimento externo para: {chave}")
        if not isinstance(chave, str) or not chave.endswith("_1"):
            return None

        row_exter = df_externo_peso[df_externo_peso['chave_pallete'] == chave]
        if row_exter.empty:
            log_callback("Não encontrado em df_externo_peso.")
            return None

        peso_real_total = pd.to_numeric(row_exter.iloc[0]['peso'], errors='coerce')
        qtd_registrada = pd.to_numeric(row_exter.iloc[0]['quantidade'], errors='coerce')
        sku_externo = pd.to_numeric(row_exter.iloc[0]['SKU'], errors='coerce')

        if pd.isna(peso_real_total) or pd.isna(qtd_registrada) or qtd_registrada == 0:
            log_callback("Peso real inválido ou dados incompletos para cálculo proporcional.")
            return None

        qtd_expedida = None
        for campo in ['quantidade', 'QUANTIDADE', 'qtd']:
            if campo in row.index:
                qtd_expedida = pd.to_numeric(row[campo], errors='coerce')
                break

        if pd.isna(qtd_expedida):
            log_callback("Quantidade expedida inválida ou ausente para recebimento externo.")
            return None


        if qtd_expedida > qtd_registrada:
            log_callback(f"Quantidade expedida ({qtd_expedida}) maior do que registrada ({qtd_registrada}) no pallet externo {chave}. Usando mesmo assim.")

        peso_proporcional = peso_real_total * (qtd_expedida / qtd_registrada)
        df_sku_filtrado = df_sku[df_sku['COD_PRODUTO'] == sku]
        if df_sku_filtrado.empty:
            log_callback(f"SKU {sku} não encontrado na base SKU.")
            return None

        peso_unit_liq = pd.to_numeric(df_sku_filtrado.iloc[0]['QTDE_PESO_LIQ'], errors='coerce')
        peso_unit_bruto = pd.to_numeric(df_sku_filtrado.iloc[0]['QTDE_PESO_BRU'], errors='coerce')

        peso_liq_teorico = float(peso_unit_liq) * qtd_expedida if pd.notna(peso_unit_liq) else 0
        peso_bruto_teorico = float(peso_unit_bruto) * qtd_expedida if pd.notna(peso_unit_bruto) else 0

        if peso_liq_teorico > 0:
            sp_real = (peso_proporcional / peso_liq_teorico) - 1
        else:
            sp_real = 0

        ajuste_sp = peso_proporcional - peso_liq_teorico
        peso_base_total += peso_bruto_teorico
        peso_base_total_liq += peso_liq_teorico

        log_callback(f"Receb. Externo: Peso real {peso_proporcional:.2f}kg | Peso teórico {peso_liq_teorico:.2f}kg → SP: {sp_real:.4f}")

        itens_detalhados.append({
            'sku': sku,
            'chave_pallet': chave,
            'sp': round(sp_real, 4),
            'ajuste_sp': round(ajuste_sp, 2),
            'origem': 'receb_ext'
        })

        return peso_base_total, peso_base_total_liq

    except Exception as e:
        log_callback(f"Erro em calcular_peso_receb_externo para chave {chave}: {e}")
        return None

def calcular_peso_final(remessa_num, peso_veiculo_vazio, qtd_paletes, df_remessa, df_sku, df_sap, df_sobrepeso_real,
                         df_base_fisica, df_frac, df_estoque_sep, df_externo_peso, log_callback):
    
    peso_veiculo_vazio = converter_para_float_seguro(peso_veiculo_vazio)
    qtd_paletes = converter_para_float_seguro(qtd_paletes)
    df_remessa['QUANTIDADE'] = df_remessa['QUANTIDADE'].apply(converter_para_float_seguro)
    df_sku['QTDE_PESO_BRU'] = df_sku['QTDE_PESO_BRU'].apply(converter_para_float_seguro)
    df_sku['QTDE_PESO_LIQ'] = df_sku['QTDE_PESO_LIQ'].apply(converter_para_float_seguro)
    try:
        remessa_num = int(remessa_num)
    except ValueError:
        log_callback("Remessa inválida.")
        return None
    colunas_checagem = ['DOCA', 'QUANTIDADE', 'DATA', 'VIDA_UTIL_EM_DIAS', 'ITEM', 'CHAVE_PALETE']
    df_com_chave = df_remessa[df_remessa['CHAVE_PALETE'].notna()]
    df_sem_chave = df_remessa[df_remessa['CHAVE_PALETE'].isna()]
    total_antes = len(df_remessa)
    com_chave_antes = len(df_com_chave)
    df_com_chave = df_com_chave.drop_duplicates(subset=colunas_checagem)
    com_chave_depois = len(df_com_chave)
    total_depois = com_chave_depois + len(df_sem_chave)
    df_remessa = pd.concat([df_com_chave, df_sem_chave], ignore_index=True)
    log_callback(f"Removidas duplicatas com CHAVE_PALETE: {com_chave_antes - com_chave_depois} removidas de {com_chave_antes} registros com chave.")
    log_callback(f"Total de linhas antes: {total_antes} → depois da limpeza: {total_depois}")
    contador_exp = defaultdict(int)
    for _, row in df_remessa.iterrows():
        key = (row['ITEM'], row['QUANTIDADE'])
        contador_exp[key] += 1

    peso_base_total = 0
    peso_base_total_liq = 0
    sp_total = 0
    itens_detalhados = []
    chaves_pallet_processadas = []

    df_expedicao_sem_pallet = df_remessa[df_remessa['CHAVE_PALETE'].isna()]
    pacotes_expedicao = df_expedicao_sem_pallet.groupby(['ITEM', 'QUANTIDADE']).size().reset_index(name='count')

    for _, pacote in pacotes_expedicao.iterrows():
        sku = pacote['ITEM']
        qtd = pd.to_numeric(pacote['QUANTIDADE'], errors='coerce')
        count_expedicao = int(pacote['count'])

        df_frac_match = df_frac[(df_frac['remessa'] == remessa_num) & (df_frac['sku'] == sku)]
        df_frac_match = df_frac_match[pd.to_numeric(df_frac_match['qtd'], errors='coerce') == qtd]
        df_frac_same_remessa = df_frac[df_frac['remessa'] == remessa_num]
        skus_frac = df_frac_same_remessa['sku'].unique()
        if len(skus_frac) > 0 and sku not in skus_frac:
            log_callback(f"Remessa {remessa_num} encontrada na base FRACAO, mas com SKUs diferentes do esperado: {skus_frac.tolist()} — SKU da linha atual: {sku}. Pode haver divergência nos dados.")

        count_fracao = len(df_frac_match)

        qtd_real = min(count_expedicao, count_fracao)
        qtd_fixo = count_expedicao - qtd_real

        log_callback(f"Processando SKU {sku} QTD {qtd}: {qtd_real} via real, {qtd_fixo} via fixo")

        for idx in range(qtd_real):
            frac_row = df_frac_match.iloc[idx]
            chave_frac = frac_row.get('chave_pallete')
            if pd.isna(chave_frac) or chave_frac == '':
                endereco = frac_row.get('endereco')
                chave_frac = buscar_chave_por_endereco(endereco, df_estoque_sep)

            if pd.isna(chave_frac):
                continue
            chaves_pallet_processadas.append(chave_frac)
            resultado_ext = calcular_peso_receb_externo(
            chave_frac, sku, frac_row, df_externo_peso, df_estoque_sep,df_sku,
            peso_base_total, peso_base_total_liq, itens_detalhados, log_callback
            )
            if resultado_ext:
                peso_base_total, peso_base_total_liq = resultado_ext
                continue
            pallet_info = df_sap[df_sap['Chave Pallet'] == chave_frac]
            if pallet_info.empty:
                log_callback(f"Chave pallet {chave_frac} da FRACAO não encontrada na base SAP.")
                continue

            pallet_info = pallet_info.iloc[0]
            df_sku_filtrado = df_sku[df_sku['COD_PRODUTO'] == sku]
            if df_sku_filtrado.empty:
                log_callback(f"SKU {sku} não encontrado na base SKU.")
                continue

            peso_liq = pd.to_numeric(df_sku_filtrado.iloc[0]['QTDE_PESO_LIQ'], errors='coerce')
            qtd_validada = pd.to_numeric(qtd, errors='coerce')

            peso_base_liq = float(peso_liq) * float(qtd_validada) if pd.notna(peso_liq) and pd.notna(qtd_validada) else 0
            peso_base_total_liq += peso_base_liq

            peso_bruto = pd.to_numeric(df_sku_filtrado.iloc[0]['QTDE_PESO_BRU'], errors='coerce')
            peso_base = float(peso_bruto) * float(qtd_validada) if pd.notna(peso_bruto) and pd.notna(qtd_validada) else 0
            peso_base_total += peso_base

            sp, origem_sp, ajuste_sp = processar_sobrepeso(chave_frac, sku, peso_base_liq, df_sap, df_sobrepeso_real, df_base_fisica, log_callback)
            sp_total += ajuste_sp

            itens_detalhados.append({
                'sku': sku,
                'chave_pallet': chave_frac,
                'sp': round(sp, 4),
                'ajuste_sp': round(ajuste_sp, 2),
                'origem': origem_sp
            })

        for _ in range(qtd_fixo):
            peso_base_total, peso_base_total_liq, sp_total = processar_sobrepeso_fixo_basico(
                sku, qtd, df_sku, df_base_fisica,
                peso_base_total, peso_base_total_liq, sp_total,
                itens_detalhados, log_callback
            )

    df_expedicao_com_pallet = df_remessa[
        (df_remessa['CHAVE_PALETE'].notna()) &
        (~df_remessa['CHAVE_PALETE'].isin(chaves_pallet_processadas))
    ]
    for _, row in df_expedicao_com_pallet.iterrows():
        sku = row['ITEM']
        chave_pallet = row['CHAVE_PALETE']
        qtd_caixas = pd.to_numeric(row['QUANTIDADE'], errors='coerce')
        if pd.isna(qtd_caixas):
            continue
        chaves_pallet_processadas.append(chave_pallet)
        resultado_ext = calcular_peso_receb_externo(
            chave_pallet, sku, row, df_externo_peso, df_estoque_sep,df_sku,
            peso_base_total, peso_base_total_liq, itens_detalhados, log_callback
        )
        if resultado_ext:
            peso_base_total, peso_base_total_liq = resultado_ext
            continue
        df_sku_filtrado = df_sku[df_sku['COD_PRODUTO'] == sku]
        if df_sku_filtrado.empty:
            log_callback(f"SKU {sku} não encontrado na base SKU.")
            continue

        peso_bruto = pd.to_numeric(df_sku_filtrado.iloc[0]['QTDE_PESO_BRU'], errors='coerce')
        peso_liq = pd.to_numeric(df_sku_filtrado.iloc[0]['QTDE_PESO_LIQ'], errors='coerce')

        peso_base = float(peso_bruto) * float(qtd_caixas) if pd.notna(peso_bruto) else 0
        peso_base_liq = float(peso_liq) * float(qtd_caixas) if pd.notna(peso_liq) else 0

        peso_base_total += peso_base
        peso_base_total_liq += peso_base_liq

        sp, origem_sp, ajuste_sp = processar_sobrepeso(chave_pallet, sku, peso_base_liq, df_sap, df_sobrepeso_real, df_base_fisica, log_callback)
        sp_total += ajuste_sp

        itens_detalhados.append({
            'sku': sku,
            'chave_pallet': chave_pallet,
            'sp': round(sp, 4),
            'ajuste_sp': round(ajuste_sp, 2),
            'origem': origem_sp
        })

    peso_com_sobrepeso = peso_base_total + sp_total
    log_callback(f"Peso com sobrepeso: {peso_com_sobrepeso:.2f} kg")
    peso_total_com_paletes = peso_com_sobrepeso + (qtd_paletes * 25) + peso_veiculo_vazio
    log_callback(f"Peso total com paletes ({qtd_paletes} x 25kg): {peso_total_com_paletes:.2f} kg")
    media_sp_geral = (sum(item['sp'] for item in itens_detalhados) / len(itens_detalhados)) if itens_detalhados else 0.0
    log_callback(f"Média geral de sobrepeso (entre {len(itens_detalhados)} itens): {media_sp_geral:.4f}")

    qtd_real = sum(1 for i in itens_detalhados if i['origem'] == 'real')
    qtd_ext  = sum(1 for i in itens_detalhados if i['origem'] == 'receb_ext')
    qtd_fixo = sum(1 for i in itens_detalhados if i['origem'] == 'fixo')

    log_callback(f"📦 Total de itens processados: {len(itens_detalhados)}")
    log_callback(f"├── reais: {sum(1 for i in itens_detalhados if i.get('origem') in ['real', 'receb_ext'])}")
    log_callback(f"├── receb_ext: {sum(1 for i in itens_detalhados if i.get('origem') == 'receb_ext')}")
    log_callback(f"└── fixos: {sum(1 for i in itens_detalhados if i.get('origem') == 'fixo')}")

    repetidas = [chave for chave, count in Counter(chaves_pallet_processadas).items() if count > 1]
    if repetidas:
        log_callback(f"⚠️ Foram encontradas chaves pallet repetidas: {', '.join(repetidas)}")
        messagebox.showwarning("Chaves Repetidas, por favor ALERTE ao Responsável pela Conferência da Remessa",
                                f"Foram encontradas {len(repetidas)} chaves_pallet repetidas no processo:\n\n" + "\n".join(repetidas))


    return peso_base_total, sp_total, peso_com_sobrepeso, peso_total_com_paletes, media_sp_geral, itens_detalhados

def calcular_limites_sobrepeso_por_quantidade(dados, itens_detalhados, df_base_familia, df_sobrepeso_tabela, df_sku, df_remessa, df_fracao, log_callback):
    total_quantidade = 0
    quantidade_com_sp_real = 0
    ponderador_pos = 0
    ponderador_neg = 0
    familia_detectada = "MIX"
    agrupado_por_sku = defaultdict(list)
    for item in itens_detalhados:
        agrupado_por_sku[item['sku']].append(item)

    for sku, itens in agrupado_por_sku.items():
        qtd_total = 0
        qtd_real = 0
        ponderador_pos_local = 0
        ponderador_neg_local = 0 
        for item in itens:
            origem = item.get('origem', 'fixo')
            sp = item.get('sp', 0)
            chave = item.get('chave_pallet', '')
            qtd = 0
            if chave in df_fracao['chave_pallete'].values:
                qtd = pd.to_numeric(df_fracao[df_fracao['chave_pallete'] == chave]['qtd'], errors='coerce').sum()
            else:
                qtd = pd.to_numeric(df_remessa[df_remessa['ITEM'] == sku]['QUANTIDADE'], errors='coerce').sum()

            qtd_total += qtd

            if origem in ['real', 'receb_ext']:
                qtd_real += qtd
                if sp > 0:
                    ponderador_pos_local += sp * qtd
                elif sp < 0:
                    ponderador_neg_local += abs(sp) * qtd

        total_quantidade += qtd_total
        quantidade_com_sp_real += qtd_real
        ponderador_pos += ponderador_pos_local
        ponderador_neg += ponderador_neg_local

    proporcao_sp_real = quantidade_com_sp_real / total_quantidade if total_quantidade > 0 else 0
    log_callback(f"Total de quantidade: {total_quantidade}, com SP Real: {quantidade_com_sp_real}, proporção: {proporcao_sp_real:.2%}")

    if proporcao_sp_real >= 0.5:
        if quantidade_com_sp_real > 0:
            media_positiva = ponderador_pos / quantidade_com_sp_real if ponderador_pos > 0 else 0.02
            media_negativa = ponderador_neg / quantidade_com_sp_real if ponderador_neg > 0 else 0.01
        else:
            media_positiva = 0.02
            media_negativa = 0.01
        log_callback("Mais de 50% da quantidade com SP Real. Usando médias ponderadas:")
        log_callback(f"Sobrepeso para mais (real): {media_positiva:.4f}")
        log_callback(f"Sobrepeso para menos (real): {media_negativa:.4f}")
    else:
        log_callback("Menos de 50% da quantidade com SP Real. Usando tabela de sobrepeso físico.")
        familias = set()
        for sku in agrupado_por_sku:
            familia_series = df_base_familia.loc[df_base_familia['CÓD'] == sku, 'FAMILIA 2']
            if not familia_series.empty:
                familia_str=str(familia_series.iloc[0])
                familias.add(familia_str)

        if len(familias) == 1:
            familia_str = list(familias)[0]
            if 'BISCOITO' in familia_str.upper():
                familia_detectada = "BISCOITO"
                row = df_sobrepeso_tabela.loc[df_sobrepeso_tabela.index.str.contains("BISCOITO", case=False)]
            elif 'MASSA' in familia_str.upper():
                familia_detectada = 'MASSA'
                row = df_sobrepeso_tabela.loc[df_sobrepeso_tabela.index.str.contains("MASSA", case=False)]
            else:
                familia_detectada = "MIX"
                row = df_sobrepeso_tabela.loc[df_sobrepeso_tabela.index.str.contains("MIX", case=False)]
        else:
            familia_detectada = "MIX"
            row = df_sobrepeso_tabela.loc[df_sobrepeso_tabela.index.str.contains("MIX", case=False)]

        if row.empty:
            log_callback("Família não encontrada na tabela. Usando MIX como fallback.")
            row = df_sobrepeso_tabela.loc[df_sobrepeso_tabela.index.str.contains("MIX", case=False)]

        media_positiva = row['(+)'].values[0]
        media_negativa = row['(-)'].values[0]

        log_callback(f"Sobrepeso para mais (físico): {media_positiva:.4f}")
        log_callback(f"Sobrepeso para menos (físico): {media_negativa:.4f}")

    return media_positiva, media_negativa, proporcao_sp_real, familia_detectada

def preencher_formulario_com_openpyxl(path_copia, dados, itens_detalhados, log_callback,df_sku, df_remessa, df_fracao):
    try:
        dados_tabela = {
        '(+)': [0.02, 0.005, 0.04],
        '(-)': [0.01, 0.01, 0.01]
        }
        index = ['CARGA COM MIX', 'EXCLUSIVO MASSAS', 'EXCLUSIVO BISCOITOS']
        df_sobrepeso_tabela = pd.DataFrame(dados_tabela, index=index)
        df_base_familia = pd.read_excel(caminho_base_fisica, 'BASE_FAMILIA')
        sp_pos, sp_neg, proporcao_sp_real, familia_detectada = calcular_limites_sobrepeso_por_quantidade(
            dados, itens_detalhados, df_base_familia, df_sobrepeso_tabela, df_sku, df_remessa, df_fracao, log_callback
        )
        wb = load_workbook(path_copia)
        ws = wb["FORMULARIO"]

        log_callback("Preenchendo cabeçalhos principais com openpyxl...")
        ws["A16"] = f"Sobrepeso para (+): {sp_pos*100:.2f}%"
        ws["A18"] = f"Sobrepeso para (-): {sp_neg*100:.2f}%"
        ws["D7"] = f"{proporcao_sp_real*100:.2f}% x {(1 - proporcao_sp_real)*100:.2f}%"
        try:
            ws["B5"] = str(familia_detectada) if familia_detectada is not None else "MIX"
        except Exception as e:
            log_callback(f"Erro ao escrever família detectada: {e}")
            ws["B5"] = "MIX" 
        ws["B4"] = dados['remessa']
        ws["B6"] = dados['qtd_skus']
        ws["B7"] = dados['placa']
        ws["B8"] = dados['turno']
        ws["B9"] = dados['peso_vazio']
        ws["B10"] = dados['peso_base']
        ws["B11"] = dados['sp_total']
        ws["B12"] = dados['peso_com_sp']
        ws["B13"] = dados['peso_total_final']
        ws["B14"] = dados['media_sp']
        ws["B16"] = dados['peso_total_final'] * (1 + sp_pos)
        ws["B17"] = dados['peso_total_final']
        ws["B18"] = dados['peso_total_final'] * (1 - sp_neg)
        ws["D4"] = dados['qtd_paletes']
        ws["D9"] = dados['qtd_paletes'] * 25

        linha_inicio = 12
        linha_fim = 46
        max_itens = linha_fim - linha_inicio + 1

        log_callback("Preenchendo SKUs e sobrepesos...")
        itens_real = [item for item in itens_detalhados if item['origem'] == 'real']
        itens_fixo = [item for item in itens_detalhados if item['origem'] == 'fixo']
        itens_nao_encontrado = [item for item in itens_detalhados if item['origem'] == 'não encontrado']

        itens_ordenados = itens_real + itens_fixo + itens_nao_encontrado
        for idx, item in enumerate(itens_ordenados):
            if idx >= max_itens:
                log_callback("Limite máximo de linhas atingido no formulário (C12 até C46). Restante será desconsiderado.")
                break
            linha = linha_inicio + idx
            sku_texto = f"{item['sku']} ({item['origem']})"
            ws[f"C{linha}"] = sku_texto
            ws[f"D{linha}"] = f"{item['sp']*100:.3f}"

        wb.save(path_copia)
        wb.close()
        log_callback("Formulário preenchido e salvo com sucesso.")

    except Exception as e:
        log_callback(f"Erro no preenchimento: {e}")
        raise

def exportar_pdf_com_comtypes(path_xlsx, aba_nome="FORMULARIO", nome_remessa="REMESSA", log_callback=None):
    try:
        if log_callback:
            log_callback("Iniciando exportação via comtypes...")

        comtypes.CoInitialize()
        excel = comtypes.client.CreateObject("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(path_xlsx)

        if aba_nome not in [sheet.Name for sheet in wb.Sheets]:
            raise Exception(f"Aba '{aba_nome}' não encontrada.")

        ws = wb.Worksheets(aba_nome)

        pdf_dir = os.path.join(fonte_dir, 'Relatório_Saida')
        os.makedirs(pdf_dir, exist_ok=True)
        pdf_path = os.path.join(pdf_dir, f"SOBREPESOSIMULADO - {nome_remessa}.pdf")
        if log_callback:
            log_callback(f"Tentando exportar para: {pdf_path}")
        ws.ExportAsFixedFormat(Type=0, Filename=pdf_path)
        wb.Close(SaveChanges=False)
        excel.Quit()
        del ws
        del wb
        del excel
        gc.collect()
        comtypes.CoUninitialize()
        time.sleep(5)
        if log_callback:
            log_callback(f"PDF exportado com sucesso: {pdf_path}")
        return pdf_path
    except Exception as e:
        if log_callback:
            log_callback(f"Erro ao exportar PDF: {e}")
        raise

def gerar_relatorio_diferenca(remessa_num, peso_final_balança, peso_veiculo_vazio, df_remessa, df_sku, peso_estimado_total, pasta_excel, log_callback):
    try:
        peso_final_balança = converter_para_float_seguro(peso_final_balança)
        peso_veiculo_vazio = converter_para_float_seguro(peso_veiculo_vazio)
        peso_estimado_total = converter_para_float_seguro(peso_estimado_total)
        
        pasta_destino = os.path.join(pasta_excel, 'Analise_divergencia')
        os.makedirs(pasta_destino, exist_ok=True)

        skus = df_remessa['ITEM'].unique()
        dados_relatorio = []
        peso_base_total_liq = 0.0

        for sku in skus:
            qtd = converter_para_float_seguro(df_remessa[df_remessa['ITEM'] == sku]['QUANTIDADE'].sum())
            
            df_sku_filtrado = df_sku[df_sku['COD_PRODUTO'] == sku]
            if df_sku_filtrado.empty:
                continue

            unidade = df_sku_filtrado.iloc[0]['DESC_UNID_MEDID']
            peso_unit_liq = converter_para_float_seguro(df_sku_filtrado.iloc[0]['QTDE_PESO_LIQ'])
            peso_total_liq = converter_para_float_seguro(peso_unit_liq * qtd)
            peso_base_total_liq += peso_total_liq

            dados_relatorio.append({
                'SKU': sku,
                'Unidade': unidade,
                'Quantidade': qtd,
                'Peso Total Líquido': peso_total_liq,
                'Peso Unit. Líquido': peso_unit_liq
            })

        df_dados = pd.DataFrame(dados_relatorio)
        peso_carga_real = converter_para_float_seguro(peso_final_balança - peso_veiculo_vazio)
        diferenca_total = converter_para_float_seguro((peso_estimado_total + peso_veiculo_vazio) - peso_final_balança)
        
        if peso_base_total_liq > 0:
            df_dados['% Peso'] = df_dados['Peso Total Líquido'] / peso_base_total_liq
        else:
            df_dados['% Peso'] = 0.0

        df_dados['Peso Proporcional Real'] = df_dados['% Peso'] * peso_carga_real
        df_dados['Quantidade Real Estimada'] = df_dados.apply(
            lambda x: round(x['Peso Proporcional Real'] / x['Peso Unit. Líquido']) if x['Peso Unit. Líquido'] > 0 else 0,
            axis=1
        )
        
        df_dados['Diferença Estimada (kg)'] = df_dados['% Peso'] * diferenca_total
        df_dados['Unid. Estimada de Divergência'] = df_dados['Quantidade Real Estimada'] - df_dados['Quantidade']

        nome_pdf = f"Análise quantitativa - {remessa_num}.pdf"
        caminho_pdf = os.path.join(pasta_destino, nome_pdf)
        fig, ax = plt.subplots(figsize=(10, 6))
        largura_barra = 0.35
        x = range(len(df_dados))

        qtd_esperada = df_dados['Quantidade']
        qtd_real = df_dados['Quantidade Real Estimada']

        ax.bar([i - largura_barra/2 for i in x], qtd_esperada, width=largura_barra, label='Quantidade Esperada')
        ax.bar([i + largura_barra/2 for i in x], qtd_real, width=largura_barra, label='Quantidade Real Estimada')

        ax.set_ylabel("Quantidade (unidades)")
        ax.set_xlabel("SKU")
        ax.set_title("Comparativo: Quantidade Esperada vs Real Estimada por SKU")
        ax.set_xticks(x)
        ax.set_xticklabels(df_dados['SKU'].astype(str))
        ax.axhline(0, color='gray', linestyle='--')
        ax.legend()

        for i in x:
            ax.text(i - largura_barra/2, qtd_esperada.iloc[i], f"{qtd_esperada.iloc[i]:.0f}", ha='center', va='bottom')
            ax.text(i + largura_barra/2, qtd_real.iloc[i], f"{qtd_real.iloc[i]:.0f}", ha='center', va='bottom')

        with PdfPages(caminho_pdf) as pdf:
            fig_tabela, ax_tabela = plt.subplots(figsize=(12, len(df_dados) * 0.5 + 3))
            ax_tabela.axis('off')
            table_data = [
                ['SKU', 'Unidade', 'Qtd. Enviada', 'Qtd. Real Estimada', 'Peso Total Líquido', '% do Peso', 'Diferença (kg)', 'Divergência (unid)']
            ] + df_dados[['SKU', 'Unidade', 'Quantidade', 'Quantidade Real Estimada', 'Peso Total Líquido', '% Peso', 'Diferença Estimada (kg)', 'Unid. Estimada de Divergência']].round(2).values.tolist()

            tabela = ax_tabela.table(cellText=table_data, colLabels=None, loc='center', cellLoc='center')
            tabela.auto_set_font_size(False)
            tabela.set_fontsize(10)
            tabela.scale(1, 1.5)

            titulo = f"Relatório Comparativo - Remessa {remessa_num}\nPeso estimado: {peso_estimado_total:.2f} kg | Peso balança: {peso_final_balança:.2f} kg | Peso veículo: {peso_veiculo_vazio:.2f} kg | Diferença: {diferenca_total:.2f} kg"
            fig_tabela.suptitle(titulo, fontsize=12)
            pdf.savefig(fig_tabela, bbox_inches='tight')
            pdf.savefig(fig, bbox_inches='tight')

        plt.close('all')
        log_callback(f"Relatório salvo em: {caminho_pdf}")
        return caminho_pdf

    except Exception as e:
        log_callback(f"Erro ao gerar relatório de divergência: {str(e)}")
        raise

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")
class EdicaoRemessaFrame(ctk.CTkFrame):
    def __init__(self, master, df_expedicao, log_callback, app, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.fonte_dir = fonte_dir
        self.remessa_editada = False
        self.df_expedicao_original = df_expedicao.dropna(subset=['ITEM']).copy()
        self.dados_remessa = pd.DataFrame(columns=['ITEM', 'QUANTIDADE', 'CHAVE_PALETE'])
        self.frame_superior = ctk.CTkFrame(self)
        self.frame_superior.pack(fill="x", padx=10, pady=(10, 0))

        self.label_status = ctk.CTkLabel(
            self.frame_superior,
            text="",
            text_color="green",
            font=("Arial", 12, "bold"),
            corner_radius=5,
            padx=10,
            pady=5
        )
        self.label_status.pack(fill="x")
        self.df_expedicao = df_expedicao
        self.log_callback = log_callback
        self.app = app
        self.remessa_var = StringVar()
        self.filtro_chave = StringVar()
        self.filtro_sku = StringVar()
        self.lista_edicao = []
        self.sku_totals = {}
        top_button_frame = ctk.CTkFrame(self)
        top_button_frame.pack(fill="x", padx=10, pady=5)
        ctk.CTkButton(
            top_button_frame, 
            text="➕ Adicionar Linha", 
            command=self.adicionar_linha,
            width=150,
            fg_color="#2a7fff"
        ).pack(side="left", padx=5)
    
        ctk.CTkButton(
            top_button_frame, 
            text="💾 Salvar Alterações", 
            command=self.salvar_alteracoes,
            width=150
        ).pack(side="right", padx=5)

        ctk.CTkLabel(top_button_frame, text="Edição de Remessa", font=("Arial", 14, "bold")).pack(side="left", padx=5)
        form_frame = ctk.CTkFrame(self)
        form_frame.pack(fill="x", padx=10, pady=5)

        ctk.CTkLabel(form_frame, text="Número da Remessa:").grid(row=0, column=0, sticky="w", padx=5)
        self.entry_remessa = ctk.CTkEntry(form_frame, textvariable=self.remessa_var)
        self.entry_remessa.grid(row=0, column=1, sticky="ew", padx=5)
        ctk.CTkButton(form_frame, text="🔎 Buscar", command=self.carregar_dados).grid(row=0, column=2, padx=5)

        ctk.CTkLabel(form_frame, text="Filtrar por Chave Pallet:").grid(row=1, column=0, sticky="w", padx=5)
        self.entry_filtro_chave = ctk.CTkEntry(form_frame, textvariable=self.filtro_chave)
        self.entry_filtro_chave.grid(row=1, column=1, sticky="ew", padx=5)
        self.entry_filtro_chave.bind("<KeyRelease>", lambda e: self.filtrar_dados())
        
        ctk.CTkLabel(form_frame, text="Filtrar por SKU:").grid(row=2, column=0, sticky="w", padx=5)
        self.entry_filtro_sku = ctk.CTkEntry(form_frame, textvariable=self.filtro_sku)
        self.entry_filtro_sku.grid(row=2, column=1, sticky="ew", padx=5)
        self.entry_filtro_sku.bind("<KeyRelease>", lambda e: self.filtrar_dados())
        
        ctk.CTkButton(form_frame, text="🔍 Aplicar Filtros", command=self.filtrar_dados).grid(row=1, column=2, rowspan=2, padx=5)
        self.frame_horizontal = ctk.CTkFrame(self)
        self.frame_horizontal.pack(fill="x", padx=10, pady=5)
        self.tabela_frame = ctk.CTkScrollableFrame(self.frame_horizontal, height=350)
        self.tabela_frame.pack(side="left", fill="both", expand=True, padx=(0, 5))
        self.totals_center_frame = ctk.CTkFrame(self.frame_horizontal, width=150, fg_color="#1a1a1a")
        self.totals_center_frame.pack(side="left", fill="y", padx=5)
        ctk.CTkLabel(
            self.totals_center_frame,
            text="Totais Gerais:",
            font=("Arial", 12, "bold"),
        ).pack(pady=(10, 5))

        self.total_itens_frame = ctk.CTkFrame(self.totals_center_frame, corner_radius=6, fg_color="#2a2a2a")
        self.total_itens_frame.pack(fill="x", padx=5, pady=2)
        ctk.CTkLabel(self.total_itens_frame, text="Itens:", text_color="#FF5555").pack(side="left", padx=5)
        self.total_itens_value = ctk.CTkLabel(self.total_itens_frame, text="0", text_color="#55FF55", font=("Arial", 10, "bold"))
        self.total_itens_value.pack(side="right", padx=5)

        self.total_qtd_frame = ctk.CTkFrame(self.totals_center_frame, corner_radius=6, fg_color="#2a2a2a")
        self.total_qtd_frame.pack(fill="x", padx=5, pady=2)
        ctk.CTkLabel(self.total_qtd_frame, text="Qtd Total:", text_color="#FF5555").pack(side="left", padx=5)
        self.total_qtd_value = ctk.CTkLabel(self.total_qtd_frame, text="0", text_color="#55FF55", font=("Arial", 10, "bold"))
        self.total_qtd_value.pack(side="right", padx=5)

        self.sku_totals_lateral = ctk.CTkFrame(
            self.frame_horizontal, 
            width=350, 
            fg_color="#1f1f1f"
        )
        self.sku_totals_lateral.pack_propagate(False) 
        self.sku_totals_lateral.pack(side="left", fill="y", padx=(5, 0))

        self.sku_totals_label = ctk.CTkLabel(self.sku_totals_lateral, text="Totais por SKU:")
        self.sku_totals_label.pack(anchor="w", padx=5, pady=(0, 5))
        self.filtro_sku_totais = StringVar()
        self.entry_filtro_sku_totais = ctk.CTkEntry(self.sku_totals_lateral, textvariable=self.filtro_sku_totais, placeholder_text="Filtrar SKU")
        self.entry_filtro_sku_totais.pack(fill="x", padx=5, pady=(0, 5))
        self.entry_filtro_sku_totais.bind("<KeyRelease>", lambda e: self.atualizar_totais_sku(self.filtro_sku_totais.get()))
        self.sku_totals_scroll = ctk.CTkScrollableFrame(self.sku_totals_lateral, height=350)
        self.sku_totals_scroll.pack(fill="both", expand=True)

    def format_number(self, num):
        if pd.isna(num):
            return "N/A"
        try:
            num_float = float(num)
            if num_float.is_integer():
                return str(int(num_float))
            return str(num_float)
        except:
            return str(num)
    
    def adicionar_linha(self):
        try:
            nova_linha = {
                'ITEM': '',          
                'QUANTIDADE': 0,     
                'CHAVE_PALETE': None 
            }
            
            df_nova_linha = pd.DataFrame([nova_linha])
            self.dados_remessa = pd.concat([self.dados_remessa, df_nova_linha], ignore_index=True)
            self.renderizar_tabela()
            self.tabela_frame._parent_canvas.yview_moveto(1.0)
            
            self.label_status.configure(
                text="Nova linha adicionada - preencha os dados",
                text_color="blue"
            )
            self.log_callback("Nova linha adicionada para edição manual")
            
        except Exception as e:
            self.log_callback(f"Erro ao adicionar linha: {str(e)}")
            self.label_status.configure(
                text=f"Erro ao adicionar linha: {str(e)}",
                text_color="red"
            )
    def atualizar_totais_sku(self, filtro=""):
        for widget in self.sku_totals_scroll.winfo_children():
            widget.destroy()
        filtro = filtro.lower().strip()
        skus_filtrados = {
            sku: total 
            for sku, total in self.sku_totals.items() 
            if filtro in str(sku).lower()
        }
        self.renderizar_totais_sku(skus_filtrados)

    def renderizar_totais_sku(self, skus_totais):
        for widget in self.sku_totals_scroll.winfo_children():
            widget.destroy()
    
        sorted_skus = sorted(skus_totais.items(), key=lambda x: str(x[0]))
        
        max_columns = 4
        row = 0
        col = 0

        for sku, total in sorted_skus:
            item_frame = ctk.CTkFrame(self.sku_totals_scroll, fg_color="#2a2a2a", corner_radius=8)
            item_frame.grid(row=row, column=col, padx=5, pady=2, sticky="w")

            sku_str = self.format_number(sku)
            total_str = self.format_number(total)

            ctk.CTkLabel(
                item_frame, 
                text=sku_str,
                text_color="#FF5555",
                font=("Arial", 10, "bold"),
                anchor="w"
            ).pack(side="left")

            ctk.CTkLabel(item_frame, text=" = ").pack(side="left")

            ctk.CTkLabel(
                item_frame, 
                text=total_str,
                text_color="#55FF55",
                font=("Arial", 10, "bold"),
                anchor="w"
            ).pack(side="left")
            
            col += 1
            if col >= max_columns:
                col = 0
                row += 1
        
    def update_sku_totals(self, event=None):
        try:
            self.sku_totals = {}
            for entry_sku, entry_qtd, _ in self.entry_widgets:
                if entry_sku:
                    sku = entry_sku.get()
                else:
                    sku = getattr(entry_qtd, 'sku_associado', '')
                if not sku:
                    continue
                qtd_str = entry_qtd.get() if entry_qtd.get() else "0"
                try:
                    qtd = float(qtd_str)
                except ValueError:
                    qtd = 0.0
                self.sku_totals[sku] = self.sku_totals.get(sku, 0) + qtd
            self.renderizar_totais_sku(self.sku_totals)
        except Exception as e:
            self.log_callback(f"Erro ao atualizar totais por SKU: {str(e)}")

    def update_totals(self, event=None):
        try:
            total_itens = len([x for x in self.entry_widgets if (x[0] and x[0].get()) or (not x[0] and x[1].sku_associado)])
            total_qtd = sum(float(entry_qtd.get()) if entry_qtd.get() else 0 
                        for _, entry_qtd, _ in self.entry_widgets)
            
            self.total_itens_value.configure(text=str(total_itens))
            self.total_qtd_value.configure(text=self.format_number(total_qtd))
            self.update_sku_totals()
        except Exception as e:
            self.log_callback(f"Erro ao atualizar totais: {str(e)}")

    def remessa_existe_na_base(self, remessa, df_base):
        """Comparação segura de remessa, ignorando .0 e diferenças de formatação"""
        if df_base.empty or 'REMESSA' not in df_base.columns:
            return False
        try:
            remessas_base = df_base['REMESSA'].dropna().astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
            remessa_busca = str(remessa).replace('.0', '').strip()
            return remessa_busca in remessas_base.values
        except Exception as e:
            self.log_callback(f"Erro ao verificar remessa: {str(e)}")
            return False
    
    def carregar_dados(self):
        remessa = self.remessa_var.get().strip()
        if not remessa:
            self.log_callback("Digite uma remessa válida.")
            return
        
        try:
            self.log_callback(f"Verificando base auxiliar para remessa {remessa}...")
            df_base_auxiliar = carregar_base_auxiliar(self.fonte_dir)
            
            if not df_base_auxiliar.empty:
                self.log_callback(f"Base auxiliar carregada com {len(df_base_auxiliar)} registros")
                remessas_aux = df_base_auxiliar['REMESSA'].dropna().astype(str).str.replace(r'\.0$', '', regex=True)
                self.log_callback(f"Remessas encontradas na base auxiliar: {', '.join(remessas_aux.unique())}")
                
                if self.remessa_existe_na_base(remessa, df_base_auxiliar):
                    self.remessa_editada = True
                    self.label_status.configure(
                        text="ATENÇÃO: Esta remessa já foi editada anteriormente!",
                        text_color="orange"
                    )
                    self.log_callback(f"Remessa {remessa} encontrada na base auxiliar - carregando dados editados")
                    df_filtrado = df_base_auxiliar[
                        df_base_auxiliar['REMESSA'].astype(str).str.replace(r'\.0$', '', regex=True) == remessa
                    ]
                else:
                    self.log_callback(f"Remessa {remessa} não encontrada na base auxiliar - buscando na base original")
            else:
                self.log_callback("Base auxiliar vazia ou não encontrada - buscando na base original")

            if not self.remessa_editada:
                self.log_callback(f"Buscando remessa {remessa} na base original...")
                df_filtrado = self.df_expedicao_original[
                    self.df_expedicao_original['REMESSA'].astype(str) == remessa
                ].dropna(subset=['ITEM']).copy()
                
                if df_filtrado.empty:
                    self.log_callback("Remessa não encontrada em nenhuma base!")
                    self.label_status.configure(
                        text="Remessa não encontrada em nenhuma base!",
                        text_color="red"
                    )
                    return
            colunas_unicas = ['ITEM', 'QUANTIDADE', 'CHAVE_PALETE']
            df_sem_duplicatas = df_filtrado.drop_duplicates(subset=colunas_unicas)
            
            if len(df_filtrado) != len(df_sem_duplicatas):
                duplicatas = len(df_filtrado) - len(df_sem_duplicatas)
                self.log_callback(f"Removidas {duplicatas} linhas duplicadas automaticamente.")
            
            self.dados_remessa = df_sem_duplicatas[['ITEM', 'QUANTIDADE', 'CHAVE_PALETE']].copy()
            self.renderizar_tabela()
            
        except Exception as e:
            error_msg = f"Erro ao carregar dados: {str(e)}"
            self.log_callback(error_msg)
            self.label_status.configure(
                text=error_msg,
                text_color="red"
            )
            self.log_callback(f"Traceback completo: {traceback.format_exc()}")
    
    def salvar_alteracoes_antes_filtro(self):
        if not hasattr(self, 'entry_widgets') or not self.entry_widgets:
            return
        
        try:
            for entry_sku, entry_qtd, entry_chave in self.entry_widgets:
                original_idx = getattr(entry_qtd, 'original_idx', None)
                
                if original_idx is None or original_idx not in self.dados_remessa.index:
                    continue
                sku = entry_sku.get() if entry_sku else getattr(entry_qtd, 'sku_associado', '')
                qtd = entry_qtd.get()
                chave = entry_chave.get()

                self.dados_remessa.at[original_idx, 'ITEM'] = sku if sku else None
                self.dados_remessa.at[original_idx, 'QUANTIDADE'] = float(qtd) if qtd else 0
                self.dados_remessa.at[original_idx, 'CHAVE_PALETE'] = chave if chave else None
                
            self.log_callback("Alterações salvas antes de aplicar filtros")
        except Exception as e:
            self.log_callback(f"Erro ao salvar alterações antes de filtrar: {str(e)}")

    def filtrar_dados(self):
        self.salvar_alteracoes_antes_filtro()
        filtro_chave = self.filtro_chave.get().strip().lower()
        filtro_sku = self.filtro_sku.get().strip().lower()
        
        df_filtrado = self.dados_remessa.copy()
        
        if filtro_chave:
            df_filtrado = df_filtrado[
                df_filtrado['CHAVE_PALETE'].fillna('').astype(str).str.lower().str.contains(filtro_chave)
            ]
        
        if filtro_sku:
            df_filtrado = df_filtrado[
                df_filtrado['ITEM'].astype(str).str.lower().str.contains(filtro_sku)
            ]
        self.renderizar_tabela(df_filtrado)

    def salvar_alteracoes(self):
        try:
            remessa = self.remessa_var.get().strip()
            if not remessa:
                self.log_callback("Nenhuma remessa selecionada para salvar")
                return
            
            self.master.update()  
            self.master.configure(cursor="watch")  

            df_completo = pd.DataFrame(columns=['ITEM', 'QUANTIDADE', 'CHAVE_PALETE'])
            
            for entry_sku, entry_qtd, entry_chave in self.entry_widgets:
           
                sku = entry_sku.get() if entry_sku else getattr(entry_qtd, 'sku_associado', '')
                qtd = entry_qtd.get() if entry_qtd.get() else "0"
                chave = entry_chave.get() if entry_chave else ""

                if not sku or str(sku).strip() == "":
                    continue
                try:
                    qtd = float(qtd)
                except ValueError:
                    qtd = 0.0
                df_completo = pd.concat([df_completo, pd.DataFrame([{
                    'ITEM': sku,
                    'QUANTIDADE': qtd,
                    'CHAVE_PALETE': chave if chave else None
                }])], ignore_index=True)
       
            df_completo['REMESSA'] = remessa
            if df_completo.empty:
                self.label_status.configure(
                    text="Nenhum dado válido para salvar!",
                    text_color="orange"
                )
                return
            salvar_em_base_auxiliar(df_completo, remessa, self.log_callback, self.fonte_dir)
            self.remessa_editada = True
            self.dados_remessa = df_completo[['ITEM', 'QUANTIDADE', 'CHAVE_PALETE']].copy()
            self.label_status.configure(
                text=f"Alterações salvas com sucesso! Itens: {len(df_completo)}",
                text_color="green"
            )
            self.log_callback(f"Remessa {remessa} salva com {len(df_completo)} itens")
            
        except Exception as e:
            self.label_status.configure(
                text=f"Erro ao salvar alterações: {str(e)}",
                text_color="red"
            )
            self.log_callback(f"[ERRO CRÍTICO] Ao salvar: {traceback.format_exc()}")
            
        finally:
            self.master.configure(cursor="")
            self.renderizar_tabela()

    def renderizar_tabela(self, df_filtrado=None):
        try:
            # Limpar completamente o frame antes de renderizar
            for widget in self.tabela_frame.winfo_children():
                widget.destroy()

            # Usar dados filtrados ou todos os dados
            df_exibicao = df_filtrado.copy() if df_filtrado is not None else self.dados_remessa.copy()
            
            # Remover duplicatas para exibição (caso haja alguma no DataFrame)
            df_exibicao = df_exibicao.drop_duplicates(
                subset=['ITEM', 'QUANTIDADE', 'CHAVE_PALETE'],
                keep='last'
            )

            if df_exibicao.empty:
                self.log_callback("Nenhum dado para exibir.")
                return
                
            # Resetar a lista de widgets
            self.entry_widgets = []

            # Cabeçalhos
            headers = ["SKU", "Quantidade", "Chave Pallet", "Ações"]
            for col, header in enumerate(headers):
                ctk.CTkLabel(self.tabela_frame, 
                            text=header, 
                            font=("Arial", 12, "bold")).grid(
                    row=0, column=col, padx=10, pady=5, sticky="ew"
                )

            # Linhas de dados
            for display_row, (idx, row) in enumerate(df_exibicao.iterrows(), start=1):
                try:
                    sku = str(row['ITEM']) if pd.notna(row['ITEM']) else ""
                    qtd = str(row['QUANTIDADE']) if pd.notna(row['QUANTIDADE']) else "0"
                    chave = str(row['CHAVE_PALETE']) if pd.notna(row['CHAVE_PALETE']) else ""

                    # Widget SKU
                    if sku == "":
                        entry_sku = ctk.CTkEntry(
                            self.tabela_frame, 
                            placeholder_text="Digite o SKU",
                            fg_color="#2a2a4a",
                            border_color="#4a4a7a"
                        )
                        entry_sku.insert(0, sku)
                        entry_sku.grid(row=display_row, column=0, padx=10, pady=2, sticky="ew")
                    else:
                        ctk.CTkLabel(
                            self.tabela_frame, 
                            text=sku
                        ).grid(row=display_row, column=0, padx=10, pady=2, sticky="w")
                        entry_sku = None

                    # Widget Quantidade
                    entry_qtd = ctk.CTkEntry(self.tabela_frame)
                    entry_qtd.insert(0, qtd)
                    entry_qtd.grid(row=display_row, column=1, padx=10, pady=2, sticky="ew")
                    entry_qtd.sku_associado = entry_sku.get() if entry_sku else sku
                    entry_qtd.original_idx = idx
                    entry_qtd.bind("<KeyRelease>", self.update_totals)

                    # Widget Chave
                    entry_chave = ctk.CTkEntry(self.tabela_frame)
                    entry_chave.insert(0, chave)
                    entry_chave.grid(row=display_row, column=2, padx=10, pady=2, sticky="ew")
                    entry_chave.original_idx = idx

                    # Botão Excluir
                    botao_excluir = ctk.CTkButton(
                        self.tabela_frame,
                        text="🗑",
                        width=30,
                        command=lambda idx=idx: self.remover_linha(idx),
                        fg_color="#d44646",
                        hover_color="#a33535"
                    )
                    botao_excluir.grid(row=display_row, column=3, padx=5, pady=2)

                    # Armazenar referências aos widgets
                    if entry_sku:
                        entry_sku.bind("<KeyRelease>", lambda e, entry_qtd=entry_qtd: setattr(entry_qtd, 'sku_associado', e.widget.get()))
                        self.entry_widgets.append((entry_sku, entry_qtd, entry_chave))
                    else:
                        self.entry_widgets.append((None, entry_qtd, entry_chave))

                except Exception as e:
                    self.log_callback(f"Erro ao renderizar linha {idx}: {str(e)}")
                    continue

            self.update_totals()
            
        except Exception as e:
            self.log_callback(f"Erro ao renderizar tabela: {str(e)}")
            raise

    def remover_linha(self, index):
        try:
            if not self.dados_remessa.empty and index in self.dados_remessa.index:
                self.dados_remessa = self.dados_remessa.drop(index)
                self.renderizar_tabela()
                self.label_status.configure(
                    text="Linha removida com sucesso!",
                    text_color="green"
                )
                self.log_callback(f"Linha {index} removida da remessa {self.remessa_var.get()}")
                self.update_totals()
            else:
                self.log_callback(f"Índice inválido para remoção: {index}")
        except Exception as e:
            self.log_callback(f"Erro ao remover linha: {str(e)}")
            self.label_status.configure(
                text=f"Erro ao remover linha: {str(e)}",
                text_color="red"
            )

    def salvar_alteracoes(self):
        try:
            remessa = self.remessa_var.get().strip()
            if not remessa:
                self.log_callback("Nenhuma remessa selecionada para salvar")
                return
            df_completo = pd.DataFrame(columns=['ITEM', 'QUANTIDADE', 'CHAVE_PALETE'])
            for entry_sku, entry_qtd, entry_chave in self.entry_widgets:
                sku = entry_sku.get() if entry_sku else getattr(entry_qtd, 'sku_associado', '')
                qtd = entry_qtd.get() if entry_qtd.get() else "0"
                chave = entry_chave.get() if entry_chave else ""

                if not sku or str(sku).strip() == "":
                    continue
                    
                try:
                    qtd = float(qtd)
                except ValueError:
                    qtd = 0.0
                new_row = pd.DataFrame([{
                    'ITEM': str(sku).strip(),
                    'QUANTIDADE': qtd,
                    'CHAVE_PALETE': str(chave).strip() if chave else None
                }])
                
                df_completo = pd.concat([df_completo, new_row], ignore_index=True)

            if df_completo.empty:
                self.label_status.configure(
                    text="Nenhum dado válido para salvar!",
                    text_color="orange"
                )
                return
            df_completo['REMESSA'] = remessa
            df_completo = df_completo.drop_duplicates(
                subset=['ITEM', 'CHAVE_PALETE'],
                keep='last'
            )
            salvar_em_base_auxiliar(df_completo, remessa, self.log_callback, self.fonte_dir)
            
            self.dados_remessa = df_completo[['ITEM', 'QUANTIDADE', 'CHAVE_PALETE']].copy()
            
            self.renderizar_tabela()
            
            self.label_status.configure(
                text=f"Alterações salvas! Itens: {len(df_completo)}",
                text_color="green"
            )
            self.log_callback(f"Remessa {remessa} salva com {len(df_completo)} itens")
            
        except Exception as e:
            self.label_status.configure(
                text=f"Erro ao salvar: {str(e)}",
                text_color="red"
            )
            self.log_callback(f"[ERRO] Ao salvar: {traceback.format_exc()}")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Simulador de Sobrepeso 3.0")
        self.placa = ctk.StringVar()
        self.turno = ctk.StringVar()
        self.remessa = ctk.StringVar()
        self.qtd_paletes = ctk.StringVar()
        self.peso_vazio = ctk.StringVar()
        self.peso_balanca = ctk.StringVar()
        self.geometry("1200x1000")
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        self.tabs = ctk.CTkTabview(self)
        self.tabs.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        self.tab_simulador = self.tabs.add("Simulador")
        self.tab_simulador.grid_rowconfigure(0, weight=1)
        self.tab_simulador.grid_columnconfigure(0, weight=1)
        simulador_main_frame = ctk.CTkFrame(self.tab_simulador)
        simulador_main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        controls_frame = ctk.CTkFrame(simulador_main_frame, width=350)
        controls_frame.pack(side="left", fill="y", padx=(0, 10), pady=5)
        logs_frame = ctk.CTkFrame(simulador_main_frame)
        logs_frame.pack(side="right", fill="both", expand=True, padx=5, pady=5)
        ctk.CTkLabel(controls_frame, 
                    text="Configuração do Simulador",
                    font=("Arial", 14, "bold")).pack(pady=(0, 10), anchor="w")

        campos = [
            ("Placa:", self.placa, None),
            ("Turno:", self.turno, ["A", "B", "C"]),
            ("Remessa:", self.remessa, None),
            ("Quantidade de Paletes:", self.qtd_paletes, None),
            ("Peso Veículo Vazio (kg):", self.peso_vazio, None),
            ("Peso Final Balança (kg):", self.peso_balanca, None)
        ]

        for texto, var, valores in campos:
            ctk.CTkLabel(controls_frame, text=texto).pack(anchor="w", pady=(5, 0))
            if valores:
                ctk.CTkComboBox(controls_frame, values=valores, variable=var).pack(fill="x", pady=(0, 5))
            else:
                ctk.CTkEntry(controls_frame, textvariable=var).pack(fill="x", pady=(0, 5))

        button_frame = ctk.CTkFrame(controls_frame, fg_color="transparent")
        button_frame.pack(fill="x", pady=10)

        ctk.CTkButton(button_frame, 
                    text="Calcular", 
                    command=self.iniciar_processamento,
                    fg_color="#2aa745").pack(side="left", expand=True, padx=2)
        
        ctk.CTkButton(button_frame, 
                    text="🔄 Refresh", 
                    command=self.atualizar_bases,
                    width=80).pack(side="right", padx=2)

        self.progress_bar = ctk.CTkProgressBar(controls_frame, mode="determinate")
        self.progress_bar.set(0)
        self.progress_bar.pack(fill="x", pady=(5, 0))
        self.progress_bar.pack_forget()

        logs_header = ctk.CTkFrame(logs_frame, fg_color="transparent")
        logs_header.pack(fill="x", pady=(0, 5))

        ctk.CTkLabel(logs_header, 
                    text="📜 Histórico de Execução",
                    font=("Arial", 14, "bold")).pack(side="left", padx=5)
        ctk.CTkButton(logs_header, 
                    text="🧹 Limpar Histórico", 
                    command=self.limpar_logs,
                    width=120,
                    fg_color="#d44646",
                    hover_color="#a33535").pack(side="right")

        self.log_display = ctk.CTkTextbox(
            logs_frame, 
            wrap="word",
            font=("Consolas", 10),  
            activate_scrollbars=True
        )
        self.log_display.pack(fill="both", expand=True)

        self.log_display.configure(state="disabled")
        self.log_text = []
        self.log_geral = []
        self.log_tecnico = []

        self.tab_edicao = self.tabs.add("Edição de Remessa")
        self.df_expedicao = pd.read_excel(
            os.path.join(fonte_dir, "expedicao.xlsx"), 
            sheet_name="dado_exp",
            dtype={"REMESSA": str}
        )
        self.df_expedicao["REMESSA"] = self.df_expedicao["REMESSA"].str.replace(r"\.0$", "", regex=True)
        self.edicao_frame = EdicaoRemessaFrame(
            master=self.tab_edicao,
            df_expedicao=self.df_expedicao,   
            log_callback=self.add_log,       
            app=self
        )
        self.edicao_frame.pack(fill="both", expand=True, padx=10, pady=10)

        footer_label = ctk.CTkLabel(
            self, 
            text="Desenvolvido por Douglas Lins - Analista de Logística", 
            font=("Arial", 10), 
            anchor="center"
        )
        footer_label.grid(row=1, column=0, columnspan=2, pady=(0, 10))

    def atualizar_bases(self):
        try:
            self.add_log("⏳ Forçando atualização das bases do OneDrive...")
            path_base_sobrepeso = os.path.join(fonte_dir, "Base_sobrepeso_real.xlsx")
            path_base_expedicao = os.path.join(fonte_dir, "expedicao.xlsx")
            path_base_sap = os.path.join(fonte_dir, "base_sap.xlsx")
            path_base_frac = os.path.join(fonte_dir, "FRACAO_1.xlsx")
            path_base_estoque = os.path.join(fonte_dir, "estoqueseparacao.xlsx")
            path_base_peso_exter = os.path.join(fonte_dir,"receb_extern_peso.xlsx")
            path_base_aux = os.path.join(fonte_dir, "expedicao_edicoes.xlsx")

            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = False
            wb = excel.Workbooks.Open(path_base_expedicao)
            self.add_log("Atualizando base da expedição, Aguarde...")
            time.sleep(30)
            wb.Close(False)
            if os.path.exists(path_base_aux):
                wb_aux = excel.Workbooks.Open(path_base_aux)
                self.add_log("Atualizando base auxiliar de edições, Aguarde...")
                time.sleep(10)
                wb_aux.Close(False)
            
            excel.Quit()

            mod_time = os.path.getmtime(path_base_expedicao)
            data_mod = datetime.fromtimestamp(mod_time).strftime("%d/%m/%Y %H:%M:%S")
            
            global df_sap, df_estoque_sep, df_externo_peso, df_frac, df_expedicao, df_sobrepeso_real
            df_externo_peso = pd.read_excel(path_base_peso_exter, sheet_name="Sheet1")
            df_estoque_sep = pd.read_excel(path_base_estoque)
            df_frac = pd.read_excel(path_base_frac, sheet_name="FRACAO")
            df_sap = pd.read_excel(path_base_sap, sheet_name="Sheet1")
            
            df_expedicao = pd.read_excel(
                path_base_expedicao, 
                sheet_name="dado_exp",
                dtype={'REMESSA': str}
            )
            df_expedicao['REMESSA'] = df_expedicao['REMESSA'].str.replace(r'\.0$', '', regex=True).astype(str)
            
            df_sobrepeso_real = pd.read_excel(path_base_sobrepeso, sheet_name="SOBREPESO")
            self.edicao_frame.df_expedicao = df_expedicao
            self.df_expedicao = df_expedicao  
            self.edicao_frame.remessa_var.set("")
            self.edicao_frame.filtro_chave.set("")
            self.edicao_frame.filtro_sku.set("")
            for widget in self.edicao_frame.tabela_frame.winfo_children():
                widget.destroy()
            
            self.add_log(f"Bases atualizadas com sucesso! Última modificação: {data_mod}")
            
        except Exception as e:
            self.add_log(f"Erro ao atualizar bases: {e}")
        
    def add_log(self, msg):
        timestamp = datetime.now().strftime("%H:%M:%S")
        entrada = f"[{timestamp}] {msg}"
        self.log_text.append(entrada)
        self.log_display.configure(state="normal")
        self.log_display.insert("end", entrada + "\n")
        self.log_display.see("end")
        self.log_display.configure(state="disabled")

    def limpar_logs(self):
        self.log_text.clear()
        self.log_geral.clear()
        self.log_display.configure(state="normal")
        self.log_display.delete("1.0", "end")
        self.log_display.configure(state="disabled")
    
    def log_callback_completo(self, mensagem):
            print(mensagem)
            self.log_geral.append(mensagem)
            self.add_log(mensagem)

    def log_callback_tecnico(self, mensagem):
        timestamp = datetime.now().strftime("%H:%M:%S")
        entrada = f"[{timestamp}] {mensagem}"
        self.log_tecnico.append(entrada)

    def iniciar_processamento(self):
        self.progress_bar.pack()  
        self.progress_bar.set(0) 
        self.add_log("Iniciando cálculo...")
        thread = threading.Thread(target=self.processar)
        thread.start()

    def processar(self):
        try:
            remessa_input = self.remessa.get()
            if '.' in remessa_input:
                remessa = int(float(remessa_input))
            else:
                remessa = int(remessa_input)
        except ValueError:
            self.log_callback_completo("Formato de remessa inválido! Digite apenas números.")
            return
        self.progress_bar.set(0.1)
        path_base_sobrepeso = os.path.join(fonte_dir, "Base_sobrepeso_real.xlsx")
        path_base_expedicao = os.path.join(fonte_dir, "expedicao.xlsx")
        path_base_sap = os.path.join(fonte_dir, "base_sap.xlsx")
        path_base_frac = os.path.join(fonte_dir, "FRACAO_1.xlsx")
        path_base_estoque = os.path.join(fonte_dir, "estoqueseparacao.xlsx")
        path_base_peso_exter = os.path.join(fonte_dir,"receb_extern_peso.xlsx")
        df_externo_peso=pd.read_excel(path_base_peso_exter, sheet_name="Sheet1")
        df_estoque_sep = pd.read_excel(path_base_estoque)
        df_frac=pd.read_excel(path_base_frac, sheet_name="FRACAO")
        df_sap = pd.read_excel(path_base_sap, sheet_name="Sheet1")
        df_expedicao = pd.read_excel(path_base_expedicao, sheet_name="dado_exp", 
                                dtype={'REMESSA': str}) 
        df_expedicao['REMESSA'] = df_expedicao['REMESSA'].str.replace('.0', '').astype(str) 
        df_sobrepeso_real = pd.read_excel(path_base_sobrepeso, sheet_name="SOBREPESO")
        df_sobrepeso_real['DataHora'] = pd.to_datetime(df_sobrepeso_real['DataHora'])
        df_base_fisica = pd.read_excel(caminho_base_fisica, sheet_name="BASE FISICA")
        df_fracao = pd.read_excel(path_base_frac, sheet_name="FRACAO")
        try:
            self.log_text.clear()
            self.log_display.configure(state="normal")
            self.log_display.delete("1.0", "end")
            self.log_display.configure(state="disabled")


            file_path = criar_copia_planilha(fonte_dir, "SIMULADOR_BALANÇA_LIMPO_2.xlsx", self.add_log)
            self.log_callback_completo(f"Abrindo planilha Excel em: {file_path}")

            xl = pd.ExcelFile(file_path)
            self.log_callback_completo("Lendo abas do arquivos...")
            with pd.ExcelFile(file_path) as xl:
                df_sku = xl.parse("dado_sku")
            self.log_callback_completo("Abas carregadas com sucesso.")
            self.progress_bar.set(0.3)
            peso_vazio = float(self.peso_vazio.get())
            peso_balança = float(self.peso_balanca.get())
            qtd_paletes = int(self.qtd_paletes.get())
            remessa = int(self.remessa.get())
            df_remessa = obter_dados_remessa(remessa, df_expedicao, log_callback=self.add_log)
            self.log_callback_completo(f"Entradas: Remessa={remessa}, Peso Vazio={peso_vazio}, Paletes={qtd_paletes}")

            self.log_callback_completo("Iniciando cálculo do peso final...")
            self.log_callback_completo(f"Tipo da remessa inserida: {type(remessa)}, Valor: {remessa}")
            self.log_callback_completo(f"Tipo da coluna REMESSA: {df_expedicao['REMESSA'].dtype}")
            self.log_callback_completo(f"Valores únicos de REMESSA: {df_expedicao['REMESSA'].dropna().unique()[:2]}")
            try:
                df_remessa = obter_dados_remessa(remessa, df_expedicao, log_callback=self.add_log)
                remessa_num = int(remessa)
            except Exception as e:
                self.log_callback_completo(f"Erro na conversão: {e}")
                messagebox.showerror("Erro", f"Formato inválido da remessa: {remessa}")
                return

            df_remessa = obter_dados_remessa(remessa, df_expedicao, log_callback=self.add_log)
            if df_remessa.empty:
                disponiveis = sorted(df_expedicao['REMESSA'].dropna().unique().tolist())
                self.log_callback_completo(f"❌ Remessa {remessa_num} não encontrada. Remessas disponíveis: {disponiveis}")
                messagebox.showwarning("Remessa não encontrada", 
                                    f"A remessa {remessa_num} não foi localizada na base.\n"
                                    f"Remessas disponíveis: {disponiveis[-10:]}\n"
                                    f"Verifique se digitou corretamente.")
                return

            resultado = calcular_peso_final(
                remessa, peso_vazio, qtd_paletes,
                df_remessa, df_sku, df_sap, df_sobrepeso_real,
                df_base_fisica, df_frac, df_estoque_sep, df_externo_peso,
                self.log_callback_tecnico
            )
            self.progress_bar.set(0.5)
            if resultado:
                peso_base, sp_total, peso_com_sp, peso_final, media_sp, itens_detalhados = resultado
            
                dados = {
                    'remessa': remessa,
                    'qtd_skus': df_expedicao[df_expedicao['REMESSA'] == remessa]['ITEM'].nunique(),
                    'placa': self.placa.get(),
                    'turno': self.turno.get(),
                    'peso_vazio': peso_vazio,
                    'peso_base': peso_base,
                    'sp_total': sp_total,
                    'peso_com_sp': peso_com_sp,
                    'peso_total_final': peso_final,
                    'media_sp': media_sp,
                    'qtd_paletes': qtd_paletes
                }

                self.log_callback_completo("Chamando preenchimento do formulário via COM...")

                preencher_formulario_com_openpyxl(
                    file_path, dados, itens_detalhados, self.add_log,
                    df_sku, df_remessa, df_fracao
                )
                self.progress_bar.set(0.7)
                self.log_callback_completo("Exportando PDF...")
                pdf_path = exportar_pdf_com_comtypes(file_path, "FORMULARIO", nome_remessa=remessa, log_callback=self.add_log)
                self.log_callback_completo(f"PDF exportado com sucesso: {pdf_path}")
                self.progress_bar.set(0.85)
                self.log_callback_completo("Gerando relatório de divergência em PDF...")
                relatorio_path = gerar_relatorio_diferenca(
                    remessa_num=remessa,
                    peso_final_balança=peso_balança,
                    peso_veiculo_vazio=peso_vazio,
                    df_remessa = obter_dados_remessa(remessa, df_expedicao, log_callback=self.add_log),
                    df_sku=df_sku,
                    peso_estimado_total=peso_com_sp,
                    pasta_excel=fonte_dir,
                    log_callback=self.log_callback_completo
                )
                self.log_callback_completo(f"Relatório adicional salvo em: {relatorio_path}")

                try:
                    self.log_callback_completo("Iniciando impressão do relatório")
                    print_pdf(pdf_path, log_callback=self.log_callback_completo)
                    print_pdf(relatorio_path, log_callback=self.log_callback_completo)
                    self.log_callback_completo("Relatório impresso com sucessos")
                    self.progress_bar.set(0.95)
                except:
                    self.log_callback_completo("Impressão do Relatório com Erro")
                try:
                    enviar_email_com_log_e_pdf(
                        pdf_path,
                        remessa,
                        log_callback=self.log_callback_completo,
                        log_geral=self.log_tecnico
                    )
                    self.log_callback_completo("Envio com sucesso para o email a tratativa")
                    self.progress_bar.set(1)
                    self.add_log("✅ Processamento concluído com sucesso!")
                except:
                    self.log_callback_completo("Erro ao enviar para o email a tratativa")
                try:
                    time.sleep(1)
                    os.remove(file_path)
                    self.log_callback_completo(f"Cópia temporária removida: {file_path}")
                except Exception as e:
                    self.log_callback_completo(f"Erro ao remover a cópia temporária: {e}")
                
                messagebox.showinfo(
                    "Sucesso",
                    f"Formulário exportado: {pdf_path}\n\nRelatório de divergência salvo:\n{relatorio_path}"
                )
            else:
                self.log_callback_completo("Falha no cálculo. Verifique os dados inseridos.")
                messagebox.showwarning("Aviso", "Cálculo não pôde ser realizado.")

        except Exception as e:
            self.log_callback_completo(f"Erro: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao processar: {str(e)}")


if __name__ == "__main__":
    app = App()
    app.mainloop()
