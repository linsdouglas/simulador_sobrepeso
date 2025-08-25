import os
import gc
import time
import shutil
import traceback
import threading
import subprocess
from datetime import datetime
from collections import defaultdict, Counter
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import customtkinter as ctk
from tkinter import messagebox, StringVar
from pathlib import Path
from shutil import copyfile
from openpyxl import load_workbook
import openpyxl
import comtypes.client
import re
import unicodedata


def _find_onedrive_subfolder(subfolder_name: str):
    user_dir = os.environ.get("USERPROFILE", "")
    for nome in os.listdir(user_dir):
        if "DIAS BRANCO" in nome.upper():
            raiz = os.path.join(user_dir, nome)
            if os.path.isdir(raiz) and subfolder_name in os.listdir(raiz):
                return os.path.join(raiz, subfolder_name)
    return None

BASE_DIR_DOCS = _find_onedrive_subfolder("GestÃ£o de Estoque - Documentos")
if not BASE_DIR_DOCS:
    raise FileNotFoundError("Pasta 'GestÃ£o de Estoque - Documentos' nÃ£o encontrada no OneDrive.")

BASE_DIR_AUD = _find_onedrive_subfolder("GestÃ£o de Estoque - GestÃ£o_Auditoria")
if not BASE_DIR_AUD:
    raise FileNotFoundError("Pasta 'GestÃ£o de Estoque - GestÃ£o_Auditoria' nÃ£o encontrada no OneDrive.")



def _norm_remessa_tuple(s):
    if s is None:
        return ("", "")
    ss = str(s).strip().strip("'").strip('"')
    ss = re.sub(r'\.0$', '', ss)              
    dig = re.sub(r'\D', '', ss)               
    if dig == "":
        return ("", "")
    sem_zeros = dig.lstrip('0') or "0"
    return (dig, sem_zeros)

def _match_remessa_series(sr, alvo):
    # forÃ§a Series
    if not isinstance(sr, pd.Series):
        sr = pd.Series(sr)
    # normaliza alvo
    a, b = _norm_remessa_tuple(alvo)
    if a == "" and b == "":
        return pd.Series(False, index=sr.index)

    vals = (sr.astype(str)
              .str.strip()
              .str.replace(r'\.0$', '', regex=True)
              .str.replace(r'\D', '', regex=True))
    vals_no0 = vals.str.lstrip('0')
    vals_no0 = vals_no0.where(vals_no0 != "", "0")

    a_no0 = (a.lstrip('0') or '0')
    b_no0 = (b.lstrip('0') or '0')

    mask = (vals.eq(a) | vals.eq(b) | vals_no0.eq(a_no0) | vals_no0.eq(b_no0))
    # garante mesmo Ã­ndice e dtype bool
    mask = mask.reindex(sr.index).fillna(False).astype(bool)
    return mask



def converter_para_float_seguro(valor):
    if pd.isna(valor):
        return 0.0
    try:
        return float(valor)
    except (ValueError, TypeError):
        return 0.0


EXPECTED_COLS = [
    "ID","LOCAL_EXPEDICAO","REMESSA","COD_ITEM","DESC_ITEM","LOTE",
    "CASEWHENA.EXCLUIDO_POR_LOGINISNULLTHENA.VOLUMEELSE-1*A.VOLUMEEND",
    "UOM","DATA_VALIDADE","COD_RASTREABILIDADE","TIPO_RASTREABILIDADE",
    "CREATED_AT","CRIADO_POR_LOGIN","ATUALIZADO_POR_LOGIN",
    "UPDATED_AT","DELETED_AT","EXCLUIDO_POR_LOGIN"
]

def _split_fix(parts, n):
    """Garante exatamente n campos por linha."""
    if len(parts) > n:
        head = parts[:n-1]
        tail = ";".join(parts[n-1:])
        return head + [tail]
    elif len(parts) < n:
        return parts + [""] * (n - len(parts))
    return parts

def ler_csv_corretamente(csv_path, log=print):
    """LÃª o CSV alinhando header e linhas para EXACTAMENTE 17 colunas."""
    for enc in ("utf-8-sig","utf-8","latin1","cp1252"):
        try:
            with open(csv_path, "r", encoding=enc, errors="replace") as f:
                lines = [ln.rstrip("\r\n") for ln in f if ln.strip()]
            break
        except Exception:
            continue
    else:
        raise RuntimeError("Falha ao ler CSV com encodings testados.")

    # encontra a linha do cabeÃ§alho (a que contÃ©m 'ID' e 'REMESSA')
    header_idx = None
    for i, ln in enumerate(lines):
        ps = ln.split(";")
        if ("ID" in ps) and ("REMESSA" in ps):
            header_idx = i
            break
    if header_idx is None:
        header_idx = 0

    header = _split_fix(lines[header_idx].split(";"), len(EXPECTED_COLS))
    rows = []
    for ln in lines[header_idx+1:]:
        rows.append(_split_fix(ln.split(";"), len(EXPECTED_COLS)))

    df = pd.DataFrame(rows, columns=EXPECTED_COLS, dtype=str)
    # limpeza leve
    df.columns = [c.strip() for c in df.columns]
    for c in ("UPDATED_AT","DELETED_AT","EXCLUIDO_POR_LOGIN"):
        if c in df.columns:
            df[c] = df[c].str.replace(r";+$", "", regex=True).str.strip()
    return df

def carregar_base_expedicao_csv(base_dir: str, log=print):
    path_csv = os.path.join(base_dir, "rastreabilidade.csv")
    if not os.path.exists(path_csv):
        raise FileNotFoundError(f"'rastreabilidade.csv' nÃ£o encontrado em: {path_csv}")

    df_raw = ler_csv_corretamente(path_csv, log=log)

    col_remessa = "REMESSA"
    col_item = "COD_ITEM"
    col_chave = "COD_RASTREABILIDADE"
    col_vol = "CASEWHENA.EXCLUIDO_POR_LOGINISNULLTHENA.VOLUMEELSE-1*A.VOLUMEEND"

    df = pd.DataFrame({
        "REMESSA": df_raw[col_remessa].astype(str).str.strip(),
        "ITEM": df_raw[col_item].astype(str).str.strip(),
        "COD_RASTREABILIDADE": df_raw[col_chave].astype(str).str.strip(),
        "QUANTIDADE": pd.to_numeric(df_raw[col_vol].astype(str).str.replace(",", ".", regex=False),
                                    errors="coerce").fillna(0.0)
    })

    df = df[(df["REMESSA"] != "") & (df["ITEM"] != "") & (df["COD_RASTREABILIDADE"] != "")]
    df = df.reset_index(drop=True)

    log(f"[loader] CSV carregado OK: {df.shape[0]} linhas")
    log(f"[loader] exemplos de remessas: {df['REMESSA'].unique()[:5]}")
    return df

def salvar_em_base_auxiliar(df_remessa, remessa, log_callback, fonte_dir):
    caminho_aux = os.path.join(fonte_dir, "expedicao_edicoes.xlsx")
    try:
        if os.path.exists(caminho_aux):
            try:
                with pd.ExcelFile(caminho_aux) as xls:
                    if "dado_exp" in xls.sheet_names:
                        df_existente = pd.read_excel(xls, sheet_name="dado_exp")
                        df_existente = df_existente[
                            df_existente["REMESSA"].astype(str).str.replace(r"\.0$", "", regex=True) != str(remessa)
                        ]
                    else:
                        df_existente = pd.DataFrame()
            except Exception as e:
                log_callback(f"Erro ao ler arquivo existente, criando novo: {str(e)}")
                df_existente = pd.DataFrame()
        else:
            df_existente = pd.DataFrame()

        df_remessa = df_remessa.copy()
        if "CHAVE_PALETE" not in df_remessa.columns and "COD_RASTREABILIDADE" in df_remessa.columns:
            df_remessa["CHAVE_PALETE"] = df_remessa["COD_RASTREABILIDADE"]
        if "COD_RASTREABILIDADE" not in df_remessa.columns and "CHAVE_PALETE" in df_remessa.columns:
            df_remessa["COD_RASTREABILIDADE"] = df_remessa["CHAVE_PALETE"]

        df_remessa["REMESSA"] = df_remessa["REMESSA"].astype(str).str.replace(r"\.0$", "", regex=True)
        df_remessa = df_remessa.dropna(subset=["ITEM"])
        df_remessa = df_remessa[df_remessa["ITEM"] != ""]

        df_remessa = df_remessa.drop_duplicates(subset=["REMESSA", "ITEM", "CHAVE_PALETE"], keep="last")

        df_atualizado = pd.concat([df_existente, df_remessa], ignore_index=True)
        df_atualizado = df_atualizado.sort_values(by=["REMESSA", "ITEM"])

        with pd.ExcelWriter(caminho_aux, engine="openpyxl") as writer:
            df_atualizado.to_excel(writer, sheet_name="dado_exp", index=False)

        log_callback(f"Remessa {remessa} salva na base auxiliar. Total de itens: {len(df_remessa)}")
        return df_atualizado

    except Exception as e:
        log_callback(f"[ERRO AO SALVAR NA BASE AUXILIAR]: {str(e)}")
        raise


def carregar_base_auxiliar(fonte_dir):
    caminho_aux = os.path.join(fonte_dir, "expedicao_edicoes.xlsx")
    try:
        if os.path.exists(caminho_aux):
            with pd.ExcelFile(caminho_aux) as xls:
                if "dado_exp" in xls.sheet_names:
                    return pd.read_excel(xls, sheet_name="dado_exp")
        return pd.DataFrame()
    except Exception as e:
        print(f"Erro ao carregar base auxiliar: {e}")
        return pd.DataFrame()

def remover_remessa_base_auxiliar(remessa, fonte_dir, log_callback):
    try:
        caminho_aux = os.path.join(fonte_dir, "expedicao_edicoes.xlsx")
        if not os.path.exists(caminho_aux):
            return

        with pd.ExcelFile(caminho_aux) as xls:
            if "dado_exp" in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name="dado_exp")
                df = df[df["REMESSA"].astype(str) != str(remessa)]

                with pd.ExcelWriter(caminho_aux, engine="openpyxl") as writer:
                    df.to_excel(writer, sheet_name="dado_exp", index=False)

                log_callback(f"Remessa {remessa} removida da base auxiliar")
    except Exception as e:
        log_callback(f"Erro ao remover remessa da base auxiliar: {e}")


def obter_dados_remessa(remessa, df_expedicao, log_callback):
    try:
        caminho_aux = os.path.join(BASE_DIR_DOCS, "expedicao_edicoes.xlsx")
        if os.path.exists(caminho_aux):
            df_aux = pd.read_excel(caminho_aux, sheet_name="dado_exp")
            if "CHAVE_PALETE" not in df_aux.columns and "COD_RASTREABILIDADE" in df_aux.columns:
                df_aux["CHAVE_PALETE"] = df_aux["COD_RASTREABILIDADE"]

            sr = df_aux['REMESSA'] if 'REMESSA' in df_aux.columns else pd.Series("", index=df_aux.index)
            mask_aux = _match_remessa_series(sr, remessa)
            df_filtrado_aux = df_aux.loc[mask_aux].copy()
            if not df_filtrado_aux.empty:
                log_callback(f"Remessa {remessa} encontrada na base auxiliar (edicoes)")
                return df_filtrado_aux

        sr_rem = df_expedicao['REMESSA'] if 'REMESSA' in df_expedicao.columns else pd.Series([], dtype=str, index=df_expedicao.index)
        mask_ori = _match_remessa_series(sr_rem, remessa)
        if not isinstance(mask_ori, pd.Series):
            raise RuntimeError("match_remessa retornou objeto nÃ£o-Series")

        df_filtrado = df_expedicao.loc[mask_ori].copy()

        if not df_filtrado.empty:
            if "CHAVE_PALETE" not in df_filtrado.columns and "COD_RASTREABILIDADE" in df_filtrado.columns:
                df_filtrado["CHAVE_PALETE"] = df_filtrado["COD_RASTREABILIDADE"]
            log_callback(f"Remessa {remessa} encontrada na base original")
            return df_filtrado

        log_callback(f"Remessa {remessa} nÃ£o encontrada em nenhuma base")
        return pd.DataFrame()

    except Exception as e:
        log_callback(f"Erro ao buscar remessa {remessa}: {str(e)}")
        return pd.DataFrame()


def criar_copia_planilha(fonte_dir, nome_arquivo, log_callback):
    try:
        origem = os.path.join(fonte_dir, nome_arquivo)
        destino_dir = os.path.join(os.environ["USERPROFILE"], "Downloads")
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        nome_copia = f"copia_temp_{timestamp}_{nome_arquivo}"
        destino = os.path.join(destino_dir, nome_copia)
        copyfile(origem, destino)
        log_callback(f"CÃ³pia criada com sucesso: {destino}")
        return destino
    except Exception as e:
        log_callback(f"Erro ao criar cÃ³pia da planilha: {e}")
        raise

def enviar_email_com_log_e_pdf(caminho_pdf, remessa, log_callback=None, log_geral=None):
    import yagmail
    email_remetente = "mdiasbrancoautomacao@gmail.com"
    token = "secwygmzlibyxhhh"
    email_destino = "douglas.lins2@mdiasbranco.com.br"

    try:
        corpo_email = "\n".join(log_geral or ["(Sem logs tÃ©cnicos disponÃ­veis)"])
        yag = yagmail.SMTP(user=email_remetente, password=token)
        assunto = f"ðŸ“¦ Simulador Sobrepeso - Remessa {remessa}"
        yag.send(to=email_destino, subject=assunto, contents=corpo_email, attachments=[caminho_pdf])
        if log_callback:
            log_callback(f"E-mail enviado com sucesso para {email_destino}")
    except Exception as e:
        if log_callback:
            log_callback(f"Erro ao enviar e-mail: {e}")

def print_pdf(file_path, impressora="VITLOG01A01", sumatra_path="C:\\Program Files\\SumatraPDF\\SumatraPDF.exe", log_callback=None):
    args = [sumatra_path, "-print-to", impressora, "-silent", file_path]
    try:
        result = subprocess.run(args, check=True, capture_output=True, text=True)
        if log_callback:
            log_callback(f"Arquivo impresso com sucesso: {file_path}")
            if result.stdout:
                log_callback(f"stdout: {result.stdout}")
            if result.stderr:
                log_callback(f"stderr: {result.stderr}")
    except subprocess.CalledProcessError as e:
        if log_callback:
            log_callback(f"Erro ao imprimir {file_path}: {e}")
            log_callback(f"Output: {e.output}")
            log_callback(f"Stderr: {e.stderr}")

def exportar_pdf_com_comtypes(path_xlsx, aba_nome="FORMULARIO", nome_remessa="REMESSA", log_callback=None):
    try:
        if log_callback:
            log_callback("Iniciando exportaÃ§Ã£o via comtypes...")

        comtypes.CoInitialize()
        excel = comtypes.client.CreateObject("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(path_xlsx)
        if aba_nome not in [sheet.Name for sheet in wb.Sheets]:
            raise Exception(f"Aba '{aba_nome}' nÃ£o encontrada.")
        ws = wb.Worksheets(aba_nome)

        pdf_dir = os.path.join(BASE_DIR_DOCS, "RelatÃ³rio_Saida")
        os.makedirs(pdf_dir, exist_ok=True)
        pdf_path = os.path.join(pdf_dir, f"SOBREPESOSIMULADO - {nome_remessa}.pdf")

        if log_callback:
            log_callback(f"Tentando exportar para: {pdf_path}")
        ws.ExportAsFixedFormat(Type=0, Filename=pdf_path)

        wb.Close(SaveChanges=False)
        excel.Quit()
        del ws
        del wb
        del excel
        gc.collect()
        comtypes.CoUninitialize()
        time.sleep(2)

        if log_callback:
            log_callback(f"PDF exportado com sucesso: {pdf_path}")
        return pdf_path
    except Exception as e:
        if log_callback:
            log_callback(f"Erro ao exportar PDF: {e}")
        raise

def _norm_colname(name: str) -> str:
    """minÃºsculas, sem acento, sÃ³ letras/dÃ­gitos/_"""
    s = unicodedata.normalize("NFKD", str(name))
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^a-zA-Z0-9]+", "_", s).strip("_").lower()
    return s

def _pick_col_flex(df: pd.DataFrame, candidatos) -> str | None:
    norm_map = {_norm_colname(c): c for c in df.columns}
    for cand in candidatos:
        key = _norm_colname(cand)
        if key in norm_map:
            return norm_map[key]
    want = {_norm_colname(c) for c in candidatos}
    for col in df.columns:
        n = _norm_colname(col)
        if any(w in n for w in want):
            return col
    return None
def carregar_base_fisica(base_dir_docs: str, log=print):
    """
    LÃª SIMULADOR_BALANÃ‡A_LIMPO_2.xlsx e devolve (df_base_fisica, df_base_familia),
    com logs de colunas e amostras para depuraÃ§Ã£o.
    """
    path = os.path.join(base_dir_docs, "SIMULADOR_BALANÃ‡A_LIMPO_2.xlsx")
    if not os.path.exists(path):
        raise FileNotFoundError(f"Arquivo nÃ£o encontrado: {path}")

    try:
        df_bf  = pd.read_excel(path, sheet_name="BASE FISICA")
        df_fam = pd.read_excel(path, sheet_name="BASE_FAMILIA")
    except Exception as e:
        log(f"[BASE_FISICA][erro] Falha ao ler planilha: {e!r}")
        raise
    log(f"[BASE_FISICA] caminho: {path}")
    log(f"[BASE_FISICA] linhas: {len(df_bf)} | colunas: {list(df_bf.columns)}")

    col_sku = _pick_col_flex(df_bf, ["CÃ“DIGO PRODUTO","CODIGO_PRODUTO","COD_PRODUTO","CÃ“D_PRODUTO"])
    col_sp  = _pick_col_flex(df_bf, ["SOBRE PESO","SOBREPESO","SOBRE_PESO","SOBRE PESO (%)","SOBREPESO_%","SOBRE_PESO_FIXO"])
    log(f"[BASE_FISICA] coluna SKU detectada: {col_sku!r} | coluna SP detectada: {col_sp!r}")

    try:
        prev = (df_bf[[c for c in [col_sku, col_sp] if c]]
                .head(5)
                .to_dict(orient="records"))
        log(f"[BASE_FISICA] amostra: {prev}")
    except Exception:
        pass

    if col_sku:
        df_bf[col_sku] = df_bf[col_sku].astype(str).str.strip()
    if col_sp:
        df_bf[col_sp]  = df_bf[col_sp].astype(str).str.strip()

    return df_bf, df_fam


def _norm_sku(s) -> str:
    return re.sub(r"\D", "", str(s).strip())

def _to_frac(x) -> float:
    if pd.isna(x):
        return 0.0
    s = str(x).strip().replace(",", ".").replace("%", "")
    try:
        v = float(s)
    except ValueError:
        return 0.0
    return v/100.0 if v > 1.0 else v


def calculo_sobrepeso_fixo(sku, df_base_fisica, df_sku, peso_base_liq, log_callback):
    try:
        sku_norm = _norm_sku(sku)
        pb = converter_para_float_seguro(peso_base_liq) or 0.0

        # UtilitÃ¡rios de coluna
        cols_sku = ["COD_PRODUTO","CODIGO_PRODUTO","CÃ“DIGO PRODUTO","CÃ“D_PRODUTO","COD_PROD"]
        cols_sp_kg = ["SOBRE PESO","SOBREPESO","SOBRE_PESO","SOBRE PESO (KG)","SOBREPESO_KG","SOBRE_PESO_FIXO"]
        cols_sp_perc = ["DIF (%)","DIF%","DIF_PERC","SOBRE PESO (%)","SOBREPESO_%"]
        cols_media = ["MEDIA","MÃ‰DIA"]
        cols_peso_sap = ["PESO SAP","PESO_SAP","PESO_SAP_KG","PESO SAP (KG)"]

        def _match_df(df):
            if df is None or df.empty:
                return None
            col_sku = _pick_col_flex(df, cols_sku)
            if not col_sku:
                return None
            serie_cod_norm = df[col_sku].astype(str).str.replace(r"\D", "", regex=True)
            hit = df.loc[serie_cod_norm == sku_norm]
            return hit if not hit.empty else None

        def _read_sp_kg(row):
            col = _pick_col_flex(row, cols_sp_kg)
            if not col:
                return None
            v = converter_para_float_seguro(row.iloc[0][col])
            return v if v is not None else None  # jÃ¡ em kg

        def _read_sp_perc(row):
            col = _pick_col_flex(row, cols_sp_perc)
            if not col:
                return None
            raw = row.iloc[0][col]
            # aceita "6", "6%", "0,06", "0.06"
            frac = _to_frac(raw)  # sua funÃ§Ã£o: 6 -> 0.06 ; "0,06" -> 0.06 ; "6%" -> 0.06
            return frac if frac is not None else None

        def _read_media_peso_sap(row):
            cmed = _pick_col_flex(row, cols_media)
            csap = _pick_col_flex(row, cols_peso_sap)
            if not (cmed and csap):
                return None, None
            med = converter_para_float_seguro(row.iloc[0][cmed])
            sap = converter_para_float_seguro(row.iloc[0][csap])
            return med, sap

        # Busca primeiro em df_sku, depois em df_base_fisica
        for origem, df in (("dado_sku", df_sku), ("base_fisica", df_base_fisica)):
            hit = _match_df(df)
            log_callback(f"[fixo/{origem}] match={'1' if hit is not None else '0'}")
            if hit is None:
                continue

            # 1) Tenta SOBRE PESO (kg)
            spkg = _read_sp_kg(hit)
            if spkg is not None:
                perc = 0.0
                # se tiver DIF(%), calcula tambÃ©m o percentual
                p = _read_sp_perc(hit)
                if p is not None:
                    perc = float(p)
                elif pb and pb > 0:
                    # fallback: tenta usar PESO SAP para inferir % com mais precisÃ£o
                    med, sap = _read_media_peso_sap(hit)
                    base_ref = sap if sap and sap > 0 else pb
                    perc = float(max(0.0, (spkg / base_ref))) if base_ref else 0.0
                log_callback(f"[fixo/{origem}] SKU {sku} â†’ SOBRE PESO (kg)={spkg:.3f} | percâ‰ˆ{perc:.4%}")
                return float(perc), float(spkg)

            # 2) Tenta DIF (%) como fraÃ§Ã£o
            p = _read_sp_perc(hit)
            if p is not None:
                perc = float(p)
                ajuste = float((pb or 0.0) * perc)
                # se houver MÃ‰DIA e PESO SAP, preferir ajuste a partir deles (mais exato que pb)
                med, sap = _read_media_peso_sap(hit)
                if med is not None and sap is not None:
                    ajuste = max(0.0, float(med - sap))
                    # recalcula % real com base no SAP
                    if sap and sap > 0:
                        perc = ajuste / float(sap)
                log_callback(f"[fixo/{origem}] SKU {sku} â†’ perc={perc:.4%} | ajusteâ‰ˆ{ajuste:.3f} kg (pb={pb:.3f})")
                return float(perc), float(ajuste)

            # 3) Tenta derivar de MÃ‰DIA e PESO SAP
            med, sap = _read_media_peso_sap(hit)
            if med is not None and sap is not None:
                ajuste = max(0.0, float(med - sap))
                perc = (ajuste / float(sap)) if sap and sap > 0 else 0.0
                log_callback(f"[fixo/{origem}] SKU {sku} â†’ med={med} sap={sap} -> ajuste={ajuste:.3f} kg | perc={perc:.4%}")
                return float(perc), float(ajuste)

        log_callback(f"[fixo] SKU {sku} sem cadastro (dado_sku/base_fisica).")
        return 0.0, 0.0

    except Exception as e:
        log_callback(f"[fixo][erro] SKU {sku}: {type(e).__name__}: {e}")
        return 0.0, 0.0

def _norm_str(x):
    if pd.isna(x): return ""
    return str(x).strip()

def _norm_chave(x):
    return re.sub(r"\s+", "", _norm_str(x))

def _safe_hour(v):
    """Extrai hora como inteiro de 0..23 a partir de datetime, time ou string."""
    if pd.isna(v): return None
    try:
        h = getattr(v, "hour", None)
        if h is not None:
            return int(h)
    except Exception:
        pass
    s = str(v).strip()
    m = re.search(r"\b(\d{1,2})\s*:\s*\d{1,2}(?::\d{1,2})?\b", s)
    if m:
        h = int(m.group(1))
        return max(0, min(23, h))
    m = re.search(r"\b(\d{1,2})\b", s)
    if m:
        h = int(m.group(1))
        return max(0, min(23, h))
    return None

def _coerce_dt(dt):
    if isinstance(dt, pd.Timestamp):
        return dt
    try:
        return pd.to_datetime(dt, dayfirst=True, errors="coerce")
    except Exception:
        return pd.NaT
def processar_sobrepeso(chave_pallet, sku, peso_base_liq, 
                        df_sap, df_sobrepeso_real, df_base_fisica, df_sku, log_callback):
    peso_base_liq = float(peso_base_liq or 0)
    sp = 0.0
    origem_sp = "nÃ£o encontrado"
    ajuste_sp = 0.0

    chave_norm = _norm_chave(chave_pallet)
    log_callback(f"[item] sku={sku} | chave={chave_norm} | peso_liqâ‰ˆ{peso_base_liq:.2f} kg")

    try:
        sap_col = None
        for c in ["Chave Pallet", "CHAVE_PALETE", "CHAVE_PALLET", "CHAVE_PALETE "]:
            if c in df_sap.columns:
                sap_col = c
                break
        if sap_col is None:
            log_callback("[real] df_sap sem coluna de chave ('Chave Pallet' / 'CHAVE_PALETE').")
        else:
            col_norm = df_sap[sap_col].astype(str).str.replace(r"\s+", "", regex=True)
            mask = col_norm.eq(chave_norm)

            found = int(mask.sum())
            log_callback(f"[real] Busca chave no SAP: col='{sap_col}' | encontrados={found}")
            if found == 0:
                sug = df_sap.loc[col_norm.str.contains(chave_norm[:6], na=False)].head(2).get(sap_col, [])
                log_callback(f"[real] Nenhuma linha com a chave exata. SugestÃµes (mesmos 6 primeiros): {list(sug)}")
            else:
                pallet_info = df_sap.loc[mask].iloc[0]

                cand_lote = None
                for c in ["Lote", "LOTE", "Lote de produÃ§Ã£o", "LOTE_PROD"]:
                    if c in df_sap.columns:
                        cand_lote = pallet_info.get(c)
                        if pd.notna(cand_lote):
                            break
                lote = _norm_str(cand_lote)
                linha_coluna = "L" + lote[-3:] if lote else ""
                if linha_coluna in ["LB06", "LB07"]:
                    linha_coluna = "LB06/07"
                log_callback(f"[real] Lote='{lote}' â†’ linha_coluna='{linha_coluna}'")

                cand_data = None
                for c in ["Data de produÃ§Ã£o", "DATA_PRODUCAO", "Data de FabricaÃ§Ã£o", "DATA FABRICACAO"]:
                    if c in df_sap.columns:
                        cand_data = pallet_info.get(c)
                        if pd.notna(cand_data):
                            break
                dt_prod = _coerce_dt(cand_data)
                h_ini = None
                h_fim = None
                for c in ["Hora de criaÃ§Ã£o", "HORA_CRIACAO", "Hora CriaÃ§Ã£o"]:
                    if c in df_sap.columns:
                        h_ini = _safe_hour(pallet_info.get(c))
                        if h_ini is not None:
                            break
                for c in ["Hora de modificaÃ§Ã£o", "HORA_MODIFICACAO", "Hora ModificaÃ§Ã£o"]:
                    if c in df_sap.columns:
                        h_fim = _safe_hour(pallet_info.get(c))
                        if h_fim is not None:
                            break
                if h_ini is None and h_fim is not None:
                    h_ini = h_fim
                if h_fim is None and h_ini is not None:
                    h_fim = h_ini
                if dt_prod is pd.NaT:
                    log_callback(f"[real][warn] Data de produÃ§Ã£o invÃ¡lida para a chave. Valor bruto='{cand_data}'")
                if h_ini is None: h_ini = 0
                if h_fim is None: h_fim = h_ini

                if pd.isna(dt_prod):
                    dt_ini = pd.NaT
                    dt_fim = pd.NaT
                else:
                    dt_ini = pd.Timestamp(dt_prod.date()) + pd.Timedelta(hours=int(h_ini))
                    dt_fim = pd.Timestamp(dt_prod.date()) + pd.Timedelta(hours=int(h_fim), minutes=59, seconds=59)

                log_callback(f"[real] janela: dt_ini='{dt_ini}' dt_fim='{dt_fim}' (h_ini={h_ini}, h_fim={h_fim})")

                if "DataHora" not in df_sobrepeso_real.columns:
                    log_callback("[real][erro] df_sobrepeso_real nÃ£o possui coluna 'DataHora'.")
                elif not linha_coluna:
                    log_callback("[real][erro] NÃ£o foi possÃ­vel derivar 'linha_coluna' a partir do lote.")
                else:
                    if not pd.api.types.is_datetime64_any_dtype(df_sobrepeso_real["DataHora"]):
                        try:
                            df_sobrepeso_real = df_sobrepeso_real.copy()
                            df_sobrepeso_real["DataHora"] = pd.to_datetime(
                                df_sobrepeso_real["DataHora"], dayfirst=True, errors="coerce"
                            )
                            log_callback("[real] Convertemos 'DataHora' para datetime.")
                        except Exception as e:
                            log_callback(f"[real][erro] Falha ao converter 'DataHora': {e}")

                    if pd.isna(dt_ini) or pd.isna(dt_fim):
                        df_sp_filtro = df_sobrepeso_real.iloc[0:0]  # vazio
                        log_callback("[real][warn] dt_ini/dt_fim invÃ¡lidos â†’ filtro vazio.")
                    else:
                        df_sp_filtro = df_sobrepeso_real[
                            (df_sobrepeso_real["DataHora"] >= dt_ini) &
                            (df_sobrepeso_real["DataHora"] <= dt_fim)
                        ]

                    log_callback(f"[real] linhas na janela: {len(df_sp_filtro)} "
                                 f"(DataHora min={df_sobrepeso_real['DataHora'].min()}, max={df_sobrepeso_real['DataHora'].max()})")

                    col_ok = linha_coluna in df_sp_filtro.columns
                    if not col_ok:
                        log_callback(f"[real][erro] Coluna '{linha_coluna}' NÃƒO existe em df_sobrepeso_real. "
                                     f"Cols disp: {list(df_sp_filtro.columns)}")
                    else:
                        sp_vals = pd.to_numeric(df_sp_filtro[linha_coluna], errors="coerce")
                        media_raw = float(sp_vals.mean()) if len(sp_vals) else 0.0
                        media_sp = media_raw/100.0 if media_raw > 1.0 else media_raw
                        log_callback(f"[real] mÃ©dia({linha_coluna}) bruta={media_raw:.6f} â†’ usada={media_sp:.6f}")

                        if media_sp > 0 and peso_base_liq > 0:
                            sp = media_sp
                            origem_sp = "real"
                            ajuste_sp = peso_base_liq * sp
                            log_callback(f"[real][OK] sp={sp:.4f} ajusteâ‰ˆ{ajuste_sp:.2f} kg")
                        else:
                            log_callback("[real] mÃ©dia<=0 ou peso_base_liq<=0 â†’ fallback fixo")

    except Exception as e:
        log_callback(f"[real][exceÃ§Ã£o] sku={sku} chave={chave_norm} -> {e}")

    if sp == 0:
        sp_frac, ajuste_fixo_kg = calculo_sobrepeso_fixo(sku, df_base_fisica, df_sku, peso_base_liq, log_callback)

        if (sp_frac is None or sp_frac == 0) and ajuste_fixo_kg and peso_base_liq > 0:
            sp_frac = float(ajuste_fixo_kg) / float(peso_base_liq)

        if (sp_frac or 0) > 0 or (ajuste_fixo_kg or 0) > 0:
            sp = float(sp_frac or 0.0)
            origem_sp = "fixo"
            ajuste_sp = float(ajuste_fixo_kg or 0.0)
            if sp == 0 and peso_base_liq > 0:
                sp = ajuste_sp / float(peso_base_liq)
            if ajuste_sp == 0 and peso_base_liq > 0 and sp > 0:
                ajuste_sp = float(peso_base_liq) * float(sp)

            log_callback(f"[fixo] adotado sp={sp:.4f} ajusteâ‰ˆ{ajuste_sp:.2f} kg")
        else:
            log_callback("[sp] nÃ£o encontrado (real/fixo).")

    if sp <= 0 and origem_sp != "fixo":
        sp, origem_sp, ajuste_sp = 0.0, "nÃ£o encontrado", 0.0

    return float(sp), origem_sp, float(ajuste_sp)


def calcular_peso_final(
    remessa_num,
    peso_veiculo_vazio,
    qtd_paletes,
    df_remessa,
    df_sku,
    df_sap,
    df_sobrepeso_real,
    df_base_fisica,
    log_callback
):
    peso_veiculo_vazio = converter_para_float_seguro(peso_veiculo_vazio)
    qtd_paletes = converter_para_float_seguro(qtd_paletes)

    try:
        remessa_num = int(remessa_num)
    except ValueError:
        log_callback("Remessa invÃ¡lida.")
        return None

    df_remessa = df_remessa.copy()
    df_remessa["QUANTIDADE"] = df_remessa["QUANTIDADE"].apply(converter_para_float_seguro)
    df_remessa = df_remessa.drop_duplicates(subset=["ITEM", "QUANTIDADE", "CHAVE_PALETE"], keep="last")
    log_callback(f"[remessa] {remessa_num} | linhas apÃ³s dedup: {len(df_remessa)}")

    peso_base_total_bruto = 0.0
    peso_base_total_liq = 0.0
    sp_total = 0.0
    itens_detalhados = []

    for i, (_, row) in enumerate(df_remessa.iterrows(), start=1):
        sku = str(row["ITEM"]).strip()
        qtd = converter_para_float_seguro(row["QUANTIDADE"])
        chave = str(row["CHAVE_PALETE"]).strip()

        if not sku or qtd <= 0:
            log_callback(f"[linha {i}] ignorada (sku vazio ou qtd<=0): sku='{sku}' qtd={qtd}")
            continue

        sku_norm = _norm_sku(str(sku))
        log_callback(f"[fixo][procura] SKU alvo: {sku} (norm={sku_norm})")
        col_cod = _pick_col_flex(
            df_sku,
            ["COD_PRODUTO","CODIGO_PRODUTO","CÃ“DIGO PRODUTO","CÃ“D_PRODUTO","COD_PROD"]
        )
        if not col_cod:
            log_callback(f"[linha {i}] Coluna de cÃ³digo de produto nÃ£o encontrada em df_sku.")
            continue

        serie_cod_norm = (
            df_sku[col_cod]
            .astype(str)
            .str.replace(r"\D", "", regex=True)
        )

        sku_norm = _norm_sku(sku)

        mask = serie_cod_norm == sku_norm
        df_sku_filtrado = df_sku[mask]

        log_callback(f"[linha {i}] SKU alvo: {sku} (norm: {sku_norm}) "
                    f"â€“ encontrados: {mask.sum()} linhas em df_sku.")

        if not df_sku_filtrado.empty:
            exemplo = df_sku_filtrado.head(1).to_dict(orient="records")[0]
            log_callback(f"[linha {i}] Exemplo da linha encontrada em df_sku: {exemplo}")
        else:
            exemplos_cod = serie_cod_norm.head(5).tolist()
            log_callback(f"[linha {i}] Nenhum match. Exemplos normalizados em df_sku: {exemplos_cod}")

        p_bruto = converter_para_float_seguro(df_sku_filtrado.iloc[0]["QTDE_PESO_BRU"])
        p_liq   = converter_para_float_seguro(df_sku_filtrado.iloc[0]["QTDE_PESO_LIQ"])

        peso_bruto = p_bruto * qtd if p_bruto > 0 else 0.0
        peso_liq   = p_liq   * qtd if p_liq   > 0 else 0.0

        log_callback(f"[linha {i}] sku={sku} qtd={qtd} chave={chave} | unit(bruto={p_bruto}, liq={p_liq}) | base(brutoâ‰ˆ{peso_bruto:.2f}, liqâ‰ˆ{peso_liq:.2f})")

        sp, origem_sp, ajuste_sp = processar_sobrepeso(
            chave, sku, peso_liq, df_sap, df_sobrepeso_real, df_base_fisica, df_sku, log_callback
        )

        peso_base_total_bruto += peso_bruto
        peso_base_total_liq   += peso_liq
        sp_total += ajuste_sp

        log_callback(f"[linha {i}] SP={sp:.4f} ({origem_sp}) | ajusteâ‰ˆ{ajuste_sp:.2f} kg | acumulados: base_brutoâ‰ˆ{peso_base_total_bruto:.2f} kg, SPâ‰ˆ{sp_total:.2f} kg")

        itens_detalhados.append({
            "sku": sku,
            "chave_pallet": chave,
            "sp": round(sp, 4),
            "ajuste_sp": round(ajuste_sp, 2),
            "origem": origem_sp
        })

    peso_com_sobrepeso = peso_base_total_bruto + sp_total
    log_callback(f"[total] base_brutoâ‰ˆ{peso_base_total_bruto:.2f} kg | SPâ‰ˆ{sp_total:.2f} kg | com_SPâ‰ˆ{peso_com_sobrepeso:.2f} kg")

    peso_total_com_paletes = peso_com_sobrepeso + (qtd_paletes * 22.0) + peso_veiculo_vazio
    log_callback(f"[total] +paletes({qtd_paletes}Ã—22) +veÃ­culo({peso_veiculo_vazio}) => finalâ‰ˆ{peso_total_com_paletes:.2f} kg")

    media_sp_geral = (sum(item["sp"] for item in itens_detalhados) / len(itens_detalhados)) if itens_detalhados else 0.0
    log_callback(f"[total] mÃ©dia(sp)={media_sp_geral:.4f} em {len(itens_detalhados)} itens")

    return (
        peso_base_total_bruto,
        sp_total,
        peso_com_sobrepeso,
        peso_total_com_paletes,
        media_sp_geral,
        itens_detalhados
    )

def _to_str(x):
    if pd.isna(x):
        return ""
    s = str(x).strip()
    # remove sufixo .0 tÃ­pico de float vindo do Excel
    s = re.sub(r"\.0$", "", s)
    return s

def _norm_digits(x: str) -> str:
    """MantÃ©m apenas dÃ­gitos. Ãštil quando SKU vem '001234.0' vs '1234'."""
    return re.sub(r"\D", "", _to_str(x))

def _pick_family_column(df_base_familia):
    """
    Tenta encontrar a coluna de famÃ­lia e a de cÃ³digo em df_base_familia.
    Retorna (col_cod, col_fam).
    """
    cols = {c.upper().strip(): c for c in df_base_familia.columns}

    # candidatos para cÃ³digo do item / SKU
    cand_cod = ["CÃ“D", "COD", "CÃ“DIGO", "CODIGO", "SKU", "ITEM", "CÃ“D_ITEM", "COD_ITEM"]
    col_cod = None
    for k in cand_cod:
        if k in cols:
            col_cod = cols[k]; break
    if col_cod is None:
        raise ValueError("NÃ£o encontrei coluna de cÃ³digo/SKU em df_base_familia.")

    # candidatos para famÃ­lia
    cand_fam = ["FAMILIA 2", "FAMÃLIA 2", "FAMILIA2", "FAMILIA", "FAMÃLIA", "FAMILIA_2"]
    col_fam = None
    for k in cand_fam:
        if k in cols:
            col_fam = cols[k]; break
    if col_fam is None:
        raise ValueError("NÃ£o encontrei coluna de famÃ­lia em df_base_familia (ex.: 'FAMILIA 2').")

    return col_cod, col_fam

def _ensure_df_sobrepeso_index(df_sobrepeso_tabela):
    """
    Garante que o df_sobrepeso_tabela pode ser consultado por famÃ­lia.
    Se o Ã­ndice nÃ£o for de famÃ­lia, tenta usar uma coluna 'FAMILIA' (variaÃ§Ãµes).
    Retorna (df, fonte_do_indice: 'index'|'col').
    """
    if df_sobrepeso_tabela.index.name is not None:
        # jÃ¡ tem Ã­ndice nomeado â€” mantemos, mas padronizamos para uppercase/strip no match
        return df_sobrepeso_tabela, "index"
    # tenta achar coluna de famÃ­lia
    fam_cols = [c for c in df_sobrepeso_tabela.columns if str(c).strip().upper() in ("FAMILIA", "FAMÃLIA", "FAMILY", "GRUPO")]
    if fam_cols:
        col = fam_cols[0]
        df = df_sobrepeso_tabela.copy()
        df["_FAM_TMP_KEY_"] = df[col].astype(str).str.upper().str.strip()
        df = df.set_index("_FAM_TMP_KEY_", drop=True)
        return df, "col"
    # se nada encontrado, usamos o Ã­ndice como estÃ¡ (e logamos depois)
    return df_sobrepeso_tabela, "index"

def _coerce_float(x, default=0.0):
    try:
        if isinstance(x, str):
            x = x.replace(",", ".")
        return float(x)
    except Exception:
        return default

# --- FunÃ§Ã£o principal com logs detalhados ---

def calcular_limites_sobrepeso_por_quantidade(
    dados,
    itens_detalhados,
    df_base_familia,
    df_sobrepeso_tabela,
    df_sku,
    df_remessa,
    df_fracao,
    log_callback
):
    total_quantidade = 0
    quantidade_com_sp_real = 0
    ponderador_pos = 0.0
    ponderador_neg = 0.0
    familia_detectada = "MIX"

    # 1) NormalizaÃ§Ã£o e diagnÃ³stico da base de famÃ­lia
    try:
        col_cod, col_fam = _pick_family_column(df_base_familia)
    except Exception as e:
        log_callback(f"[ERRO] Base famÃ­lia sem colunas esperadas: {e}. ForÃ§ando MIX.")
        # fallback hard
        col_cod, col_fam = None, None

    if col_cod and col_fam:
        df_fam = df_base_familia[[col_cod, col_fam]].copy()
        df_fam["_COD_NORM_"] = df_fam[col_cod].map(_norm_digits)
        df_fam["_FAM_NORM_"] = df_fam[col_fam].astype(str).str.upper().str.strip()
        # remove vazios Ã³bvios
        df_fam = df_fam[df_fam["_COD_NORM_"] != ""]
        log_callback(f"[DEBUG] Base famÃ­lia normalizada: linhas={len(df_fam)}, col_cod='{col_cod}', col_fam='{col_fam}'.")
    else:
        df_fam = pd.DataFrame(columns=["_COD_NORM_", "_FAM_NORM_"])

    # 2) Garantir tabela de sobrepeso consultÃ¡vel por famÃ­lia
    df_sp_tab, sp_index_source = _ensure_df_sobrepeso_index(df_sobrepeso_tabela)
    debug_index_hint = f"fonte_index='{sp_index_source}', indice_name='{df_sp_tab.index.name}'"
    log_callback(f"[DEBUG] df_sobrepeso_tabela preparado ({debug_index_hint}), linhas={len(df_sp_tab)}.")

    # 3) Agrupar itens por SKU para ponderaÃ§Ãµes
    agrupado_por_sku = defaultdict(list)
    for item in itens_detalhados:
        agrupado_por_sku[item.get("sku")] = agrupado_por_sku[item.get("sku")] + [item]

    # 4) PonderaÃ§Ã£o por quantidade e SP real
    for sku, itens in agrupado_por_sku.items():
        qtd_total = 0.0
        qtd_real = 0.0
        ponderador_pos_local = 0.0
        ponderador_neg_local = 0.0

        for item in itens:
            origem = item.get("origem", "fixo")
            sp = _coerce_float(item.get("sp", 0.0), default=0.0)
            chave = _to_str(item.get("chave_pallet", ""))

            # quantidade: prioriza df_fracao por chave_pallete
            if (df_fracao is not None) and (not df_fracao.empty) and ("chave_pallete" in df_fracao.columns):
                mask_frac = df_fracao["chave_pallete"].astype(str).eq(chave)
                qtd = pd.to_numeric(df_fracao.loc[mask_frac, "qtd"], errors="coerce").fillna(0).sum()
            else:
                # cai para df_remessa por ITEM == sku
                if (df_remessa is not None) and (not df_remessa.empty) and ("ITEM" in df_remessa.columns) and ("QUANTIDADE" in df_remessa.columns):
                    mask_rem = df_remessa["ITEM"].astype(str).map(_norm_digits).eq(_norm_digits(sku))
                    qtd = pd.to_numeric(df_remessa.loc[mask_rem, "QUANTIDADE"], errors="coerce").fillna(0).sum()
                else:
                    qtd = 0.0

            qtd_total += qtd

            if origem == "real":
                qtd_real += qtd
                if sp > 0:
                    ponderador_pos_local += sp * qtd
                elif sp < 0:
                    ponderador_neg_local += abs(sp) * qtd

        total_quantidade += qtd_total
        quantidade_com_sp_real += qtd_real
        ponderador_pos += ponderador_pos_local
        ponderador_neg += ponderador_neg_local

        log_callback(f"[SKU] sku='{sku}' qtd_total={qtd_total:.3f} qtd_real={qtd_real:.3f} "
                     f"pond_pos_local={ponderador_pos_local:.3f} pond_neg_local={ponderador_neg_local:.3f}")

    proporcao_sp_real = (quantidade_com_sp_real / total_quantidade) if total_quantidade > 0 else 0.0
    log_callback(f"[AGREGADO] Total_qtde={total_quantidade:.3f} Qtde_com_SP_Real={quantidade_com_sp_real:.3f} "
                 f"ProporÃ§Ã£o_SP_Real={proporcao_sp_real:.2%}")

    # 5) Caminho A: maioria real (>=50%) â†’ mÃ©dias a partir do realizado
    if proporcao_sp_real >= 0.5:
        if quantidade_com_sp_real > 0:
            media_positiva = (ponderador_pos / quantidade_com_sp_real) if ponderador_pos > 0 else 0.02
            media_negativa = (ponderador_neg / quantidade_com_sp_real) if ponderador_neg > 0 else 0.01
        else:
            media_positiva = 0.02
            media_negativa = 0.01
        log_callback("[REGRA] >=50% com SP Real â†’ usando mÃ©dias ponderadas do realizado.")
        return media_positiva, media_negativa, proporcao_sp_real, "REAL"

    # 6) Caminho B: minoria real (<50%) â†’ decidir por famÃ­lia
    #    Descobrir famÃ­lias dos SKUs envolvidos
    familias = set()
    if not df_fam.empty:
        for sku in agrupado_por_sku.keys():
            sku_norm = _norm_digits(sku)
            fam_rows = df_fam.loc[df_fam["_COD_NORM_"] == sku_norm, "_FAM_NORM_"]
            if fam_rows.empty:
                log_callback(f"[FAM] SKU '{sku}' (norm='{sku_norm}') sem famÃ­lia mapeada na base.")
            else:
                fam_val = str(fam_rows.iloc[0])  # se houver duplicata, pega a 1Âª
                familias.add(fam_val)
                log_callback(f"[FAM] SKU '{sku}' (norm='{sku_norm}') â†’ famÃ­lia='{fam_val}'.")
    else:
        log_callback("[FAM] df_base_familia vazio ou sem colunas esperadas. SÃ³ MIX possÃ­vel agora.")

    # Se todos os SKUs convergem para a mesma famÃ­lia, usamos a famÃ­lia; caso contrÃ¡rio, MIX
    row = pd.DataFrame()
    if len(familias) == 1:
        fam = list(familias)[0].upper()
        # normalizaÃ§Ãµes simples para as 2 famÃ­lias principais
        is_biscoito = "BISCOITO" in fam
        is_massa = "MASSA" in fam

        if is_biscoito:
            familia_detectada = "BISCOITO"
            # Match no Ã­ndice padronizado (upper/strip via contains case-insensitive)
            row = df_sp_tab.loc[df_sp_tab.index.astype(str).str.contains("BISCOITO", case=False, regex=False)]
            log_callback("[TAB] FamÃ­lia detectada: BISCOITO (buscando linha em df_sobrepeso_tabela).")
        elif is_massa:
            familia_detectada = "MASSA"
            row = df_sp_tab.loc[df_sp_tab.index.astype(str).str.contains("MASSA", case=False, regex=False)]
            log_callback("[TAB] FamÃ­lia detectada: MASSA (buscando linha em df_sobrepeso_tabela).")
        else:
            familia_detectada = fam  # mantÃ©m a famÃ­lia real encontrada
            row = df_sp_tab.loc[df_sp_tab.index.astype(str).str.contains(fam, case=False, regex=False)]
            log_callback(f"[TAB] FamÃ­lia detectada: '{fam}' (buscando linha correspondente).")

        if row.empty:
            log_callback(f"[WARN] NÃ£o achei linha para famÃ­lia '{familia_detectada}' em df_sobrepeso_tabela. Fallback para MIX.")
            familia_detectada = "MIX"
    else:
        if len(familias) == 0:
            log_callback("[FAM] Nenhuma famÃ­lia encontrada para os SKUs â†’ Fallback MIX.")
        else:
            log_callback(f"[FAM] FamÃ­lias mÃºltiplas detectadas ({len(familias)}): {sorted(familias)} â†’ MIX.")
        familia_detectada = "MIX"

    # Se ainda nÃ£o temos 'row', buscar MIX
    if row.empty:
        # tenta por MIX
        row = df_sp_tab.loc[df_sp_tab.index.astype(str).str.contains("MIX", case=False, regex=False)]
        if row.empty:
            log_callback("[ERRO] df_sobrepeso_tabela nÃ£o contÃ©m entrada para 'MIX'. Usando defaults 0.02/0.01.")
            media_positiva = 0.02
            media_negativa = 0.01
            return media_positiva, media_negativa, proporcao_sp_real, familia_detectada

    # Extrair (+) e (-)
    if "(+)" not in row.columns or "(-)" not in row.columns:
        # Em alguns arquivos, pode estar como '+', '-', 'POS', 'NEG' etc. Tentar alternativas
        cand_pos = [c for c in row.columns if str(c).strip().upper() in ["(+)", "+", "POS", "POSITIVO", "SOBREPOS_POS", "SOBREPOS+"]] or ["(+)" ]
        cand_neg = [c for c in row.columns if str(c).strip().upper() in ["(-)", "-", "NEG", "NEGATIVO", "SOBREPOS_NEG", "SOBREPOS-"]] or ["(-)"]

        col_pos = cand_pos[0] if cand_pos[0] in row.columns else None
        col_neg = cand_neg[0] if cand_neg[0] in row.columns else None

        if (col_pos is None) or (col_neg is None):
            log_callback(f"[ERRO] Colunas de sobrepeso '(+)'/'(-)' nÃ£o encontradas em df_sobrepeso_tabela. "
                         f"Cols disponÃ­veis: {list(row.columns)}. Usando defaults 0.02/0.01.")
            media_positiva = 0.02
            media_negativa = 0.01
            return media_positiva, media_negativa, proporcao_sp_real, familia_detectada
    else:
        col_pos, col_neg = "(+)", "(-)"

    val_pos = _coerce_float(row.iloc[0][col_pos], default=0.02)
    val_neg = _coerce_float(row.iloc[0][col_neg], default=0.01)

    media_positiva = val_pos
    media_negativa = val_neg

    log_callback(f"[RESULT] FamÃ­lia='{familia_detectada}' | Sobrepeso fÃ­sico (+)={media_positiva:.4f} | (-)={media_negativa:.4f}")

    return media_positiva, media_negativa, proporcao_sp_real, familia_detectada


def preencher_formulario_com_openpyxl(path_copia, dados, itens_detalhados, log_callback, df_sku, df_remessa, df_fracao):
    try:
        dados_tabela = {'(+)': [0.02, 0.005, 0.04], '(-)': [0.01, 0.01, 0.01]}
        index = ['CARGA COM MIX', 'EXCLUSIVO MASSAS', 'EXCLUSIVO BISCOITOS']
        df_sobrepeso_tabela = pd.DataFrame(dados_tabela, index=index)

        sp_pos, sp_neg, proporcao_sp_real, familia_detectada = calcular_limites_sobrepeso_por_quantidade(
            dados, itens_detalhados, df_base_familia, df_sobrepeso_tabela, df_sku, df_remessa, df_fracao, log_callback
        )

        wb = load_workbook(path_copia)
        ws = wb["FORMULARIO"]

        ws["A16"] = f"Sobrepeso para (+): {sp_pos*100:.2f}%"
        ws["A18"] = f"Sobrepeso para (-): {sp_neg*100:.2f}%"
        ws["D7"] = f"{proporcao_sp_real*100:.2f}% x {(1 - proporcao_sp_real)*100:.2f}%"
        ws["B5"] = str(familia_detectada or "MIX")

        ws["B4"] = dados["remessa"]
        ws["B6"] = dados["qtd_skus"]
        ws["B7"] = dados["placa"]
        ws["B8"] = dados["turno"]
        ws["B9"] = dados["peso_vazio"]
        ws["B10"] = dados["peso_base"]
        ws["B11"] = dados["sp_total"]
        ws["B12"] = dados["peso_com_sp"]
        ws["B13"] = dados["peso_total_final"]
        ws["B14"] = dados["media_sp"]
        ws["B16"] = dados["peso_total_final"] * (1 + sp_pos)
        ws["B17"] = dados["peso_total_final"]
        ws["B18"] = dados["peso_total_final"] * (1 - sp_neg)
        ws["D4"] = dados["qtd_paletes"]
        ws["D9"] = dados["qtd_paletes"] * 22

        linha_inicio = 12
        linha_fim = 46
        max_itens = linha_fim - linha_inicio + 1

        itens_ordenados = (
            [i for i in itens_detalhados if i["origem"] == "real"] +
            [i for i in itens_detalhados if i["origem"] == "fixo"] +
            [i for i in itens_detalhados if i["origem"] == "nÃ£o encontrado"]
        )

        for idx, item in enumerate(itens_ordenados[:max_itens]):
            linha = linha_inicio + idx
            ws[f"C{linha}"] = f"{item['sku']} ({item['origem']})"
            ws[f"D{linha}"] = f"{item['sp']*100:.3f}"

        wb.save(path_copia)
        wb.close()
        log_callback("FormulÃ¡rio preenchido e salvo com sucesso.")
    except Exception as e:
        log_callback(f"Erro no preenchimento: {e}")
        raise

def gerar_relatorio_diferenca(remessa_num, peso_final_balanÃ§a, peso_veiculo_vazio, df_remessa, df_sku, peso_estimado_total, pasta_excel, log_callback):
    try:
        peso_final_balanÃ§a = converter_para_float_seguro(peso_final_balanÃ§a)
        peso_veiculo_vazio = converter_para_float_seguro(peso_veiculo_vazio)
        peso_estimado_total = converter_para_float_seguro(peso_estimado_total)

        pasta_destino = os.path.join(pasta_excel, "Analise_divergencia")
        os.makedirs(pasta_destino, exist_ok=True)

        skus = df_remessa["ITEM"].unique()
        dados_relatorio = []
        peso_base_total_liq = 0.0

        for sku in skus:
            qtd = converter_para_float_seguro(df_remessa[df_remessa["ITEM"] == sku]["QUANTIDADE"].sum())
            df_sku_filtrado = df_sku[df_sku["COD_PRODUTO"] == sku]
            if df_sku_filtrado.empty:
                continue

            unidade = df_sku_filtrado.iloc[0]["DESC_UNID_MEDID"]
            peso_unit_liq = converter_para_float_seguro(df_sku_filtrado.iloc[0]["QTDE_PESO_LIQ"])
            peso_total_liq = converter_para_float_seguro(peso_unit_liq * qtd)
            peso_base_total_liq += peso_total_liq

            dados_relatorio.append({
                "SKU": sku,
                "Unidade": unidade,
                "Quantidade": qtd,
                "Peso Total LÃ­quido": peso_total_liq,
                "Peso Unit. LÃ­quido": peso_unit_liq
            })

        df_dados = pd.DataFrame(dados_relatorio)
        # --- sanidade: remover colunas duplicadas e tipar numÃ©ricos ---
        dups = df_dados.columns[df_dados.columns.duplicated()].tolist()
        if dups:
            log_callback(f"[div] colunas duplicadas detectadas no df_dados: {dups} â†’ mantendo a 1Âª ocorrÃªncia")
            df_dados = df_dados.loc[:, ~df_dados.columns.duplicated()].copy()

        for col in ["Quantidade", "Peso Total LÃ­quido", "Peso Unit. LÃ­quido"]:
            if col in df_dados.columns:
                df_dados[col] = pd.to_numeric(df_dados[col], errors="coerce").fillna(0.0)

        peso_carga_real = converter_para_float_seguro(peso_final_balanÃ§a - peso_veiculo_vazio)
        diferenca_total = converter_para_float_seguro((peso_estimado_total + peso_veiculo_vazio) - peso_final_balanÃ§a)

        if peso_base_total_liq > 0:
            df_dados["% Peso"] = df_dados["Peso Total LÃ­quido"] / peso_base_total_liq
        else:
            df_dados["% Peso"] = 0.0

        df_dados["Peso Proporcional Real"] = df_dados["% Peso"] * float(peso_carga_real)

        num = df_dados["Peso Proporcional Real"]
        den = df_dados["Peso Unit. LÃ­quido"].replace(0, np.nan)

        qre = np.floor((num / den).round(0)).fillna(0.0)  
        df_dados["Quantidade Real Estimada"] = qre.astype(int)


        df_dados["DiferenÃ§a Estimada (kg)"] = df_dados["% Peso"] * diferenca_total
        df_dados["Unid. Estimada de DivergÃªncia"] = df_dados["Quantidade Real Estimada"] - df_dados["Quantidade"]

        nome_pdf = f"AnÃ¡lise quantitativa - {remessa_num}.pdf"
        caminho_pdf = os.path.join(pasta_destino, nome_pdf)

        fig, ax = plt.subplots(figsize=(10, 6))
        largura_barra = 0.35
        x = range(len(df_dados))
        qtd_esperada = df_dados["Quantidade"]
        qtd_real = df_dados["Quantidade Real Estimada"]

        ax.bar([i - largura_barra/2 for i in x], qtd_esperada, width=largura_barra, label="Quantidade Esperada")
        ax.bar([i + largura_barra/2 for i in x], qtd_real, width=largura_barra, label="Quantidade Real Estimada")
        ax.set_ylabel("Quantidade (unidades)")
        ax.set_xlabel("SKU")
        ax.set_title("Comparativo: Quantidade Esperada vs Real Estimada por SKU")
        ax.set_xticks(list(x))
        ax.set_xticklabels(df_dados["SKU"].astype(str))
        ax.axhline(0)
        ax.legend()

        for i in x:
            ax.text(i - largura_barra/2, qtd_esperada.iloc[i], f"{qtd_esperada.iloc[i]:.0f}", ha="center", va="bottom")
            ax.text(i + largura_barra/2, qtd_real.iloc[i], f"{qtd_real.iloc[i]:.0f}", ha="center", va="bottom")

        with PdfPages(caminho_pdf) as pdf:
            fig_tabela, ax_tabela = plt.subplots(figsize=(12, len(df_dados) * 0.5 + 3))
            ax_tabela.axis("off")
            table_data = [
                ["SKU", "Unidade", "Qtd. Enviada", "Qtd. Real Estimada", "Peso Total LÃ­quido", "% do Peso", "DiferenÃ§a (kg)", "DivergÃªncia (unid)"]
            ] + df_dados[["SKU", "Unidade", "Quantidade", "Quantidade Real Estimada", "Peso Total LÃ­quido", "% Peso", "DiferenÃ§a Estimada (kg)", "Unid. Estimada de DivergÃªncia"]].round(2).values.tolist()

            tabela = ax_tabela.table(cellText=table_data, colLabels=None, loc="center", cellLoc="center")
            tabela.auto_set_font_size(False)
            tabela.set_fontsize(10)
            tabela.scale(1, 1.5)

            titulo = (
                f"RelatÃ³rio Comparativo - Remessa {remessa_num}\n"
                f"Peso estimado: {peso_estimado_total:.2f} kg | "
                f"Peso balanÃ§a: {peso_final_balanÃ§a:.2f} kg | "
                f"Peso veÃ­culo: {peso_veiculo_vazio:.2f} kg | "
                f"DiferenÃ§a: {diferenca_total:.2f} kg"
            )
            fig_tabela.suptitle(titulo, fontsize=12)
            pdf.savefig(fig_tabela, bbox_inches="tight")
            pdf.savefig(fig, bbox_inches="tight")

        plt.close("all")
        log_callback(f"RelatÃ³rio salvo em: {caminho_pdf}")
        return caminho_pdf

    except Exception as e:
        log_callback(f"Erro ao gerar relatÃ³rio de divergÃªncia: {str(e)}")
        raise

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")

class EdicaoRemessaFrame(ctk.CTkFrame):
    def __init__(self, master, df_expedicao, log_callback, app, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.fonte_dir = BASE_DIR_AUD
        self.remessa_editada = False
        self.df_expedicao_original = df_expedicao.dropna(subset=["ITEM"]).copy()
        # garante alias
        if "CHAVE_PALETE" not in self.df_expedicao_original.columns and "COD_RASTREABILIDADE" in self.df_expedicao_original.columns:
            self.df_expedicao_original["CHAVE_PALETE"] = self.df_expedicao_original["COD_RASTREABILIDADE"]

        self.df_expedicao_original["REMESSA"] = (
            self.df_expedicao_original["REMESSA"].astype(str)
            .str.strip()
            .str.replace(r"\.0$", "", regex=True)
        )

        self.dados_remessa = pd.DataFrame(columns=["ITEM", "QUANTIDADE", "CHAVE_PALETE"])

        self.frame_superior = ctk.CTkFrame(self)
        self.frame_superior.pack(fill="x", padx=10, pady=(10, 0))
        self.label_status = ctk.CTkLabel(self.frame_superior, text="", text_color="green", font=("Arial", 12, "bold"), corner_radius=5, padx=10, pady=5)
        self.label_status.pack(fill="x")

        self.df_expedicao = df_expedicao
        self.log_callback = log_callback
        self.app = app

        self.remessa_var = StringVar()
        self.filtro_chave = StringVar()
        self.filtro_sku = StringVar()

        top_button_frame = ctk.CTkFrame(self)
        top_button_frame.pack(fill="x", padx=10, pady=5)
        ctk.CTkButton(top_button_frame, text="âž• Adicionar Linha", command=self.adicionar_linha, width=150, fg_color="#2a7fff").pack(side="left", padx=5)
        ctk.CTkButton(top_button_frame, text="ðŸ’¾ Salvar AlteraÃ§Ãµes", command=self.salvar_alteracoes, width=150).pack(side="right", padx=5)
        ctk.CTkLabel(top_button_frame, text="EdiÃ§Ã£o de Remessa", font=("Arial", 14, "bold")).pack(side="left", padx=5)

        form_frame = ctk.CTkFrame(self)
        form_frame.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(form_frame, text="NÃºmero da Remessa:").grid(row=0, column=0, sticky="w", padx=5)
        self.entry_remessa = ctk.CTkEntry(form_frame, textvariable=self.remessa_var)
        self.entry_remessa.grid(row=0, column=1, sticky="ew", padx=5)
        ctk.CTkButton(form_frame, text="ðŸ”Ž Buscar", command=self.carregar_dados).grid(row=0, column=2, padx=5)

        ctk.CTkLabel(form_frame, text="Filtrar por Chave Pallet:").grid(row=1, column=0, sticky="w", padx=5)
        self.entry_filtro_chave = ctk.CTkEntry(form_frame, textvariable=self.filtro_chave)
        self.entry_filtro_chave.grid(row=1, column=1, sticky="ew", padx=5)
        self.entry_filtro_chave.bind("<KeyRelease>", lambda e: self.filtrar_dados())

        ctk.CTkLabel(form_frame, text="Filtrar por SKU:").grid(row=2, column=0, sticky="w", padx=5)
        self.entry_filtro_sku = ctk.CTkEntry(form_frame, textvariable=self.filtro_sku)
        self.entry_filtro_sku.grid(row=2, column=1, sticky="ew", padx=5)
        self.entry_filtro_sku.bind("<KeyRelease>", lambda e: self.filtrar_dados())
        ctk.CTkButton(form_frame, text="ðŸ” Aplicar Filtros", command=self.filtrar_dados).grid(row=1, column=2, rowspan=2, padx=5)

        self.frame_horizontal = ctk.CTkFrame(self)
        self.frame_horizontal.pack(fill="x", padx=10, pady=5)
        self.tabela_frame = ctk.CTkScrollableFrame(self.frame_horizontal, height=350)
        self.tabela_frame.pack(side="left", fill="both", expand=True, padx=(0, 5))

        self.totals_center_frame = ctk.CTkFrame(self.frame_horizontal, width=150, fg_color="#1a1a1a")
        self.totals_center_frame.pack(side="left", fill="y", padx=5)
        ctk.CTkLabel(self.totals_center_frame, text="Totais Gerais:", font=("Arial", 12, "bold")).pack(pady=(10, 5))

        self.total_itens_value = ctk.CTkLabel(self.totals_center_frame, text="0", text_color="#55FF55", font=("Arial", 10, "bold"))
        ctk.CTkLabel(self.totals_center_frame, text="Itens:", text_color="#FF5555").pack(side="top")
        self.total_itens_value.pack(side="top")

        self.total_qtd_value = ctk.CTkLabel(self.totals_center_frame, text="0", text_color="#55FF55", font=("Arial", 10, "bold"))
        ctk.CTkLabel(self.totals_center_frame, text="Qtd Total:", text_color="#FF5555").pack(side="top")
        self.total_qtd_value.pack(side="top")

        self.sku_totals_lateral = ctk.CTkFrame(self.frame_horizontal, width=350, fg_color="#1f1f1f")
        self.sku_totals_lateral.pack_propagate(False)
        self.sku_totals_lateral.pack(side="left", fill="y", padx=(5, 0))
        self.sku_totals_label = ctk.CTkLabel(self.sku_totals_lateral, text="Totais por SKU:")
        self.sku_totals_label.pack(anchor="w", padx=5, pady=(0, 5))
        self.filtro_sku_totais = StringVar()
        self.entry_filtro_sku_totais = ctk.CTkEntry(self.sku_totals_lateral, textvariable=self.filtro_sku_totais, placeholder_text="Filtrar SKU")
        self.entry_filtro_sku_totais.pack(fill="x", padx=5, pady=(0, 5))
        self.entry_filtro_sku_totais.bind("<KeyRelease>", lambda e: self.atualizar_totais_sku(self.filtro_sku_totais.get()))
        self.sku_totals_scroll = ctk.CTkScrollableFrame(self.sku_totals_lateral, height=350)
        self.sku_totals_scroll.pack(fill="both", expand=True)


    def format_number(self, num):
        if pd.isna(num):
            return "N/A"
        try:
            num_float = float(num)
            return str(int(num_float)) if num_float.is_integer() else str(num_float)
        except:
            return str(num)

    def adicionar_linha(self):
        try:
            df_nova = pd.DataFrame([{'ITEM': '', 'QUANTIDADE': 0, 'CHAVE_PALETE': None}])
            self.dados_remessa = pd.concat([self.dados_remessa, df_nova], ignore_index=True)
            self.renderizar_tabela()
            self.tabela_frame._parent_canvas.yview_moveto(1.0)
            self.label_status.configure(text="Nova linha adicionada - preencha os dados", text_color="blue")
            self.log_callback("Nova linha adicionada para ediÃ§Ã£o manual")
        except Exception as e:
            self.log_callback(f"Erro ao adicionar linha: {str(e)}")
            self.label_status.configure(text=f"Erro ao adicionar linha: {str(e)}", text_color="red")

    def atualizar_totais_sku(self, filtro=""):
        for w in self.sku_totals_scroll.winfo_children():
            w.destroy()
        filtro = filtro.lower().strip()
        skus_totais = {}
        for entry_sku, entry_qtd, _ in getattr(self, "entry_widgets", []):
            sku = entry_sku.get() if entry_sku else getattr(entry_qtd, "sku_associado", "")
            if not sku:
                continue
            qtd = converter_para_float_seguro(entry_qtd.get())
            skus_totais[sku] = skus_totais.get(sku, 0) + qtd
        skus_filtrados = {k: v for k, v in skus_totais.items() if filtro in str(k).lower()}
        for sku, total in sorted(skus_filtrados.items(), key=lambda x: str(x[0])):
            linha = ctk.CTkLabel(self.sku_totals_scroll, text=f"{sku} = {self.format_number(total)}", anchor="w")
            linha.pack(fill="x", padx=6, pady=1)

    def update_totals(self, event=None):
        total_itens = 0
        total_qtd = 0.0
        for entry_sku, entry_qtd, _ in getattr(self, "entry_widgets", []):
            sku = entry_sku.get() if entry_sku else getattr(entry_qtd, "sku_associado", "")
            if sku:
                total_itens += 1
            total_qtd += converter_para_float_seguro(entry_qtd.get())
        self.total_itens_value.configure(text=str(total_itens))
        self.total_qtd_value.configure(text=self.format_number(total_qtd))
        self.atualizar_totais_sku()

    def remessa_existe_na_base(self, remessa, df_base):
        if df_base.empty or 'REMESSA' not in df_base.columns:
            return False
        try:
            mask = _match_remessa_series(df_base['REMESSA'], remessa)
            return bool(mask.any())
        except Exception as e:
            self.log_callback(f"Erro ao verificar remessa: {str(e)}")
            return False


    def carregar_dados(self):
        remessa = self.remessa_var.get().strip()
        if not remessa:
            self.log_callback("Digite uma remessa vÃ¡lida.")
            return

        try:
            self.log_callback(f"Verificando base auxiliar para remessa {remessa}...")
            df_base_auxiliar = carregar_base_auxiliar(self.fonte_dir)

            df_filtrado = pd.DataFrame()
            self.remessa_editada = False

            if not df_base_auxiliar.empty:
                mask_aux = _match_remessa_series(df_base_auxiliar['REMESSA'], remessa)
                if mask_aux.any():
                    self.remessa_editada = True
                    self.label_status.configure(
                        text="ATENÃ‡ÃƒO: Esta remessa jÃ¡ foi editada anteriormente!",
                        text_color="orange"
                    )
                    self.log_callback(f"Remessa {remessa} encontrada na base auxiliar - carregando dados editados")
                    df_filtrado = df_base_auxiliar.loc[mask_aux]
                else:
                    self.log_callback(f"Remessa {remessa} nÃ£o encontrada na base auxiliar - buscando na base original")
            else:
                self.log_callback("Base auxiliar vazia ou nÃ£o encontrada - buscando na base original")

            if df_filtrado.empty:
                self.log_callback(f"Buscando remessa {remessa} na base original...")
                mask_ori = _match_remessa_series(self.df_expedicao_original['REMESSA'], remessa)
                df_filtrado = self.df_expedicao_original.loc[mask_ori].dropna(subset=['ITEM']).copy()
                if df_filtrado.empty:
                    self.log_callback("Remessa nÃ£o encontrado em nenhuma base!")
                    self.label_status.configure(text="Remessa nÃ£o encontrada em nenhuma base!", text_color="red")
                    return

            colunas_unicas = ['ITEM', 'QUANTIDADE', 'CHAVE_PALETE']
            df_sem_duplicatas = df_filtrado.drop_duplicates(subset=colunas_unicas)
            if len(df_filtrado) != len(df_sem_duplicatas):
                duplicatas = len(df_filtrado) - len(df_sem_duplicatas)
                self.log_callback(f"Removidas {duplicatas} linhas duplicadas automaticamente.")

            self.dados_remessa = df_sem_duplicatas[['ITEM', 'QUANTIDADE', 'CHAVE_PALETE']].copy()
            self.renderizar_tabela()

        except Exception as e:
            error_msg = f"Erro ao carregar dados: {str(e)}"
            self.log_callback(error_msg)
            self.label_status.configure(text=error_msg, text_color="red")
            self.log_callback(f"Traceback completo: {traceback.format_exc()}")


    def salvar_alteracoes_antes_filtro(self):
        if not hasattr(self, "entry_widgets") or not self.entry_widgets:
            return
        try:
            for entry_sku, entry_qtd, entry_chave in self.entry_widgets:
                original_idx = getattr(entry_qtd, "original_idx", None)
                if original_idx is None or original_idx not in self.dados_remessa.index:
                    continue
                sku = entry_sku.get() if entry_sku else getattr(entry_qtd, "sku_associado", "")
                qtd = converter_para_float_seguro(entry_qtd.get())
                chave = entry_chave.get()
                self.dados_remessa.at[original_idx, "ITEM"] = sku or None
                self.dados_remessa.at[original_idx, "QUANTIDADE"] = qtd
                self.dados_remessa.at[original_idx, "CHAVE_PALETE"] = chave or None
            self.log_callback("AlteraÃ§Ãµes salvas antes de aplicar filtros")
        except Exception as e:
            self.log_callback(f"Erro ao salvar alteraÃ§Ãµes antes de filtrar: {str(e)}")

    def filtrar_dados(self):
        self.salvar_alteracoes_antes_filtro()
        filtro_chave = self.filtro_chave.get().strip().lower()
        filtro_sku = self.filtro_sku.get().strip().lower()
        df_filtrado = self.dados_remessa.copy()
        if filtro_chave:
            df_filtrado = df_filtrado[df_filtrado["CHAVE_PALETE"].fillna("").astype(str).str.lower().str.contains(filtro_chave)]
        if filtro_sku:
            df_filtrado = df_filtrado[df_filtrado["ITEM"].astype(str).str.lower().str.contains(filtro_sku)]
        self.renderizar_tabela(df_filtrado)

    def salvar_alteracoes(self):
        try:
            remessa = self.remessa_var.get().strip()
            if not remessa:
                self.log_callback("Nenhuma remessa selecionada para salvar")
                return

            df_completo = pd.DataFrame(columns=["ITEM", "QUANTIDADE", "CHAVE_PALETE"])
            for entry_sku, entry_qtd, entry_chave in self.entry_widgets:
                sku = entry_sku.get() if entry_sku else getattr(entry_qtd, "sku_associado", "")
                if not sku:
                    continue
                qtd = converter_para_float_seguro(entry_qtd.get())
                chave = entry_chave.get() or None
                df_completo = pd.concat([df_completo, pd.DataFrame([{"ITEM": str(sku).strip(), "QUANTIDADE": qtd, "CHAVE_PALETE": (str(chave).strip() if chave else None)}])], ignore_index=True)

            if df_completo.empty:
                self.label_status.configure(text="Nenhum dado vÃ¡lido para salvar!", text_color="orange")
                return

            df_completo["REMESSA"] = remessa
            df_completo = df_completo.drop_duplicates(subset=["ITEM", "CHAVE_PALETE"], keep="last")
            salvar_em_base_auxiliar(df_completo, remessa, self.log_callback, self.fonte_dir)

            self.dados_remessa = df_completo[["ITEM", "QUANTIDADE", "CHAVE_PALETE"]].copy()
            self.renderizar_tabela()
            self.label_status.configure(text=f"AlteraÃ§Ãµes salvas! Itens: {len(df_completo)}", text_color="green")
            self.log_callback(f"Remessa {remessa} salva com {len(df_completo)} itens")

        except Exception as e:
            self.label_status.configure(text=f"Erro ao salvar: {str(e)}", text_color="red")
            self.log_callback(f"[ERRO] Ao salvar: {traceback.format_exc()}")

    def renderizar_tabela(self, df_filtrado=None):
        for widget in self.tabela_frame.winfo_children():
            widget.destroy()
        df_exibicao = (df_filtrado.copy() if df_filtrado is not None else self.dados_remessa.copy()).drop_duplicates(subset=["ITEM", "QUANTIDADE", "CHAVE_PALETE"], keep="last")
        if df_exibicao.empty:
            self.log_callback("Nenhum dado para exibir.")
            return

        self.entry_widgets = []
        headers = ["SKU", "Quantidade", "Chave Pallet", "AÃ§Ãµes"]
        for col, header in enumerate(headers):
            ctk.CTkLabel(self.tabela_frame, text=header, font=("Arial", 12, "bold")).grid(row=0, column=col, padx=10, pady=5, sticky="ew")

        for display_row, (idx, row) in enumerate(df_exibicao.iterrows(), start=1):
            try:
                sku = str(row["ITEM"]) if pd.notna(row["ITEM"]) else ""
                qtd = str(row["QUANTIDADE"]) if pd.notna(row["QUANTIDADE"]) else "0"
                chave = str(row["CHAVE_PALETE"]) if pd.notna(row["CHAVE_PALETE"]) else ""

                if sku == "":
                    entry_sku = ctk.CTkEntry(self.tabela_frame, placeholder_text="Digite o SKU", fg_color="#2a2a4a", border_color="#4a4a7a")
                    entry_sku.insert(0, sku)
                    entry_sku.grid(row=display_row, column=0, padx=10, pady=2, sticky="ew")
                else:
                    ctk.CTkLabel(self.tabela_frame, text=sku).grid(row=display_row, column=0, padx=10, pady=2, sticky="w")
                    entry_sku = None

                entry_qtd = ctk.CTkEntry(self.tabela_frame)
                entry_qtd.insert(0, qtd)
                entry_qtd.grid(row=display_row, column=1, padx=10, pady=2, sticky="ew")
                entry_qtd.sku_associado = entry_sku.get() if entry_sku else sku
                entry_qtd.original_idx = idx
                entry_qtd.bind("<KeyRelease>", self.update_totals)

                entry_chave = ctk.CTkEntry(self.tabela_frame)
                entry_chave.insert(0, chave)
                entry_chave.grid(row=display_row, column=2, padx=10, pady=2, sticky="ew")
                entry_chave.original_idx = idx

                botao_excluir = ctk.CTkButton(self.tabela_frame, text="ðŸ—‘", width=30, command=lambda idx=idx: self.remover_linha(idx), fg_color="#d44646", hover_color="#a33535")
                botao_excluir.grid(row=display_row, column=3, padx=5, pady=2)

                if entry_sku:
                    entry_sku.bind("<KeyRelease>", lambda e, entry_qtd=entry_qtd: setattr(entry_qtd, "sku_associado", e.widget.get()))
                    self.entry_widgets.append((entry_sku, entry_qtd, entry_chave))
                else:
                    self.entry_widgets.append((None, entry_qtd, entry_chave))
            except Exception as e:
                self.log_callback(f"Erro ao renderizar linha {idx}: {str(e)}")
                continue

        self.update_totals()

    def remover_linha(self, index):
        try:
            if not self.dados_remessa.empty and index in self.dados_remessa.index:
                self.dados_remessa = self.dados_remessa.drop(index)
                self.renderizar_tabela()
                self.label_status.configure(text="Linha removida com sucesso!", text_color="green")
                self.log_callback(f"Linha {index} removida da remessa {self.remessa_var.get()}")
                self.update_totals()
            else:
                self.log_callback(f"Ãndice invÃ¡lido para remoÃ§Ã£o: {index}")
        except Exception as e:
            self.log_callback(f"Erro ao remover linha: {str(e)}")
            self.label_status.configure(text=f"Erro ao remover linha: {str(e)}", text_color="red")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Simulador de Sobrepeso 4.0")
        self.geometry("1200x1000")

        self.placa = ctk.StringVar()
        self.turno = ctk.StringVar()
        self.remessa = ctk.StringVar()
        self.qtd_paletes = ctk.StringVar()
        self.peso_vazio = ctk.StringVar()
        self.peso_balanca = ctk.StringVar()

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.tabs = ctk.CTkTabview(self)
        self.tabs.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        self.tab_simulador = self.tabs.add("Simulador")
        simulador_main_frame = ctk.CTkFrame(self.tab_simulador)
        simulador_main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        controls_frame = ctk.CTkFrame(simulador_main_frame, width=350)
        controls_frame.pack(side="left", fill="y", padx=(0, 10), pady=5)

        logs_frame = ctk.CTkFrame(simulador_main_frame)
        logs_frame.pack(side="right", fill="both", expand=True, padx=5, pady=5)

        ctk.CTkLabel(controls_frame, text="ConfiguraÃ§Ã£o do Simulador", font=("Arial", 14, "bold")).pack(pady=(0, 10), anchor="w")

        campos = [
            ("Placa:", self.placa, None),
            ("Turno:", self.turno, ["A", "B", "C"]),
            ("Remessa:", self.remessa, None),
            ("Quantidade de Paletes:", self.qtd_paletes, None),
            ("Peso VeÃ­culo Vazio (kg):", self.peso_vazio, None),
            ("Peso Final BalanÃ§a (kg):", self.peso_balanca, None)
        ]
        for texto, var, valores in campos:
            ctk.CTkLabel(controls_frame, text=texto).pack(anchor="w", pady=(5, 0))
            if valores:
                ctk.CTkComboBox(controls_frame, values=valores, variable=var).pack(fill="x", pady=(0, 5))
            else:
                ctk.CTkEntry(controls_frame, textvariable=var).pack(fill="x", pady=(0, 5))

        button_frame = ctk.CTkFrame(controls_frame, fg_color="transparent")
        button_frame.pack(fill="x", pady=10)
        ctk.CTkButton(button_frame, text="Calcular", command=self.iniciar_processamento, fg_color="#2aa745").pack(side="left", expand=True, padx=2)
        ctk.CTkButton(button_frame, text="ðŸ”„ Refresh", command=self.atualizar_bases, width=80).pack(side="right", padx=2)

        self.progress_bar = ctk.CTkProgressBar(controls_frame, mode="determinate")
        self.progress_bar.set(0)
        self.progress_bar.pack(fill="x", pady=(5, 0))
        self.progress_bar.pack_forget()

        logs_header = ctk.CTkFrame(logs_frame, fg_color="transparent")
        logs_header.pack(fill="x", pady=(0, 5))
        ctk.CTkLabel(logs_header, text="ðŸ“œ HistÃ³rico de ExecuÃ§Ã£o", font=("Arial", 14, "bold")).pack(side="left", padx=5)
        ctk.CTkButton(logs_header, text="ðŸ§¹ Limpar HistÃ³rico", command=self.limpar_logs, width=120, fg_color="#d44646", hover_color="#a33535").pack(side="right")

        self.log_display = ctk.CTkTextbox(logs_frame, wrap="word", font=("Consolas", 10), activate_scrollbars=True)
        self.log_display.pack(fill="both", expand=True)
        self.log_display.configure(state="disabled")

        self.log_text = []
        self.log_geral = []
        self.log_tecnico = []

        self.tab_edicao = self.tabs.add("EdiÃ§Ã£o de Remessa")
        self.df_expedicao = carregar_base_expedicao_csv(BASE_DIR_AUD, log=self.add_log)
        self.add_log(f"Base de expediÃ§Ã£o (CSV Auditoria) carregada com {len(self.df_expedicao)} linhas "
             f"e {self.df_expedicao['REMESSA'].nunique()} remessas.")

        self.edicao_frame = EdicaoRemessaFrame(master=self.tab_edicao, df_expedicao=self.df_expedicao, log_callback=self.add_log, app=self)
        self.edicao_frame.pack(fill="both", expand=True, padx=10, pady=10)

        footer_label = ctk.CTkLabel(self, text="Desenvolvido por Douglas Lins - Analista de LogÃ­stica", font=("Arial", 10), anchor="center")
        footer_label.grid(row=1, column=0, columnspan=2, pady=(0, 10))



    def atualizar_bases(self):
        try:
            self.add_log("â³ Atualizando bases do OneDrive...")
            path_base_sobrepeso = os.path.join(BASE_DIR_DOCS, "Base_sobrepeso_real.xlsx")
            path_base_sap       = os.path.join(BASE_DIR_DOCS, "base_sap.xlsx")

            df_expedicao = carregar_base_expedicao_csv(BASE_DIR_AUD, log=self.add_log)

            _ = pd.read_excel(path_base_sap, sheet_name="Sheet1")       
            _ = pd.read_excel(path_base_sobrepeso, sheet_name="SOBREPESO") 

            self.edicao_frame.df_expedicao = df_expedicao
            self.df_expedicao = df_expedicao

           
            self.edicao_frame.remessa_var.set("")
            self.edicao_frame.filtro_chave.set("")
            self.edicao_frame.filtro_sku.set("")
            for widget in self.edicao_frame.tabela_frame.winfo_children():
                widget.destroy()

            self.add_log("Bases atualizadas com sucesso!")
        except Exception as e:
            self.add_log(f"Erro ao atualizar bases: {e}")

    def add_log(self, msg):
        timestamp = datetime.now().strftime("%H:%M:%S")
        entrada = f"[{timestamp}] {msg}"
        self.log_text.append(entrada)
        self.log_display.configure(state="normal")
        self.log_display.insert("end", entrada + "\n")
        self.log_display.see("end")
        self.log_display.configure(state="disabled")

    def limpar_logs(self):
        self.log_text.clear()
        self.log_geral.clear()
        self.log_display.configure(state="normal")
        self.log_display.delete("1.0", "end")
        self.log_display.configure(state="disabled")

    def log_callback_completo(self, mensagem):
        print(mensagem)
        self.log_geral.append(mensagem)
        self.add_log(mensagem)

    def log_callback_tecnico(self, mensagem):
        timestamp = datetime.now().strftime("%H:%M:%S")
        txt = f"[{timestamp}] {mensagem}"
        self.log_tecnico.append(txt)
        self.add_log(mensagem)


    def iniciar_processamento(self):
        self.progress_bar.pack()
        self.progress_bar.set(0)
        self.add_log("Iniciando cÃ¡lculo...")
        threading.Thread(target=self.processar).start()

    def processar(self):
        try:
            self.progress_bar.set(0.1)

            remessa = int(float(self.remessa.get()))
            peso_vazio = float(self.peso_vazio.get())
            peso_balanca = float(self.peso_balanca.get())
            qtd_paletes = int(float(self.qtd_paletes.get()))

            path_base_sobrepeso = os.path.join(BASE_DIR_DOCS, "Base_sobrepeso_real.xlsx")
            path_base_sap       = os.path.join(BASE_DIR_DOCS, "base_sap.xlsx")
            file_path = criar_copia_planilha(BASE_DIR_DOCS, "SIMULADOR_BALANÃ‡A_LIMPO_2.xlsx", self.add_log)

            df_base_fisica, df_base_familia_local = carregar_base_fisica(BASE_DIR_DOCS, log=self.add_log)

            global df_base_familia
            df_base_familia = df_base_familia_local

            with pd.ExcelFile(file_path) as xl:
                df_sku = xl.parse("dado_sku")


            df_expedicao = carregar_base_expedicao_csv(BASE_DIR_AUD, log=self.add_log)
            df_sap = pd.read_excel(path_base_sap, sheet_name="Sheet1")
            df_sobrepeso_real = pd.read_excel(path_base_sobrepeso, sheet_name="SOBREPESO")
            df_sobrepeso_real["DataHora"] = pd.to_datetime(df_sobrepeso_real["DataHora"])

            df_remessa = obter_dados_remessa(remessa, df_expedicao, log_callback=self.add_log)
            if df_remessa.empty:
                disponiveis = sorted(df_expedicao["REMESSA"].dropna().unique().tolist())
                self.log_callback_completo(f"âŒ Remessa {remessa} nÃ£o encontrada. Remessas disponÃ­veis: {disponiveis[-10:]}")
                messagebox.showwarning("Remessa nÃ£o encontrada", f"A remessa {remessa} nÃ£o foi localizada.")
                return

            self.progress_bar.set(0.5)

            resultado = calcular_peso_final(
                remessa, peso_vazio, qtd_paletes,
                df_remessa, df_sku, df_sap, df_sobrepeso_real, df_base_fisica,
                self.log_callback_completo,   
            )

            if not resultado:
                self.log_callback_completo("Falha no cÃ¡lculo. Verifique os dados inseridos.")
                messagebox.showwarning("Aviso", "CÃ¡lculo nÃ£o pÃ´de ser realizado.")
                return

            peso_base, sp_total, peso_com_sp, peso_final, media_sp, itens_detalhados = resultado
            mask_qtd = _match_remessa_series(df_expedicao['REMESSA'], remessa)
            
            dados = {
                'remessa': remessa,
                'qtd_skus': df_expedicao.loc[mask_qtd, 'ITEM'].nunique(),
                'placa': self.placa.get(),
                'turno': self.turno.get(),
                'peso_vazio': peso_vazio,
                'peso_base': peso_base,
                'sp_total': sp_total,
                'peso_com_sp': peso_com_sp,
                'peso_total_final': peso_final,
                'media_sp': media_sp,
                'qtd_paletes': qtd_paletes
            }

            df_fracao_vazio = pd.DataFrame(columns=['chave_pallete', 'qtd'])

            preencher_formulario_com_openpyxl(
                file_path, dados, itens_detalhados, self.add_log, df_sku, df_remessa, df_fracao_vazio
            )

            self.progress_bar.set(0.7)
            self.log_callback_completo("Exportando PDF...")
            pdf_path = exportar_pdf_com_comtypes(file_path, "FORMULARIO", nome_remessa=remessa, log_callback=self.add_log)
            self.log_callback_completo(f"PDF exportado com sucesso: {pdf_path}")

            self.progress_bar.set(0.85)
            self.log_callback_completo("Gerando relatÃ³rio de divergÃªncia em PDF...")

            # relatorio_path = gerar_relatorio_diferenca(
            #     remessa_num=remessa,
            #     peso_final_balanÃ§a=peso_balanca,
            #     peso_veiculo_vazio=peso_vazio,
            #     df_remessa=df_remessa,
            #     df_sku=df_sku,
            #     peso_estimado_total=peso_com_sp,
            #     pasta_excel=BASE_DIR_DOCS,
            #     log_callback=self.log_callback_completo
            # )
            # RelatÃ³rio de divergÃªncia desativado temporariamente
            relatorio_path = None
            self.log_callback_completo("RelatÃ³rio de divergÃªncia desconsiderado temporariamente")

            try:
                self.log_callback_completo("Iniciando impressÃ£o do relatÃ³rio")
                print_pdf(pdf_path, log_callback=self.log_callback_completo)
                print_pdf(relatorio_path, log_callback=self.log_callback_completo)
                self.log_callback_completo("RelatÃ³rio impresso com sucesso")
                self.progress_bar.set(0.95)
            except:
                self.log_callback_completo("ImpressÃ£o do RelatÃ³rio com Erro")

            try:
                enviar_email_com_log_e_pdf(
                    pdf_path,
                    remessa,
                    log_callback=self.log_callback_completo,
                    log_geral=self.log_tecnico
                )
                self.log_callback_completo("Envio com sucesso para o e-mail de tratativa")
                self.progress_bar.set(1)
                self.add_log("âœ… Processamento concluÃ­do com sucesso!")
            except:
                self.log_callback_completo("Erro ao enviar para o e-mail de tratativa")

            try:
                time.sleep(1)
                os.remove(file_path)
                self.log_callback_completo(f"CÃ³pia temporÃ¡ria removida: {file_path}")
            except Exception as e:
                self.log_callback_completo(f"Erro ao remover a cÃ³pia temporÃ¡ria: {e}")

            messagebox.showinfo(
                "Sucesso",
            )

        except Exception as e:
            tb = traceback.format_exc()
            self.log_callback_completo(f"Erro: {e!r}")
            self.log_callback_completo(tb)
            messagebox.showerror("Erro", f"Erro ao processar: {e!r}")



if __name__ == "__main__":
    app = App()
    app.mainloop()
