import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from datetime import datetime
import customtkinter as ctk
from tkinter import messagebox, StringVar
import win32com.client as win32
import comtypes.client
from pathlib import Path
import winreg
from shutil import copyfile
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from collections import defaultdict
import numpy as np
import os
import threading
import glob
import subprocess
from collections import defaultdict
import yagmail
import time
import gc

log_geral = []
 
def encontrar_pasta_onedrive_empresa():
    user_dir = os.environ["USERPROFILE"]
    possiveis = os.listdir(user_dir)
    for nome in possiveis:
        if "DIAS BRANCO" in nome.upper():
            caminho_completo = os.path.join(user_dir, nome)
            if os.path.isdir(caminho_completo) and "Gest√£o de Estoque - Documentos" in os.listdir(caminho_completo):
                return os.path.join(caminho_completo, "Gest√£o de Estoque - Documentos")
    return None

fonte_dir = encontrar_pasta_onedrive_empresa()
if not fonte_dir:
    raise FileNotFoundError("N√£o foi poss√≠vel localizar a pasta sincronizada do SharePoint via OneDrive.")

caminho_base_fisica = os.path.join(fonte_dir, "SIMULADOR_BALAN√áA_LIMPO_2.xlsx")
df_base_fisica = pd.read_excel(caminho_base_fisica, sheet_name="BASE FISICA")
df_base_familia = pd.read_excel(caminho_base_fisica, 'BASE_FAMILIA')

def criar_copia_planilha(fonte_dir, nome_arquivo, log_callback):
    try:
        origem = os.path.join(fonte_dir, nome_arquivo)
        destino_dir = os.path.join(os.environ['USERPROFILE'], 'Downloads')
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        nome_copia = f"copia_temp_{timestamp}_{nome_arquivo}"
        destino = os.path.join(destino_dir, nome_copia)
        copyfile(origem, destino)
        log_callback(f"C√≥pia criada com sucesso: {destino}")
        return destino
    except Exception as e:
        log_callback(f"Erro ao criar c√≥pia da planilha: {e}")
        raise

def enviar_email_com_log_e_pdf(caminho_pdf, remessa, log_callback=None, log_geral=None):
    import yagmail
    email_remetente = 'mdiasbrancoautomacao@gmail.com'
    token = 'secwygmzlibyxhhh'  
    email_destino = 'douglas.lins2@mdiasbranco.com.br'

    try:
        corpo_email = "\n".join(log_geral or ["(Sem logs t√©cnicos dispon√≠veis)"])
        yag = yagmail.SMTP(user=email_remetente, password=token)
        assunto = f"üì¶ Simulador Sobrepeso - Remessa {remessa}"
        yag.send(
            to=email_destino,
            subject=assunto,
            contents=corpo_email,
            attachments=[caminho_pdf]
        )
        if log_callback:
            log_callback(f"E-mail enviado com sucesso para {email_destino}")
    except Exception as e:
        if log_callback:
            log_callback(f"Erro ao enviar e-mail: {e}")

def print_pdf(file_path, impressora="VITLOG01A01", sumatra_path="C:\\Program Files\\SumatraPDF\\SumatraPDF.exe", log_callback=None):
    import subprocess
    args = [sumatra_path, "-print-to", impressora, "-silent", file_path]
    try:
        result = subprocess.run(args, check=True, capture_output=True, text=True)
        if log_callback:
            log_callback(f"Arquivo impresso com sucesso: {file_path}")
            if result.stdout:
                log_callback(f"stdout: {result.stdout}")
            if result.stderr:
                log_callback(f"stderr: {result.stderr}")
    except subprocess.CalledProcessError as e:
        if log_callback:
            log_callback(f"Erro ao imprimir {file_path}: {e}")
            log_callback(f"Output: {e.output}")
            log_callback(f"Stderr: {e.stderr}")

def integrar_itens_detalhados(df_remessa, df_sap, df_sobrepeso_real, df_sku, df_base_fisica, log_callback):
    itens_detalhados = []

    for _, row in df_remessa.iterrows():
        sku = row['ITEM']
        chave_pallet = row.get('CHAVE_PALETE', None)
        qtd = pd.to_numeric(row.get('QUANTIDADE', 0), errors='coerce')
        sp = 0.0
        origem = 'fixo'
        ajuste_sp = 0.0

        peso_base_liq = 0
        peso_caixa_liq = pd.to_numeric(
            df_sku[df_sku['COD_PRODUTO'] == sku].iloc[0]['QTDE_PESO_LIQ'],
            errors='coerce'
        ) if not df_sku[df_sku['COD_PRODUTO'] == sku].empty else 0

        if pd.notna(qtd) and pd.notna(peso_caixa_liq):
            peso_base_liq = float(qtd) * float(peso_caixa_liq)

        if pd.notna(chave_pallet) and chave_pallet in df_sap['Chave Pallet'].values:
            try:
                lote_info = df_sap[df_sap['Chave Pallet'] == chave_pallet].iloc[0]
                lote = lote_info['Lote']
                data_producao = lote_info['Data de produ√ß√£o']
                hora_inicio = f"{lote_info['Hora de cria√ß√£o'].hour:02d}:00:00"
                hora_fim = f"{lote_info['Hora de modifica√ß√£o'].hour:02d}:00:00"
                linha_coluna = "L" + lote[-3:] if isinstance(lote, str) and len(lote) >= 2 else "LB00"
                if linha_coluna in ['LB06', 'LB07']:
                    linha_coluna = 'LB06/07'
                log_callback(f"Linha produzida ajustada: {linha_coluna}")

                df_sp_filtro = df_sobrepeso_real[
                    (df_sobrepeso_real['DataHora'] >= pd.to_datetime(f"{data_producao} {hora_inicio}")) &
                    (df_sobrepeso_real['DataHora'] <= pd.to_datetime(f"{data_producao} {hora_fim}"))
                ]

                if linha_coluna in df_sp_filtro.columns:
                    sp_valores = df_sp_filtro[linha_coluna].fillna(0)
                    if not sp_valores.empty:
                        media_sp = sp_valores.mean() / 100
                        sp = media_sp
                        origem = 'real'
                        ajuste_sp = peso_base_liq * sp

            except Exception as e:
                log_callback(f"Erro ao calcular SP real para pallet {chave_pallet}: {e}")

        if origem == 'fixo':
            sp, ajuste_sp = calculo_sobrepeso_fixo(sku, df_base_fisica, peso_base_liq, log_callback)

        itens_detalhados.append({
            'sku': sku,
            'chave_pallet': chave_pallet if pd.notna(chave_pallet) else 'N/A',
            'sp': round(sp, 4),
            'ajuste_sp': round(ajuste_sp, 2),
            'origem': origem
        })
    log_callback(f"Total de itens detalhados integrados: {len(itens_detalhados)} de {len(df_remessa)} linhas da remessa.")
    return itens_detalhados


def calculo_sobrepeso_fixo(sku, df_base_fisica, peso_base_liq, log_callback):
    try:
        sp_row = df_base_fisica[df_base_fisica['C√ìDIGO PRODUTO'] == sku]
        if not sp_row.empty:
            sp_fixo = pd.to_numeric(sp_row.iloc[0]['SOBRE PESO'], errors='coerce') / 100
            peso_base_liq = pd.to_numeric(peso_base_liq, errors='coerce')

            if isinstance(peso_base_liq, pd.Series):
                peso_base_liq = peso_base_liq.iloc[0]

            if pd.notna(peso_base_liq) and pd.notna(sp_fixo):
                ajuste = float(peso_base_liq) * float(sp_fixo)
            else:
                ajuste = 0.0
            return sp_fixo, ajuste
        else:
            log_callback(f"Nenhum sobrepeso fixo encontrado para SKU {sku}.")
            return 0.0, 0.0
    except Exception as e:
        log_callback(f"Erro ao buscar sobrepeso fixo para SKU {sku}: {e}")
        return 0.0, 0.0


def buscar_chave_por_endereco(endereco, df_estoque_sep):
    if pd.isna(endereco) or endereco == '':
        return None
    df_filtrado = df_estoque_sep[df_estoque_sep['endereco'] == endereco]
    if df_filtrado.empty:
        return None
    df_ordenado = df_filtrado.sort_values(by='Criado', ascending=False)
    return df_ordenado.iloc[0]['chave_pallete']


def processar_sobrepeso(chave_pallet, sku, peso_base_liq, df_sap, df_sobrepeso_real, df_base_fisica, log_callback):
    sp = 0
    origem_sp = 'n√£o encontrado'
    ajuste_sp = 0

    if pd.notna(chave_pallet) and chave_pallet in df_sap['Chave Pallet'].values:
        pallet_info = df_sap[df_sap['Chave Pallet'] == chave_pallet].iloc[0]
        lote = pallet_info['Lote']
        data_producao = pallet_info['Data de produ√ß√£o']
        hora_inicio = f"{pallet_info['Hora de cria√ß√£o'].hour:02d}:00:00"
        hora_fim = f"{pallet_info['Hora de modifica√ß√£o'].hour:02d}:00:00"
        linha_coluna = "L" + lote[-3:]
        if linha_coluna in ['LB06', 'LB07']:
            linha_coluna = 'LB06/07'

        log_callback(f"Processando pallet {chave_pallet} para SKU {sku}.")

        df_sp_filtro = df_sobrepeso_real[
            (df_sobrepeso_real['DataHora'] >= pd.to_datetime(f"{data_producao} {hora_inicio}")) &
            (df_sobrepeso_real['DataHora'] <= pd.to_datetime(f"{data_producao} {hora_fim}"))
        ]

        if linha_coluna in df_sp_filtro.columns:
            sp_valores = df_sp_filtro[linha_coluna].fillna(0)
            if not sp_valores.empty:
                media_sp = sp_valores.mean() / 100
                log_callback(f"[{linha_coluna}] M√©dia SP: {media_sp:.4f}")

                peso_base_liq = pd.to_numeric(peso_base_liq, errors='coerce')
                sp = pd.to_numeric(media_sp, errors='coerce')

                if isinstance(peso_base_liq, pd.Series):
                    peso_base_liq = peso_base_liq.iloc[0]
                if isinstance(sp, pd.Series):
                    sp = sp.iloc[0]

                if pd.notna(peso_base_liq) and pd.notna(sp) and float(sp)>0:
                    ajuste_sp = float(peso_base_liq) * float(sp)
                    origem_sp = 'real'
                else:
                    sp=0
                    ajuste_sp = 0.0
                    origem_sp='fixo'
                    log_callback(f"Erro: peso_base_liq ou sp inv√°lido (verificar se h√° a data ou chave pallete na base) para SKU {sku}. Ajuste SP definido como 0.")
        else:
            log_callback(f"Coluna {linha_coluna} n√£o encontrada na base de sobrepeso.")

    if sp == 0:
        sp_valor, ajuste_fixo = calculo_sobrepeso_fixo(sku, df_base_fisica, peso_base_liq, log_callback)
        if sp_valor != 0:
            sp = sp_valor
            origem_sp = 'fixo'
            ajuste_sp = ajuste_fixo
            if origem_sp == 'fixo':
                log_callback(f"SOBREPESO FIXO aplicado para SKU {sku}: {sp:.4f} ‚Üí Ajuste: {ajuste_sp:.2f} kg")
        else:
            log_callback(f"Nenhum sobrepeso encontrado para SKU {sku}.")

    return sp, origem_sp, ajuste_sp

def processar_sobrepeso_fixo_basico(sku, qtd, df_sku, df_base_fisica, peso_base_total, peso_base_total_liq, sp_total, itens_detalhados, log_callback):
    try:
        df_sku_filtrado = df_sku[df_sku['COD_PRODUTO'] == sku]
        if df_sku_filtrado.empty:
            log_callback(f"SKU {sku} n√£o encontrado na base SKU.")
            return peso_base_total, peso_base_total_liq, sp_total

        peso_por_caixa_bruto = pd.to_numeric(df_sku_filtrado.iloc[0]['QTDE_PESO_BRU'], errors='coerce')
        peso_por_caixa_liquido = pd.to_numeric(df_sku_filtrado.iloc[0]['QTDE_PESO_LIQ'], errors='coerce')
        qtd_validada = pd.to_numeric(qtd, errors='coerce')

        peso_base = peso_base_liq = 0.0

        if pd.notna(peso_por_caixa_bruto) and pd.notna(qtd_validada):
            peso_base = float(peso_por_caixa_bruto) * float(qtd_validada)
            peso_base_total += peso_base

        if pd.notna(peso_por_caixa_liquido) and pd.notna(qtd_validada):
            peso_base_liq = float(peso_por_caixa_liquido) * float(qtd_validada)
            peso_base_total_liq += peso_base_liq

        sp_valor, ajuste_sp = calculo_sobrepeso_fixo(sku, df_base_fisica, peso_base_liq, log_callback)
        sp_total += ajuste_sp

        log_callback(f"Aplicando SP fixo para SKU {sku} ‚Üí ajuste: {ajuste_sp:.2f}kg")

        itens_detalhados.append({
            'sku': sku,
            'chave_pallet': 'N/A',
            'sp': round(sp_valor, 4),
            'ajuste_sp': round(ajuste_sp, 2),
            'origem': 'fixo'
        })

        return peso_base_total, peso_base_total_liq, sp_total

    except Exception as e:
        log_callback(f"Erro ao aplicar SP fixo para SKU {sku}: {e}")
        return peso_base_total, peso_base_total_liq, sp_total
    
def calcular_peso_teorico_receb_ext(sku, qtd, df_sku):
    try:
        df_sku_filtrado = df_sku[df_sku['COD_PRODUTO'] == sku]
        if df_sku_filtrado.empty:
            return 0.0

        peso_liq = pd.to_numeric(df_sku_filtrado.iloc[0]['QTDE_PESO_LIQ'], errors='coerce')
        qtd = pd.to_numeric(qtd, errors='coerce')

        if pd.notna(peso_liq) and pd.notna(qtd):
            return float(peso_liq) * float(qtd)
        return 0.0
    except:
        return 0.0

def calcular_peso_receb_externo(chave, sku, row, df_externo_peso, df_estoque_sep, df_sku,
                                peso_base_total, peso_base_total_liq, itens_detalhados, log_callback):
    try:
        log_callback(f"üîé Verificando recebimento externo para: {chave}")
        if not isinstance(chave, str) or not chave.endswith("_1"):
            return None

        row_exter = df_externo_peso[df_externo_peso['chave_pallete'] == chave]
        if row_exter.empty:
            log_callback("N√£o encontrado em df_externo_peso.")
            return None

        peso_real_total = pd.to_numeric(row_exter.iloc[0]['peso'], errors='coerce')
        qtd_registrada = pd.to_numeric(row_exter.iloc[0]['quantidade'], errors='coerce')
        sku_externo = pd.to_numeric(row_exter.iloc[0]['SKU'], errors='coerce')

        if pd.isna(peso_real_total) or pd.isna(qtd_registrada) or qtd_registrada == 0:
            log_callback("Peso real inv√°lido ou dados incompletos para c√°lculo proporcional.")
            return None

        qtd_expedida = None
        for campo in ['quantidade', 'QUANTIDADE', 'qtd']:
            if campo in row.index:
                qtd_expedida = pd.to_numeric(row[campo], errors='coerce')
                break

        if pd.isna(qtd_expedida):
            log_callback("Quantidade expedida inv√°lida ou ausente para recebimento externo.")
            return None


        if qtd_expedida > qtd_registrada:
            log_callback(f"Quantidade expedida ({qtd_expedida}) maior do que registrada ({qtd_registrada}) no pallet externo {chave}. Usando mesmo assim.")

        peso_proporcional = peso_real_total * (qtd_expedida / qtd_registrada)
        df_sku_filtrado = df_sku[df_sku['COD_PRODUTO'] == sku]
        if df_sku_filtrado.empty:
            log_callback(f"SKU {sku} n√£o encontrado na base SKU.")
            return None

        peso_unit_liq = pd.to_numeric(df_sku_filtrado.iloc[0]['QTDE_PESO_LIQ'], errors='coerce')
        peso_unit_bruto = pd.to_numeric(df_sku_filtrado.iloc[0]['QTDE_PESO_BRU'], errors='coerce')

        peso_liq_teorico = float(peso_unit_liq) * qtd_expedida if pd.notna(peso_unit_liq) else 0
        peso_bruto_teorico = float(peso_unit_bruto) * qtd_expedida if pd.notna(peso_unit_bruto) else 0

        if peso_liq_teorico > 0:
            sp_real = (peso_proporcional / peso_liq_teorico) - 1
        else:
            sp_real = 0

        ajuste_sp = peso_proporcional - peso_liq_teorico
        peso_base_total += peso_bruto_teorico
        peso_base_total_liq += peso_liq_teorico

        log_callback(f"Receb. Externo: Peso real {peso_proporcional:.2f}kg | Peso te√≥rico {peso_liq_teorico:.2f}kg ‚Üí SP: {sp_real:.4f}")

        itens_detalhados.append({
            'sku': sku,
            'chave_pallet': chave,
            'sp': round(sp_real, 4),
            'ajuste_sp': round(ajuste_sp, 2),
            'origem': 'receb_ext'
        })

        return peso_base_total, peso_base_total_liq

    except Exception as e:
        log_callback(f"Erro em calcular_peso_receb_externo para chave {chave}: {e}")
        return None

def calcular_peso_final(remessa_num, peso_veiculo_vazio, qtd_paletes, df_expedicao, df_sku, df_sap, df_sobrepeso_real,
                         df_base_fisica, df_frac, df_estoque_sep, df_externo_peso,log_callback):
    from collections import defaultdict
    try:
        remessa_num = int(remessa_num)
    except ValueError:
        log_callback("Remessa inv√°lida.")
        return None

    df_remessa = df_expedicao[df_expedicao['REMESSA'] == remessa_num]
    if df_remessa.empty:
        log_callback("Remessa n√£o encontrada em data_exp.")
        return None

    contador_exp = defaultdict(int)
    for _, row in df_remessa.iterrows():
        key = (row['ITEM'], row['QUANTIDADE'])
        contador_exp[key] += 1

    peso_base_total = 0
    peso_base_total_liq = 0
    sp_total = 0
    itens_detalhados = []
    chaves_pallet_processadas = set()

    df_expedicao_sem_pallet = df_remessa[df_remessa['CHAVE_PALETE'].isna()]
    pacotes_expedicao = df_expedicao_sem_pallet.groupby(['ITEM', 'QUANTIDADE']).size().reset_index(name='count')

    for _, pacote in pacotes_expedicao.iterrows():
        sku = pacote['ITEM']
        qtd = pd.to_numeric(pacote['QUANTIDADE'], errors='coerce')
        count_expedicao = int(pacote['count'])

        df_frac_match = df_frac[(df_frac['remessa'] == remessa_num) & (df_frac['sku'] == sku)]
        df_frac_match = df_frac_match[pd.to_numeric(df_frac_match['qtd'], errors='coerce') == qtd]
        df_frac_same_remessa = df_frac[df_frac['remessa'] == remessa_num]
        skus_frac = df_frac_same_remessa['sku'].unique()
        if len(skus_frac) > 0 and sku not in skus_frac:
            log_callback(f"Remessa {remessa_num} encontrada na base FRACAO, mas com SKUs diferentes do esperado: {skus_frac.tolist()} ‚Äî SKU da linha atual: {sku}. Pode haver diverg√™ncia nos dados.")

        count_fracao = len(df_frac_match)

        qtd_real = min(count_expedicao, count_fracao)
        qtd_fixo = count_expedicao - qtd_real

        log_callback(f"Processando SKU {sku} QTD {qtd}: {qtd_real} via real, {qtd_fixo} via fixo")

        for idx in range(qtd_real):
            frac_row = df_frac_match.iloc[idx]
            chave_frac = frac_row.get('chave_pallete')
            if pd.isna(chave_frac) or chave_frac == '':
                endereco = frac_row.get('endereco')
                chave_frac = buscar_chave_por_endereco(endereco, df_estoque_sep)

            if chave_frac in chaves_pallet_processadas or pd.isna(chave_frac):
                continue
            chaves_pallet_processadas.add(chave_frac)
            resultado_ext = calcular_peso_receb_externo(
            chave_frac, sku, frac_row, df_externo_peso, df_estoque_sep,df_sku,
            peso_base_total, peso_base_total_liq, itens_detalhados, log_callback
            )
            if resultado_ext:
                peso_base_total, peso_base_total_liq = resultado_ext
                continue
            pallet_info = df_sap[df_sap['Chave Pallet'] == chave_frac]
            if pallet_info.empty:
                log_callback(f"Chave pallet {chave_frac} da FRACAO n√£o encontrada na base SAP.")
                continue

            pallet_info = pallet_info.iloc[0]
            df_sku_filtrado = df_sku[df_sku['COD_PRODUTO'] == sku]
            if df_sku_filtrado.empty:
                log_callback(f"SKU {sku} n√£o encontrado na base SKU.")
                continue

            peso_liq = pd.to_numeric(df_sku_filtrado.iloc[0]['QTDE_PESO_LIQ'], errors='coerce')
            qtd_validada = pd.to_numeric(qtd, errors='coerce')

            peso_base_liq = float(peso_liq) * float(qtd_validada) if pd.notna(peso_liq) and pd.notna(qtd_validada) else 0
            peso_base_total_liq += peso_base_liq

            peso_bruto = pd.to_numeric(df_sku_filtrado.iloc[0]['QTDE_PESO_BRU'], errors='coerce')
            peso_base = float(peso_bruto) * float(qtd_validada) if pd.notna(peso_bruto) and pd.notna(qtd_validada) else 0
            peso_base_total += peso_base

            sp, origem_sp, ajuste_sp = processar_sobrepeso(chave_frac, sku, peso_base_liq, df_sap, df_sobrepeso_real, df_base_fisica, log_callback)
            sp_total += ajuste_sp

            itens_detalhados.append({
                'sku': sku,
                'chave_pallet': chave_frac,
                'sp': round(sp, 4),
                'ajuste_sp': round(ajuste_sp, 2),
                'origem': origem_sp
            })

        for _ in range(qtd_fixo):
            peso_base_total, peso_base_total_liq, sp_total = processar_sobrepeso_fixo_basico(
                sku, qtd, df_sku, df_base_fisica,
                peso_base_total, peso_base_total_liq, sp_total,
                itens_detalhados, log_callback
            )

    df_expedicao_com_pallet = df_remessa[df_remessa['CHAVE_PALETE'].notna()]
    for _, row in df_expedicao_com_pallet.iterrows():
        sku = row['ITEM']
        chave_pallet = row['CHAVE_PALETE']
        qtd_caixas = pd.to_numeric(row['QUANTIDADE'], errors='coerce')
        if chave_pallet in chaves_pallet_processadas or pd.isna(qtd_caixas):
            continue
        chaves_pallet_processadas.add(chave_pallet)
        resultado_ext = calcular_peso_receb_externo(
            chave_pallet, sku, row, df_externo_peso, df_estoque_sep,df_sku,
            peso_base_total, peso_base_total_liq, itens_detalhados, log_callback
        )
        if resultado_ext:
            peso_base_total, peso_base_total_liq = resultado_ext
            continue
        df_sku_filtrado = df_sku[df_sku['COD_PRODUTO'] == sku]
        if df_sku_filtrado.empty:
            log_callback(f"SKU {sku} n√£o encontrado na base SKU.")
            continue

        peso_bruto = pd.to_numeric(df_sku_filtrado.iloc[0]['QTDE_PESO_BRU'], errors='coerce')
        peso_liq = pd.to_numeric(df_sku_filtrado.iloc[0]['QTDE_PESO_LIQ'], errors='coerce')

        peso_base = float(peso_bruto) * float(qtd_caixas) if pd.notna(peso_bruto) else 0
        peso_base_liq = float(peso_liq) * float(qtd_caixas) if pd.notna(peso_liq) else 0

        peso_base_total += peso_base
        peso_base_total_liq += peso_base_liq

        sp, origem_sp, ajuste_sp = processar_sobrepeso(chave_pallet, sku, peso_base_liq, df_sap, df_sobrepeso_real, df_base_fisica, log_callback)
        sp_total += ajuste_sp

        itens_detalhados.append({
            'sku': sku,
            'chave_pallet': chave_pallet,
            'sp': round(sp, 4),
            'ajuste_sp': round(ajuste_sp, 2),
            'origem': origem_sp
        })

    peso_com_sobrepeso = peso_base_total + sp_total
    log_callback(f"Peso com sobrepeso: {peso_com_sobrepeso:.2f} kg")

    peso_total_com_paletes = peso_com_sobrepeso + (qtd_paletes * 23) + peso_veiculo_vazio
    log_callback(f"Peso total com paletes ({qtd_paletes} x 26kg): {peso_total_com_paletes:.2f} kg")

    itens_detalhados_integrados = integrar_itens_detalhados(
        df_remessa, df_sap, df_sobrepeso_real, df_sku, df_base_fisica, log_callback
    )

    media_sp_geral = (sum(item['sp'] for item in itens_detalhados_integrados) / len(itens_detalhados_integrados)) if itens_detalhados_integrados else 0.0
    log_callback(f"M√©dia geral de sobrepeso (entre {len(itens_detalhados_integrados)} itens): {media_sp_geral:.4f}")

    qtd_real = sum(1 for i in itens_detalhados_integrados if i['origem'] == 'real')
    qtd_ext  = sum(1 for i in itens_detalhados_integrados if i['origem'] == 'receb_ext')
    qtd_fixo = sum(1 for i in itens_detalhados_integrados if i['origem'] == 'fixo')

    log_callback(f"üì¶ Total de itens processados: {len(itens_detalhados)}")
    log_callback(f"‚îú‚îÄ‚îÄ reais: {sum(1 for i in itens_detalhados if i.get('origem') in ['real', 'receb_ext'])}")
    log_callback(f"‚îú‚îÄ‚îÄ receb_ext: {sum(1 for i in itens_detalhados if i.get('origem') == 'receb_ext')}")
    log_callback(f"‚îî‚îÄ‚îÄ fixos: {sum(1 for i in itens_detalhados if i.get('origem') == 'fixo')}")


    return peso_base_total, sp_total, peso_com_sobrepeso, peso_total_com_paletes, media_sp_geral, itens_detalhados_integrados

def calcular_limites_sobrepeso_por_quantidade(dados, itens_detalhados, df_base_familia, df_sobrepeso_tabela, df_sku, df_remessa, df_fracao, log_callback):
    total_quantidade = 0
    quantidade_com_sp_real = 0
    ponderador_pos = 0
    ponderador_neg = 0
    familia_detectada = "MIX"
    agrupado_por_sku = defaultdict(list)
    for item in itens_detalhados:
        agrupado_por_sku[item['sku']].append(item)

    for sku, itens in agrupado_por_sku.items():
        qtd_total = 0
        qtd_real = 0
        ponderador_pos_local = 0
        ponderador_neg_local = 0 
        for item in itens:
            origem = item.get('origem', 'fixo')
            sp = item.get('sp', 0)
            chave = item.get('chave_pallet', '')
            qtd = 0
            if chave in df_fracao['chave_pallete'].values:
                qtd = pd.to_numeric(df_fracao[df_fracao['chave_pallete'] == chave]['qtd'], errors='coerce').sum()
            else:
                qtd = pd.to_numeric(df_remessa[df_remessa['ITEM'] == sku]['QUANTIDADE'], errors='coerce').sum()

            qtd_total += qtd

            if origem in ['real', 'receb_ext']:
                qtd_real += qtd
                if sp > 0:
                    ponderador_pos_local += sp * qtd
                elif sp < 0:
                    ponderador_neg_local += abs(sp) * qtd

        total_quantidade += qtd_total
        quantidade_com_sp_real += qtd_real
        ponderador_pos += ponderador_pos_local
        ponderador_neg += ponderador_neg_local

    proporcao_sp_real = quantidade_com_sp_real / total_quantidade if total_quantidade > 0 else 0
    log_callback(f"Total de quantidade: {total_quantidade}, com SP Real: {quantidade_com_sp_real}, propor√ß√£o: {proporcao_sp_real:.2%}")

    if proporcao_sp_real >= 0.5:
        if quantidade_com_sp_real > 0:
            media_positiva = ponderador_pos / quantidade_com_sp_real if ponderador_pos > 0 else 0.02
            media_negativa = ponderador_neg / quantidade_com_sp_real if ponderador_neg > 0 else 0.01
        else:
            media_positiva = 0.02
            media_negativa = 0.01
        log_callback("Mais de 50% da quantidade com SP Real. Usando m√©dias ponderadas:")
        log_callback(f"Sobrepeso para mais (real): {media_positiva:.4f}")
        log_callback(f"Sobrepeso para menos (real): {media_negativa:.4f}")
    else:
        log_callback("Menos de 50% da quantidade com SP Real. Usando tabela de sobrepeso f√≠sico.")
        familias = set()
        for sku in agrupado_por_sku:
            familia_series = df_base_familia.loc[df_base_familia['C√ìD'] == sku, 'FAMILIA 2']
            if not familia_series.empty:
                familia_str=str(familia_series.iloc[0])
                familias.add(familia_str)

        if len(familias) == 1:
            familia_str = list(familias)[0]
            if 'BISCOITO' in familia_str.upper():
                familia_detectada = "BISCOITO"
                row = df_sobrepeso_tabela.loc[df_sobrepeso_tabela.index.str.contains("BISCOITO", case=False)]
            elif 'MASSA' in familia_str.upper():
                familia_detectada = 'MASSA'
                row = df_sobrepeso_tabela.loc[df_sobrepeso_tabela.index.str.contains("MASSA", case=False)]
            else:
                familia_detectada = "MIX"
                row = df_sobrepeso_tabela.loc[df_sobrepeso_tabela.index.str.contains("MIX", case=False)]
        else:
            familia_detectada = "MIX"
            row = df_sobrepeso_tabela.loc[df_sobrepeso_tabela.index.str.contains("MIX", case=False)]

        if row.empty:
            log_callback("Fam√≠lia n√£o encontrada na tabela. Usando MIX como fallback.")
            row = df_sobrepeso_tabela.loc[df_sobrepeso_tabela.index.str.contains("MIX", case=False)]

        media_positiva = row['(+)'].values[0]
        media_negativa = row['(-)'].values[0]

        log_callback(f"Sobrepeso para mais (f√≠sico): {media_positiva:.4f}")
        log_callback(f"Sobrepeso para menos (f√≠sico): {media_negativa:.4f}")

    return media_positiva, media_negativa, proporcao_sp_real, familia_detectada

def preencher_formulario_com_openpyxl(path_copia, dados, itens_detalhados, log_callback,df_sku, df_remessa, df_fracao):
    try:
        dados_tabela = {
        '(+)': [0.02, 0.005, 0.04],
        '(-)': [0.01, 0.01, 0.01]
        }
        index = ['CARGA COM MIX', 'EXCLUSIVO MASSAS', 'EXCLUSIVO BISCOITOS']
        df_sobrepeso_tabela = pd.DataFrame(dados_tabela, index=index)
        df_base_familia = pd.read_excel(caminho_base_fisica, 'BASE_FAMILIA')
        sp_pos, sp_neg, proporcao_sp_real, familia_detectada = calcular_limites_sobrepeso_por_quantidade(
            dados, itens_detalhados, df_base_familia, df_sobrepeso_tabela, df_sku, df_remessa, df_fracao, log_callback
        )
        wb = load_workbook(path_copia)
        ws = wb["FORMULARIO"]

        log_callback("Preenchendo cabe√ßalhos principais com openpyxl...")
        ws["A16"] = f"Sobrepeso para (+): {sp_pos*100:.2f}%"
        ws["A18"] = f"Sobrepeso para (-): {sp_neg*100:.2f}%"
        ws["D7"] = f"{proporcao_sp_real*100:.2f}% x {(1 - proporcao_sp_real)*100:.2f}%"
        try:
            ws["B5"] = str(familia_detectada) if familia_detectada is not None else "MIX"
        except Exception as e:
            log_callback(f"Erro ao escrever fam√≠lia detectada: {e}")
            ws["B5"] = "MIX" 
        ws["B4"] = dados['remessa']
        ws["B6"] = dados['qtd_skus']
        ws["B7"] = dados['placa']
        ws["B8"] = dados['turno']
        ws["B9"] = dados['peso_vazio']
        ws["B10"] = dados['peso_base']
        ws["B11"] = dados['sp_total']
        ws["B12"] = dados['peso_com_sp']
        ws["B13"] = dados['peso_total_final']
        ws["B14"] = dados['media_sp']
        ws["B16"] = dados['peso_total_final'] * (1 + sp_pos)
        ws["B17"] = dados['peso_total_final']
        ws["B18"] = dados['peso_total_final'] * (1 - sp_neg)
        ws["D4"] = dados['qtd_paletes']
        ws["D9"] = dados['qtd_paletes'] * 26

        linha_inicio = 12
        linha_fim = 46
        max_itens = linha_fim - linha_inicio + 1

        log_callback("Preenchendo SKUs e sobrepesos...")
        itens_real = [item for item in itens_detalhados if item['origem'] == 'real']
        itens_fixo = [item for item in itens_detalhados if item['origem'] == 'fixo']
        itens_nao_encontrado = [item for item in itens_detalhados if item['origem'] == 'n√£o encontrado']

        itens_ordenados = itens_real + itens_fixo + itens_nao_encontrado
        for idx, item in enumerate(itens_ordenados):
            if idx >= max_itens:
                log_callback("Limite m√°ximo de linhas atingido no formul√°rio (C12 at√© C46). Restante ser√° desconsiderado.")
                break
            linha = linha_inicio + idx
            sku_texto = f"{item['sku']} ({item['origem']})"
            ws[f"C{linha}"] = sku_texto
            ws[f"D{linha}"] = f"{item['sp']*100:.3f}"

        wb.save(path_copia)
        wb.close()
        log_callback("Formul√°rio preenchido e salvo com sucesso.")

    except Exception as e:
        log_callback(f"Erro no preenchimento: {e}")
        raise

def exportar_pdf_com_comtypes(path_xlsx, aba_nome="FORMULARIO", nome_remessa="REMESSA", log_callback=None):
    try:
        if log_callback:
            log_callback("Iniciando exporta√ß√£o via comtypes...")

        comtypes.CoInitialize()
        excel = comtypes.client.CreateObject("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(path_xlsx)

        if aba_nome not in [sheet.Name for sheet in wb.Sheets]:
            raise Exception(f"Aba '{aba_nome}' n√£o encontrada.")

        ws = wb.Worksheets(aba_nome)

        pdf_dir = os.path.join(fonte_dir, 'Relat√≥rio_Saida')
        os.makedirs(pdf_dir, exist_ok=True)
        pdf_path = os.path.join(pdf_dir, f"SOBREPESOSIMULADO - {nome_remessa}.pdf")
        if log_callback:
            log_callback(f"Tentando exportar para: {pdf_path}")
        ws.ExportAsFixedFormat(Type=0, Filename=pdf_path)
        wb.Close(SaveChanges=False)
        excel.Quit()
        del ws
        del wb
        del excel
        gc.collect()
        comtypes.CoUninitialize()
        time.sleep(5)
        if log_callback:
            log_callback(f"PDF exportado com sucesso: {pdf_path}")
        return pdf_path
    except Exception as e:
        if log_callback:
            log_callback(f"Erro ao exportar PDF: {e}")
        raise

def gerar_relatorio_diferenca(remessa_num, peso_final_balan√ßa, peso_veiculo_vazio, df_remessa, df_sku, peso_estimado_total, pasta_excel,log_callback):
    import os
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages
    import pandas as pd

    pasta_destino = os.path.join(pasta_excel, 'Analise_divergencia')
    os.makedirs(pasta_destino, exist_ok=True)

    skus = df_remessa['ITEM'].unique()
    dados_relatorio = []
    peso_base_total_liq = 0

    for sku in skus:
        qtd = df_remessa[df_remessa['ITEM'] == sku]['QUANTIDADE'].sum()
        df_sku_filtrado = df_sku[df_sku['COD_PRODUTO'] == sku]

        if df_sku_filtrado.empty:
            continue

        unidade = df_sku_filtrado.iloc[0]['DESC_UNID_MEDID']
        peso_unit_liq = df_sku_filtrado.iloc[0]['QTDE_PESO_LIQ']
        peso_total_liq = qtd * peso_unit_liq
        peso_base_total_liq += peso_total_liq

        dados_relatorio.append({
            'SKU': sku,
            'Unidade': unidade,
            'Quantidade': qtd,
            'Peso Total L√≠quido': peso_total_liq,
            'Peso Unit. L√≠quido': peso_unit_liq
        })

    df_dados = pd.DataFrame(dados_relatorio)

    diferenca_total = diferenca_total = (peso_estimado_total + peso_veiculo_vazio) - peso_final_balan√ßa
    peso_carga_real = peso_final_balan√ßa - peso_veiculo_vazio
    df_dados['% Peso'] = df_dados['Peso Total L√≠quido'] / peso_base_total_liq
    df_dados['Peso Proporcional Real'] = df_dados['% Peso'] * peso_carga_real
    df_dados['Quantidade Real Estimada'] = (df_dados['Peso Proporcional Real'] / df_dados['Peso Unit. L√≠quido']).round()
    df_dados['Diferen√ßa Estimada (kg)'] = df_dados['% Peso'] * diferenca_total
    df_dados['Unid. Estimada de Diverg√™ncia'] = df_dados['Quantidade Real Estimada'] - df_dados['Quantidade']

    nome_pdf = f"An√°lise quantitativa - {remessa_num}.pdf"
    caminho_pdf = os.path.join(pasta_destino, nome_pdf)

    fig, ax = plt.subplots(figsize=(10, 6))
    largura_barra = 0.35
    x = range(len(df_dados))

    qtd_esperada = df_dados['Quantidade']
    qtd_real = df_dados['Quantidade Real Estimada']

    ax.bar([i - largura_barra/2 for i in x], qtd_esperada, width=largura_barra, label='Quantidade Esperada')
    ax.bar([i + largura_barra/2 for i in x], qtd_real, width=largura_barra, label='Quantidade Real Estimada')

    ax.set_ylabel("Quantidade (unidades)")
    ax.set_xlabel("SKU")
    ax.set_title("Comparativo: Quantidade Esperada vs Real Estimada por SKU")
    ax.set_xticks(x)
    ax.set_xticklabels(df_dados['SKU'].astype(str))
    ax.axhline(0, color='gray', linestyle='--')
    ax.legend()

    for i in x:
        ax.text(i - largura_barra/2, qtd_esperada.iloc[i], f"{qtd_esperada.iloc[i]:.0f}", ha='center', va='bottom')
        ax.text(i + largura_barra/2, qtd_real.iloc[i], f"{qtd_real.iloc[i]:.0f}", ha='center', va='bottom')

    with PdfPages(caminho_pdf) as pdf:
        fig_tabela, ax_tabela = plt.subplots(figsize=(12, len(df_dados) * 0.5 + 3))
        ax_tabela.axis('off')
        table_data = [
            ['SKU', 'Unidade', 'Qtd. Enviada', 'Qtd. Real Estimada', 'Peso Total L√≠quido', '% do Peso', 'Diferen√ßa (kg)', 'Diverg√™ncia (unid)']
        ] + df_dados[['SKU', 'Unidade', 'Quantidade', 'Quantidade Real Estimada', 'Peso Total L√≠quido', '% Peso', 'Diferen√ßa Estimada (kg)', 'Unid. Estimada de Diverg√™ncia']].round(2).values.tolist()

        tabela = ax_tabela.table(cellText=table_data, colLabels=None, loc='center', cellLoc='center')
        tabela.auto_set_font_size(False)
        tabela.set_fontsize(10)
        tabela.scale(1, 1.5)

        titulo = f"Relat√≥rio Comparativo - Remessa {remessa_num}\nPeso estimado: {peso_estimado_total:.2f} kg | Peso balan√ßa: {peso_final_balan√ßa:.2f} kg | Peso ve√≠culo: {peso_veiculo_vazio:.2f} kg | Diferen√ßa: {diferenca_total:.2f} kg"
        fig_tabela.suptitle(titulo, fontsize=12)
        pdf.savefig(fig_tabela, bbox_inches='tight')
        pdf.savefig(fig, bbox_inches='tight')

    plt.close('all')
    log_callback(f"Relat√≥rio salvo em: {caminho_pdf}")
    return caminho_pdf

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Simulador de Sobrepeso")
        self.geometry("1000x600")

        self.log_tecnico = []
        self.log_geral = []
        self.placa = StringVar()
        self.turno = StringVar()
        self.remessa = StringVar()
        self.qtd_paletes = StringVar()
        self.peso_vazio = StringVar()
        self.peso_balanca = StringVar()
        self.log_text = []
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        formulario_frame = ctk.CTkFrame(self)
        formulario_frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        footer_label = ctk.CTkLabel(self, text="Desenvolvido por Douglas Lins - Analista de Log√≠stica", font=("Arial", 10), anchor="center")
        footer_label.grid(row=1, column=0, columnspan=2, pady=(0, 10))
        ctk.CTkLabel(formulario_frame, text="Placa:").pack(anchor="w")
        ctk.CTkEntry(formulario_frame, textvariable=self.placa).pack(fill="x")
        ctk.CTkLabel(formulario_frame, text="Turno:").pack(anchor="w")
        ctk.CTkComboBox(formulario_frame, values=["A", "B", "C"], variable=self.turno).pack(fill="x")
        ctk.CTkLabel(formulario_frame, text="Remessa:").pack(anchor="w")
        ctk.CTkEntry(formulario_frame, textvariable=self.remessa).pack(fill="x")
        ctk.CTkLabel(formulario_frame, text="Quantidade de Paletes:").pack(anchor="w")
        ctk.CTkEntry(formulario_frame, textvariable=self.qtd_paletes).pack(fill="x")
        ctk.CTkLabel(formulario_frame, text="Peso Ve√≠culo Vazio:").pack(anchor="w")
        ctk.CTkEntry(formulario_frame, textvariable=self.peso_vazio).pack(fill="x")
        ctk.CTkLabel(formulario_frame, text="Peso Final Balan√ßa:").pack(anchor="w")
        ctk.CTkEntry(formulario_frame, textvariable=self.peso_balanca).pack(fill="x")
        ctk.CTkButton(formulario_frame, text="Calcular", command=self.iniciar_processamento).pack(pady=10)
        self.progress_bar = ctk.CTkProgressBar(formulario_frame, mode="determinate")
        self.progress_bar.set(0)
        self.progress_bar.pack(pady=10, fill="x")
        self.progress_bar.pack_forget()
        logs_frame = ctk.CTkFrame(self)
        logs_frame.grid(row=0, column=1, sticky="nsew", padx=20, pady=20)
        logs_frame.grid_rowconfigure(1, weight=1)
        logs_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(logs_frame, text="üìú Hist√≥rico de Logs:").grid(row=0, column=0, sticky="w")
        self.log_display = ctk.CTkTextbox(logs_frame, wrap="word", width=400, height=400)
        self.log_display.grid(row=1, column=0, sticky="nsew")
        ctk.CTkButton(logs_frame, text="üßπ Limpar Hist√≥rico", command=self.limpar_logs).grid(row=2, column=0, pady=10, sticky="e")

    def add_log(self, msg):
        timestamp = datetime.now().strftime("%H:%M:%S")
        entrada = f"[{timestamp}] {msg}"
        self.log_text.append(entrada)
        self.log_display.configure(state="normal")
        self.log_display.insert("end", entrada + "\n")
        self.log_display.see("end")
        self.log_display.configure(state="disabled")

    def limpar_logs(self):
        self.log_text.clear()
        self.log_geral.clear()
        self.log_display.configure(state="normal")
        self.log_display.delete("1.0", "end")
        self.log_display.configure(state="disabled")
    

    def log_callback_completo(self, mensagem):
            print(mensagem)
            self.log_geral.append(mensagem)
            self.add_log(mensagem)

    def log_callback_tecnico(self, mensagem):
        timestamp = datetime.now().strftime("%H:%M:%S")
        entrada = f"[{timestamp}] {mensagem}"
        self.log_tecnico.append(entrada)

    def iniciar_processamento(self):
        self.progress_bar.pack()  
        self.progress_bar.set(0) 
        self.add_log("Iniciando c√°lculo...")
        thread = threading.Thread(target=self.processar)
        thread.start()

    def processar(self):
        self.progress_bar.set(0.1)
        path_base_sobrepeso = os.path.join(fonte_dir, "Base_sobrepeso_real.xlsx")
        path_base_expedicao = os.path.join(fonte_dir, "expedicao.xlsx")
        path_base_sap = os.path.join(fonte_dir, "base_sap.xlsx")
        path_base_frac = os.path.join(fonte_dir, "FRACAO_1.xlsx")
        path_base_estoque = os.path.join(fonte_dir, "estoqueseparacao.xlsx")
        path_base_peso_exter = os.path.join(fonte_dir,"receb_extern_peso.xlsx")
        df_externo_peso=pd.read_excel(path_base_peso_exter, sheet_name="Sheet1")
        df_estoque_sep = pd.read_excel(path_base_estoque)
        df_frac=pd.read_excel(path_base_frac, sheet_name="FRACAO")
        df_sap = pd.read_excel(path_base_sap, sheet_name="Sheet1")
        df_expedicao = pd.read_excel(path_base_expedicao, sheet_name="dado_exp")
        df_sobrepeso_real = pd.read_excel(path_base_sobrepeso, sheet_name="SOBREPESO")
        df_sobrepeso_real['DataHora'] = pd.to_datetime(df_sobrepeso_real['DataHora'])
        df_base_fisica = pd.read_excel(caminho_base_fisica, sheet_name="BASE FISICA")
        df_fracao = pd.read_excel(path_base_frac, sheet_name="FRACAO")
        try:
            self.log_text.clear()
            self.log_display.configure(state="normal")
            self.log_display.delete("1.0", "end")
            self.log_display.configure(state="disabled")


            file_path = criar_copia_planilha(fonte_dir, "SIMULADOR_BALAN√áA_LIMPO_2.xlsx", self.add_log)
            self.log_callback_completo(f"Abrindo planilha Excel em: {file_path}")

            xl = pd.ExcelFile(file_path)
            self.log_callback_completo("Lendo abas do arquivos...")
            with pd.ExcelFile(file_path) as xl:
                df_sku = xl.parse("dado_sku")
            self.log_callback_completo("Abas carregadas com sucesso.")
            self.progress_bar.set(0.3)
            peso_vazio = float(self.peso_vazio.get())
            peso_balan√ßa = float(self.peso_balanca.get())
            qtd_paletes = int(self.qtd_paletes.get())
            remessa = int(self.remessa.get())
            df_remessa = df_expedicao[df_expedicao['REMESSA'] == remessa]

            self.log_callback_completo(f"Entradas: Remessa={remessa}, Peso Vazio={peso_vazio}, Paletes={qtd_paletes}")

            self.log_callback_completo("Iniciando c√°lculo do peso final...")
            resultado = calcular_peso_final(
                remessa, peso_vazio, qtd_paletes,
                df_expedicao, df_sku, df_sap, df_sobrepeso_real,df_base_fisica, df_frac,df_estoque_sep,df_externo_peso,
                self.log_callback_tecnico
            )
            self.progress_bar.set(0.5)
            if resultado:
                peso_base, sp_total, peso_com_sp, peso_final, media_sp, itens_detalhados = resultado
            
                dados = {
                    'remessa': remessa,
                    'qtd_skus': df_expedicao[df_expedicao['REMESSA'] == remessa]['ITEM'].nunique(),
                    'placa': self.placa.get(),
                    'turno': self.turno.get(),
                    'peso_vazio': peso_vazio,
                    'peso_base': peso_base,
                    'sp_total': sp_total,
                    'peso_com_sp': peso_com_sp,
                    'peso_total_final': peso_final,
                    'media_sp': media_sp,
                    'qtd_paletes': qtd_paletes
                }

                self.log_callback_completo("Chamando preenchimento do formul√°rio via COM...")
                itens_detalhados_integrados = integrar_itens_detalhados(
                    df_remessa, df_sap, df_sobrepeso_real, df_sku, df_base_fisica, self.add_log
                )

                preencher_formulario_com_openpyxl(
                    file_path, dados, itens_detalhados_integrados, self.add_log,
                    df_sku, df_remessa, df_fracao
                )
                self.progress_bar.set(0.7)
                self.log_callback_completo("Exportando PDF...")
                pdf_path = exportar_pdf_com_comtypes(file_path, "FORMULARIO", nome_remessa=remessa, log_callback=self.add_log)
                self.log_callback_completo(f"PDF exportado com sucesso: {pdf_path}")
                self.progress_bar.set(0.85)
                self.log_callback_completo("Gerando relat√≥rio de diverg√™ncia em PDF...")
                relatorio_path = gerar_relatorio_diferenca(
                    remessa_num=remessa,
                    peso_final_balan√ßa=peso_balan√ßa,
                    peso_veiculo_vazio=peso_vazio,
                    df_remessa=df_expedicao[df_expedicao['REMESSA'] == remessa],
                    df_sku=df_sku,
                    peso_estimado_total=peso_com_sp,
                    pasta_excel=fonte_dir,
                    log_callback=self.log_callback_completo
                )
                self.log_callback_completo(f"Relat√≥rio adicional salvo em: {relatorio_path}")

                try:
                    self.log_callback_completo("Iniciando impress√£o do relat√≥rio")
                    print_pdf(pdf_path, log_callback=self.log_callback_completo)
                    print_pdf(relatorio_path, log_callback=self.log_callback_completo)
                    self.log_callback_completo("Relat√≥rio impresso com sucessos")
                    self.progress_bar.set(0.95)
                except:
                    self.log_callback_completo("Impress√£o do Relat√≥rio com Erro")
                try:
                    enviar_email_com_log_e_pdf(
                        pdf_path,
                        remessa,
                        log_callback=self.log_callback_completo,
                        log_geral=self.log_tecnico
                    )
                    self.log_callback_completo("Envio com sucesso para o email a tratativa")
                    self.progress_bar.set(1)
                    self.add_log("‚úÖ Processamento conclu√≠do com sucesso!")
                except:
                    self.log_callback_completo("Erro ao enviar para o email a tratativa")
                try:
                    time.sleep(1)
                    os.remove(file_path)
                    self.log_callback_completo(f"C√≥pia tempor√°ria removida: {file_path}")
                except Exception as e:
                    self.log_callback_completo(f"Erro ao remover a c√≥pia tempor√°ria: {e}")
                
                messagebox.showinfo(
                    "Sucesso",
                    f"Formul√°rio exportado: {pdf_path}\n\nRelat√≥rio de diverg√™ncia salvo:\n{relatorio_path}"
                )

            else:
                self.log_callback_completo("Falha no c√°lculo. Verifique os dados inseridos.")
                messagebox.showwarning("Aviso", "C√°lculo n√£o p√¥de ser realizado.")


        except Exception as e:
            self.log_callback_completo(f"Erro: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao processar: {str(e)}")


if __name__ == "__main__":
    app = App()
    app.mainloop()
