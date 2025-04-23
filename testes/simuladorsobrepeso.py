import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from datetime import datetime
import customtkinter as ctk
from tkinter import messagebox, StringVar
import os
import win32com.client as win32
import comtypes.client
from pathlib import Path



download_dir = os.path.join(os.environ['USERPROFILE'], 'Downloads')
MODELO_FORMULARIO = os.path.join(download_dir, "SIMULADOR_BALANÇA_LIMPO.xlsx")

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")

def calcular_peso_final(remessa_num, peso_veiculo_vazio, qtd_paletes, df_exp, df_sku, df_sap, df_sobrepeso, log_callback):
    try:
        remessa_num = int(remessa_num)
    except ValueError:
        log_callback("Remessa inválida.")
        return None

    df_remessa = df_exp[df_exp['REMESSA'] == remessa_num]
    if df_remessa.empty:
        log_callback("Remessa não encontrada em data_exp.")
        return None

    skus = df_remessa['ITEM'].unique()
    qtd_caixas_total = df_remessa['QUANTIDADE'].sum()

    peso_base_total = 0
    sobrepesos_por_item = {}
    sp_total = 0

    for sku in skus:
        qtd_caixas = df_remessa[df_remessa['ITEM'] == sku]['QUANTIDADE'].sum()
        df_sku_filtrado = df_sku[(df_sku['COD_PRODUTO'] == sku) & (df_sku['DESC_UNID_MEDID'] == 'Caixa')]
        if df_sku_filtrado.empty:
            continue

        peso_por_caixa = df_sku_filtrado.iloc[0]['QTDE_PESO_LIQ']
        peso_base = qtd_caixas * peso_por_caixa
        peso_base_total += peso_base

        chaves_pallet = df_remessa[df_remessa['ITEM'] == sku]['CHAVE_PALETE'].unique()
        df_pallets = df_sap[df_sap['Chave Pallet'].isin(chaves_pallet)]

        total_overweight = 0
        count_sp = 0

        for idx, row in df_pallets.iterrows():
            try:
                log_callback(f"Processando pallet {idx+1}/{len(df_pallets)}...")

                lote = row['Lote']
                log_callback(f"Lote: {lote}")
                
                data_producao = row['Data de produção']
                log_callback(f"Data produção: {data_producao}")

                hora_inicio = f"{row['Hora de criação'].hour:02d}:00:00"
                hora_fim = f"{row['Hora de modificação'].hour:02d}:00:00"
                log_callback(f"Intervalo hora: {hora_inicio} - {hora_fim}")

                linha_produzida = "L" + lote[-3:]
                log_callback(f"Linha produzida: {linha_produzida}")

                df_sp_filtro = df_sobrepeso[
                    (df_sobrepeso['LINHA'] == linha_produzida) &
                    (df_sobrepeso['Data e hora'] >= f"{data_producao} {hora_inicio}") &
                    (df_sobrepeso['Data e hora'] <= f"{data_producao} {hora_fim}")
                ]

                log_callback(f"Linhas sobrepeso encontradas: {len(df_sp_filtro)}")

                if not df_sp_filtro.empty:
                    media_sp = df_sp_filtro['sobrepesohora'].mean() / 100
                    log_callback(f"Média SP: {media_sp:.4f}")
                    ajuste = peso_base * media_sp
                    log_callback(f"Ajuste aplicado ao peso base ({peso_base:.2f}): {ajuste:.2f}kg")
                    total_overweight += media_sp
                    count_sp += 1

            except Exception as e:
                log_callback(f"Erro ao processar pallet {idx+1}: {e}")


        if count_sp > 0:
            sp_medio = total_overweight / count_sp
            sobrepesos_por_item[sku] = sp_medio
            sp_total_ajuste = peso_base * sp_medio
            log_callback(f"Total SP médio do SKU {sku}: {sp_medio:.4f}, ajuste total: {sp_total_ajuste:.2f}kg")
            sp_total += sp_total_ajuste

    peso_com_sobrepeso = peso_base_total + sp_total
    log_callback(f"Peso com sobrepeso: {peso_com_sobrepeso:.2f} kg")
    peso_total_com_paletes = peso_com_sobrepeso + (qtd_paletes * 26)
    log_callback(f"Peso total com paletes ({qtd_paletes} x 26kg): {peso_total_com_paletes:.2f} kg")
    if sobrepesos_por_item:
        media_sp_geral = sum(sobrepesos_por_item.values()) / len(sobrepesos_por_item)
    else:
        media_sp_geral = 0.0
    log_callback(f"Média geral de sobrepeso (entre {len(sobrepesos_por_item)} itens): {media_sp_geral:.4f}")

    return peso_base_total, sp_total, peso_com_sobrepeso, peso_total_com_paletes, media_sp_geral, sobrepesos_por_item

def preencher_formulario_com_openpyxl(path_copia, dados, sobrepesos_por_item, log_callback):
    try:
        wb = load_workbook(path_copia)
        ws = wb["FORMULARIO"]

        log_callback("Preenchendo cabeçalhos principais com openpyxl...")
        ws["B4"] = dados['remessa']
        ws["B6"] = dados['qtd_skus']
        ws["B7"] = dados['placa']
        ws["B8"] = dados['turno']
        ws["B9"] = dados['peso_vazio']
        ws["B10"] = dados['peso_base']
        ws["B11"] = dados['sp_total']
        ws["B12"] = dados['peso_com_sp']
        ws["B13"] = dados['peso_total_final']
        ws["B14"] = dados['media_sp']
        ws["B16"] = dados['peso_total_final'] * 1.02
        ws["B17"] = dados['peso_total_final']
        ws["B18"] = dados['peso_total_final'] * 0.99
        ws["D4"] = dados['qtd_paletes']
        ws["D9"] = dados['qtd_paletes'] * 26

        linha = 12
        log_callback("Preenchendo SKUs e sobrepesos...")
        for sku, sp in sobrepesos_por_item.items():
            if pd.notna(sp):
                ws[f"C{linha}"] = sku
                ws[f"D{linha}"] = sp
                linha += 1

        wb.save(path_copia)
        log_callback("Formulário preenchido e salvo com sucesso.")

    except Exception as e:
        log_callback(f"Erro no preenchimento: {e}")
        raise

def exportar_pdf_com_comtypes(path_xlsx, aba_nome="FORMULARIO", log_callback=None):
    try:
        if log_callback:
            log_callback("Iniciando exportação via comtypes...")

        comtypes.CoInitialize()
        excel = comtypes.client.CreateObject("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(path_xlsx)

        if aba_nome not in [sheet.Name for sheet in wb.Sheets]:
            raise Exception(f"Aba '{aba_nome}' não encontrada.")

        ws = wb.Worksheets(aba_nome)

        # Cria diretório temporário se não existir
        pdf_dir = "C:\\temp"
        os.makedirs(pdf_dir, exist_ok=True)

        pdf_path = os.path.join(pdf_dir, f"{Path(path_xlsx).stem}_{aba_nome}.pdf")
        if log_callback:
            log_callback(f"Tentando exportar para: {pdf_path}")

        ws.ExportAsFixedFormat(Type=0, Filename=pdf_path)
        wb.Close(False)
        excel.Quit()
        comtypes.CoUninitialize()

        if log_callback:
            log_callback(f"PDF exportado com sucesso: {pdf_path}")

        return pdf_path

    except Exception as e:
        if log_callback:
            log_callback(f"Erro ao exportar PDF: {e}")
        raise


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Simulador de Sobrepeso")
        self.geometry("700x600")

        self.placa = StringVar()
        self.turno = StringVar()
        self.remessa = StringVar()
        self.qtd_paletes = StringVar()
        self.peso_vazio = StringVar()
        self.peso_balanca = StringVar()
        self.log_text = []

        ctk.CTkLabel(self, text="Placa:").pack()
        ctk.CTkEntry(self, textvariable=self.placa).pack()

        ctk.CTkLabel(self, text="Turno:").pack()
        ctk.CTkComboBox(self, values=["Manhã", "Tarde", "Noite"], variable=self.turno).pack()

        ctk.CTkLabel(self, text="Remessa:").pack()
        ctk.CTkEntry(self, textvariable=self.remessa).pack()

        ctk.CTkLabel(self, text="Quantidade de Paletes:").pack()
        ctk.CTkEntry(self, textvariable=self.qtd_paletes).pack()

        ctk.CTkLabel(self, text="Peso Veículo Vazio:").pack()
        ctk.CTkEntry(self, textvariable=self.peso_vazio).pack()

        ctk.CTkLabel(self, text="Peso Final Balança:").pack()
        ctk.CTkEntry(self, textvariable=self.peso_balanca).pack()

        ctk.CTkButton(self, text="Calcular", command=self.processar).pack(pady=10)

        ctk.CTkLabel(self, text="Histórico de Logs:").pack(pady=5)
        self.log_display = ctk.CTkLabel(self, text="", wraplength=600, justify="left")
        self.log_display.pack(pady=5)

    def add_log(self, msg):
        print(msg)
        timestamp = datetime.now().strftime("%H:%M:%S")
        entrada = f"[{timestamp}] {msg}"
        self.log_text.append(entrada)
        self.log_display.configure(text="\n".join(self.log_text[-10:]))

    def processar(self):
        try:
            self.log_text.clear()
            self.log_display.configure(text="")
            file_path = os.path.join(download_dir,"SIMULADOR_BALANÇA_LIMPO.xlsx")
            self.add_log(f"Abrindo planilha Excel em: {file_path}")

            xl = pd.ExcelFile(file_path)
            self.add_log("Lendo abas do arquivo...")
            df_exp = xl.parse("dado_exp")
            df_sap = xl.parse("dado_sap")
            df_sku = xl.parse("dado_sku")
            df_sobrepeso = xl.parse("dado_sobrepeso")
            self.add_log("Abas carregadas com sucesso.")

            df_sobrepeso = df_sobrepeso[~df_sobrepeso['Data e hora'].astype(str).str.contains("Redimensionar", na=False)]
            df_sobrepeso['Data e hora'] = pd.to_datetime(df_sobrepeso['Data e hora'], errors='coerce')
            df_sobrepeso = df_sobrepeso.dropna(subset=['Data e hora'])
            self.add_log("Pré-processamento do sobrepeso finalizado.")

            peso_vazio = float(self.peso_vazio.get())
            qtd_paletes = int(self.qtd_paletes.get())
            remessa = int(self.remessa.get())
            self.add_log(f"Entradas: Remessa={remessa}, Peso Vazio={peso_vazio}, Paletes={qtd_paletes}")

            self.add_log("Iniciando cálculo do peso final...")
            resultado = calcular_peso_final(
                remessa, peso_vazio, qtd_paletes,
                df_exp, df_sku, df_sap, df_sobrepeso,
                self.add_log
            )

            if resultado:
                peso_base, sp_total, peso_com_sp, peso_final, media_sp, sp_itens = resultado

                dados = {
                    'remessa': remessa,
                    'qtd_skus': df_exp[df_exp['REMESSA'] == remessa]['ITEM'].nunique(),
                    'placa': self.placa.get(),
                    'turno': self.turno.get(),
                    'peso_vazio': peso_vazio,
                    'peso_base': peso_base,
                    'sp_total': sp_total,
                    'peso_com_sp': peso_com_sp,
                    'peso_total_final': peso_final,
                    'media_sp': media_sp,
                    'qtd_paletes': qtd_paletes
                }

                self.add_log("Chamando preenchimento do formulário via COM...")
                preencher_formulario_com_openpyxl(file_path, dados, sp_itens, self.add_log)

                self.add_log("Exportando PDF...")
                pdf_path = exportar_pdf_com_comtypes(file_path, "FORMULARIO")
                self.add_log(f"PDF exportado com sucesso: {pdf_path}")

                messagebox.showinfo("Sucesso", f"Resultado salvo e exportado para PDF:\n{pdf_path}")
            else:
                self.add_log("Falha no cálculo. Verifique os dados inseridos.")
                messagebox.showwarning("Aviso", "Cálculo não pôde ser realizado.")

        except Exception as e:
            self.add_log(f"Erro: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao processar: {str(e)}")

if __name__ == "__main__":
    app = App()
    app.mainloop()
