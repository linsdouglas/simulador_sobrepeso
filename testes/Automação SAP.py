#Automação SAP
import time
import math
import  datetime
import threading
import tkinter as tk
from tkinter import messagebox, scrolledtext
from tkinter import ttk
import customtkinter as ctk
from PIL import Image
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import ElementClickInterceptedException, TimeoutException
import pyautogui as pt
import yagmail
from openpyxl import load_workbook
import comtypes.client
import glob
import shutil
import subprocess
import pandas as pd
import os
import sys
import requests
from openpyxl.styles import numbers

base_sap="base_sap"
download_dir = os.path.join(os.environ['USERPROFILE'], 'Downloads')
if not os.path.exists(download_dir):
    os.makedirs(download_dir)

def encontrar_pasta_onedrive_empresa():
    user_dir = os.environ["USERPROFILE"]
    possiveis = os.listdir(user_dir)
    for nome in possiveis:
        if "DIAS BRANCO" in nome.upper():
            caminho_completo = os.path.join(user_dir, nome)
            if os.path.isdir(caminho_completo) and "Gestão de Estoque - Documentos" in os.listdir(caminho_completo):
                return os.path.join(caminho_completo, "Gestão de Estoque - Documentos")
    return None

fonte_dir = encontrar_pasta_onedrive_empresa()
if not fonte_dir:
    raise FileNotFoundError("Não foi possível localizar a pasta sincronizada do SharePoint via OneDrive.")

url = "https://s4.mdiasbranco.com.br:44380/sap/bc/gui/sap/its/webgui#"
chrome_options = webdriver.ChromeOptions()
#chrome_options.debugger_address = "localhost:9222"
driver = webdriver.Chrome(options=chrome_options)
#chrome_options.add_argument("--headless=new")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--window-size=1920x1080")
actions = ActionChains(driver)
prefs = {
    "download.default_directory": download_dir,
    "download.prompt_for_download": False,
    "directory_upgrade": True
}
chrome_options.add_experimental_option("prefs", prefs)
def login_sap():
    try:
        driver.get(url)
        time.sleep(2)
        username_field = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, "//input[@type='email' and @name='loginfmt']"))
        )
        print("Campo de usuário encontrado, efetuando printin...")
        username_field.clear()
        username_field.send_keys("xql80316@mdb.com.br")
        password_field = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, "//input[@type='password' and @name='passwd']"))
        )
        password_field.clear()
        password_field.send_keys("8583Doug#")
        password_field.send_keys(Keys.ENTER)
        print("printin efetuado com sucesso.")
        time.sleep(5)
    except Exception as e:
        print("Login não requerido após o refresh ou elemento não encontrado")

def interacoes_sap(driver,actions):
    driver.set_window_size(1920, 2000)
    try:
        WebDriverWait(driver, 5).until(lambda d: d.execute_script("return document.readyState") == "complete")
        time.sleep(2)
        elemento_pesq = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, 'ToolbarOkCode')))
        safe_click(driver,(By.ID, 'ToolbarOkCode'),nome_elemento="barra de pesquisa principal do sap")
        time.sleep(1)
        elemento_pesq.clear()
        elemento_pesq.send_keys("/nZPM003")
        time.sleep(1)
        actions.send_keys(Keys.ENTER).perform()
        elemento_centro_trabalho = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'M0:46:::4:34')))
        safe_click(driver,(By.ID, 'M0:46:::4:34'),nome_elemento="centro de trabalho")
        time.sleep(1)
        elemento_centro_trabalho.send_keys("M431")
        hoje = datetime.date.today()
        data_inicial = hoje - datetime.timedelta(days=60)
        data_inicial_input = driver.find_element(By.ID, "M0:46:::2:34")
        data_inicial_input.send_keys(data_inicial.strftime("%d.%m.%Y"))
        data_final_input = driver.find_element(By.ID, "M0:46:::2:59")
        data_final_input.send_keys(hoje.strftime("%d.%m.%Y"))
        executar_button = driver.find_element(By.ID, "M0:50::btn[8]")
        safe_click(driver,(By.ID,"M0:50::btn[8]"),nome_elemento="botão executar pesquisa")
        time.sleep(3)
        WebDriverWait(driver, 5).until(lambda d: d.execute_script("return document.readyState") == "complete")
        elemento_pesq = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'grid#C114#4,4#if-r')))
        actions.key_down(Keys.SHIFT).send_keys(Keys.F9).key_up(Keys.SHIFT).perform()       
        elemento_relatorio = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'M1:46:::2:3')))
        safe_click(driver,(By.ID, 'M1:50::btn[0]'),nome_elemento="exportar o relatório")
        WebDriverWait(driver, 5).until(lambda d: d.execute_script("return document.readyState") == "complete")
        try:
            WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, 'webguiPopupWindow10000-header-title-txt')))
            print("elemento_mostra_relatório encontrado")
        except Exception as e:
            print("elemento_mostra_relatório não encontrado")
        try:
            safe_click(driver,(By.ID, 'UpDownDialogChoose'),nome_elemento="confirmar exportação")
            print("confirmado a exportação com sucesso")
        except Exception as e:
            print("erro ao clicar no ok de exportação", e)
        wait = WebDriverWait(driver, 70)
        wait.until(lambda driver: 
            "Download" in driver.find_element(By.CSS_SELECTOR, "[id='wnd[0]/sbar_msg-txt']").text and 
            "EXPORT.XLSX" in driver.find_element(By.CSS_SELECTOR, "[id='wnd[0]/sbar_msg-txt']").text
        )
        msg_elemento = wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "[id='wnd[0]/sbar_msg-txt']"))
        )
        mensagem = driver.find_element(By.CSS_SELECTOR, "[id='wnd[0]/sbar_msg-txt']").text
        print("Mensagem recebida:", mensagem)
        print("Download confirmado. Seguindo para o próximo passo.")
        time.sleep(4)
        try:
            safe_click(driver,(By.XPATH,
        "//div[@role='button' and @title='Ações e configurações GUI']"),nome_elemento="configuração pra abrir o relatório de downloads")
        except Exception as e:
            print("não encontrado elemento de configuração")
        wait = WebDriverWait(driver, 10)
        browser_item = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//tr[@role='menuitem' and contains(@aria-label, 'Browser de arquivo SAP GUI for HTML')]")
        ))
        safe_click(driver,(By.XPATH, "//tr[@role='menuitem' and contains(@aria-label, 'Browser de arquivo SAP GUI for HTML')]"),nome_elemento="relatórios de download")
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'popupDialogExportBtn')))
        time.sleep(5)
        input_export = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//input[@value='EXPORT.XLSX' and @readonly='true']"))
        )
        safe_click(driver,(By.XPATH, "//input[@value='EXPORT.XLSX' and @readonly='true']"),nome_elemento="linha onde o relatório está")
        time.sleep(2)
        safe_click(driver,(By.ID,'popupDialogExportBtn'),nome_elemento="botão final para exportar")
    except Exception as e:
        print("erro ao interagir")


def safe_click(driver, by_locator,nome_elemento="Elemento", timeout=10):
    try:
        element = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable(by_locator)
        )
        element.click()
        print(f"Clique padrão realizado com sucesso no elemento:{nome_elemento}")
    except (ElementClickInterceptedException, TimeoutException) as e:
        print(f"Clique padrão falhou: {repr(e)}. Tentando com JavaScript...")
        try:
            element = WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located(by_locator)
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", element)
            time.sleep(0.5)
            driver.execute_script("arguments[0].click();", element)
            print("Clique forçado via JavaScript realizado com sucesso.")
        except Exception as js_e:
            print(f"Erro ao clicar com JavaScript: {repr(js_e)}")
    except Exception as e:
        print(f"Erro inesperado ao tentar clicar: {repr(e)}")

def encontrar_arquivo_export():
    arquivos = glob.glob(os.path.join(download_dir, "EXPORT*.XLSX"))
    print("Arquivos encontrados:", arquivos)
    if not arquivos:
        return None
    arquivos.sort(key=os.path.getmtime, reverse=True)
    return arquivos[0]

def aguardar_download_final(nome_base="EXPORT", timeout=30):
    tempo_inicial = time.time()
    while time.time() - tempo_inicial < timeout:
        arquivos = glob.glob(os.path.join(download_dir, f"{nome_base}*.XLSX"))
        if arquivos and all(not a.endswith(".crdownload") for a in arquivos):
            return True
        time.sleep(1)
    return False

def envio_base_sap():
    try:
        caminho_export = encontrar_arquivo_export()
        if not caminho_export or not os.path.exists(caminho_export):
            print("Arquivo EXPORT.XLSX não encontrado ou ainda não foi salvo completamente.")
            return

        caminho_base = os.path.join(fonte_dir, "base_sap.xlsx")
        if not os.path.exists(caminho_base):
            print("Arquivo base_sap.xlsx não encontrado no OneDrive.")
            return

        df_novos = pd.read_excel(caminho_export)
        df_novos = df_novos.iloc[1:].reset_index(drop=True)
        if "CHAVE_PALLET" not in df_novos.columns:
            print("Coluna 'CHAVE_PALLET' não encontrada no arquivo exportado.")
            return

        wb = load_workbook(caminho_base)
        ws = wb["dado_sap"]

        colunas_base = [cell.value for cell in ws[1]]
        if "CHAVE_PALLET" not in colunas_base or "DATA_ENTRADA" not in colunas_base:
            print("Coluna 'CHAVE_PALLET' ou 'DATA_ENTRADA' não encontrada na planilha base.")
            return

        idx_chave = colunas_base.index("CHAVE_PALLET") + 1
        idx_data = colunas_base.index("DATA_ENTRADA") + 1

        data_limite = (datetime.today() - datetime.timedelta(days=60))
        linhas_para_apagar = []

        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            valor_data = row[idx_data - 1].value
            if isinstance(valor_data, datetime):
                valor_data = valor_data.date()
            if valor_data == data_limite:
                linhas_para_apagar.append(i)

        for offset, linha in enumerate(linhas_para_apagar):
            ws.delete_rows(linha - offset)

        print(f"{len(linhas_para_apagar)} linhas com DATA_ENTRADA = {data_limite.strftime('%d/%m/%Y')} foram removidas.")
        chaves_existentes = set()
        for row in ws.iter_rows(min_row=2, max_col=idx_chave):
            valor = row[idx_chave - 1].value
            if valor:
                chaves_existentes.add(str(valor))

        df_filtrado = df_novos[~df_novos["CHAVE_PALLET"].astype(str).isin(chaves_existentes)]
        if df_filtrado.empty:
            print("Nenhum novo registro a ser inserido.")
            wb.save(caminho_base)
            wb.close()
            return

        linha_inicio = ws.max_row + 1
        colunas_data = ["DATA_ENTRADA", "DATA_VENCIMENTO", "DATA_PRODUCAO", "DATA_CRIACAO", "DATA_MODIFICACAO"]

        for i, row in df_filtrado.iterrows():
            for j, value in enumerate(row):
                nome_coluna = df_filtrado.columns[j]

                if nome_coluna.upper() in colunas_data:
                    if isinstance(value, pd.Timestamp):
                        value = value.date()
                    elif isinstance(value, str):
                        try:
                            value = datetime.strptime(value, "%d/%m/%Y").date()
                        except:
                            pass

                celula = ws.cell(row=linha_inicio + i, column=j + 1, value=value)
                if isinstance(value, datetime.date):
                    celula.number_format = "DD/MM/YYYY"

        wb.save(caminho_base)
        wb.close()
        print(f"{len(df_filtrado)} novos registros adicionados com base na coluna CHAVE_PALLET.")

    except Exception as e:
        print(f"Erro ao colar os dados na base SAP: {e}")

    try:
        if caminho_export and os.path.exists(caminho_export):
            os.remove(caminho_export)
            print(f"Relatório {os.path.basename(caminho_export)} apagado com sucesso.")
    except Exception as e:
        print(f"Erro ao apagar o arquivo do relatório SAP: {e}")
        
if __name__ == "__main__":
    login_sap()
    actions = ActionChains(driver)
    interacoes_sap(driver, actions)
    aguardar_download_final
    envio_base_sap()
    driver.quit()

