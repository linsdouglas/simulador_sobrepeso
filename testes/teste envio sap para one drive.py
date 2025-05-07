import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta, date

import os


fonte_export = os.path.join(os.environ["USERPROFILE"], "Downloads")
caminho_export = os.path.join(fonte_export, "EXPORT.xlsx")

def encontrar_pasta_onedrive_empresa():
    user_dir = os.environ["USERPROFILE"]
    possiveis = os.listdir(user_dir)
    for nome in possiveis:
        if "DIAS BRANCO" in nome.upper():
            caminho_completo = os.path.join(user_dir, nome)
            if os.path.isdir(caminho_completo) and "Gestão de Estoque - Documentos" in os.listdir(caminho_completo):
                return os.path.join(caminho_completo, "Gestão de Estoque - Documentos")
    return None

fonte_dir = encontrar_pasta_onedrive_empresa()
caminho_base = os.path.join(fonte_dir, "base_sap.xlsx")
if not fonte_dir:
    raise FileNotFoundError("Não foi possível localizar a pasta sincronizada do SharePoint via OneDrive.")

def envio_base_sap_teste():
    print("[INÍCIO] Script de envio base SAP iniciado.")
    try:
        if not os.path.exists(caminho_export):
            print("Arquivo EXPORT_TESTE.xlsx não encontrado.")
            return
        else:
            print("Arquivo exportado encontrado.")

        if not os.path.exists(caminho_base):
            print("Arquivo base_sap.xlsx não encontrado.")
            return
        else:
            print("Diretório no OneDrive da base SAP encontrado.")

        df_novos = pd.read_excel(caminho_export)
        print("Leitura do Excel feita com sucesso.")
        df_novos = df_novos.iloc[1:].reset_index(drop=True)

        if "Chave Pallet" not in df_novos.columns:
            print("Coluna 'Chave Pallet' não encontrada no arquivo exportado.")
            return
        else:
            print("Coluna 'Chave Pallet' OK.")

        wb = load_workbook(caminho_base)
        ws = wb["dado_sap"]

        colunas_base = [cell.value for cell in ws[1]]
        if "Chave Pallet" not in colunas_base or "Data de entrada" not in colunas_base:
            print("Coluna 'Chave Pallet' ou 'Data de entrada' não encontrada na planilha base.")
            return
        else:
            print("Colunas necessárias encontradas.")

        idx_chave = colunas_base.index("Chave Pallet") + 1
        idx_data = colunas_base.index("Data de entrada") + 1
        print("Identificadores de coluna definidos.")

        data_limite = (datetime.today() - timedelta(days=60)).date()
        linha_primeira_valida = None

        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            valor_data = row[idx_data - 1].value
            if isinstance(valor_data, datetime):
                valor_data = valor_data.date()

            if valor_data and valor_data >= data_limite:
                linha_primeira_valida = i
                break

        if linha_primeira_valida and linha_primeira_valida > 2:
            total_apagar = linha_primeira_valida - 2
            ws.delete_rows(2, total_apagar)
            print(f"{total_apagar} linhas com 'Data de entrada' anterior a {data_limite.strftime('%d/%m/%Y')} foram removidas.")
        else:
            print("Nenhuma linha antiga encontrada para remover.")

        chaves_existentes = set()
        for row in ws.iter_rows(min_row=2, max_col=idx_chave):
            valor = row[idx_chave - 1].value
            if valor:
                chaves_existentes.add(str(valor))

        df_filtrado = df_novos[~df_novos["Chave Pallet"].astype(str).isin(chaves_existentes)]
        if df_filtrado.empty:
            print("Nenhum novo registro a ser inserido.")
            wb.save(caminho_base)
            wb.close()
            return

        linha_inicio = ws.max_row + 1
        colunas_data = ["Data de entrada", "Data do vencimento", "Data de produção", "Data de criação", "Data de modificação"]

        for i, row in df_filtrado.iterrows():
            for j, value in enumerate(row):
                nome_coluna = df_filtrado.columns[j]

                if nome_coluna in colunas_data:
                    if isinstance(value, pd.Timestamp):
                        value = value.date()
                    elif isinstance(value, str):
                        try:
                            value = datetime.strptime(value, "%d/%m/%Y").date()
                        except:
                            pass

                celula = ws.cell(row=linha_inicio + i, column=j + 1, value=value)
                if isinstance(value, date):
                    celula.number_format = "DD/MM/YYYY"

        wb.save(caminho_base)
        wb.close()
        print(f"{len(df_filtrado)} novos registros adicionados com base na coluna 'Chave Pallet'.")

    except Exception as e:
        print(f"Erro no processo de teste: {e}")


if __name__ == "__main__":
    envio_base_sap_teste()
